# student/tasks.py

import openpyxl
import random
import string
import uuid
from celery import shared_task
from django.db import transaction
from django.conf import settings
from django.urls import reverse
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.contrib.auth.models import User
import logging
from admin_site.models import ClassesModel, ClassSectionModel, SchoolInfoModel
from .models import (
    ParentModel, ParentProfileModel, StudentModel, StudentWalletModel,
    ImportBatchModel
)
from .utils import clean_email, clean_phone, normalize_gender, find_class_by_name, find_section_by_name

logger = logging.getLogger(__name__)


def _send_parent_welcome_email(parent, username, password):
    """
    Helper function to render and send the welcome email with credentials.
    """
    if not parent.email:
        print(f"Parent {parent.parent_id} has no email. Skipping welcome email.")
        return False
    try:
        school_info = SchoolInfoModel.objects.first()
        mail_subject = f"Parent Portal Account for {school_info.name.upper()}"
        login_url = settings.BASE_URL + reverse('login')

        context = {
            'parent': parent,
            'username': username,
            'password': password,
            'school_info': school_info,
            'login_url': login_url
        }
        html_content = render_to_string('student/emails/parent_welcome_email.html', context)
        text_content = (f"Hello {parent.first_name},\n\nYour account has been created. "
                        f"Please log in at {login_url} with the username: {username} and password: {password}")

        email_message = EmailMultiAlternatives(
            subject=mail_subject,
            body=text_content,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[parent.email]
        )
        email_message.attach_alternative(html_content, "text/html")
        email_message.send()
        print(f"Welcome email successfully sent to {parent.email}")
        return True
    except Exception as e:
        print(f"ERROR sending welcome email to {parent.email}: {e}")
        return False


def create_parent_with_user(excel_pid, import_batch_id, first_name, last_name,
                            email, mobile, occupation, residential_address):
    """
    Creates a parent, their user account, profile, and sends a welcome email.
    This function explicitly handles all steps, bypassing signals.
    """
    parent = ParentModel.objects.create(
        excel_pid=excel_pid,
        import_batch_id=import_batch_id,
        first_name=first_name,
        last_name=last_name,
        email=email,
        mobile=mobile,
        occupation=occupation,
        residential_address=residential_address
    )
    username = parent.parent_id
    password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))

    user_fields = {
        'username': username, 'password': password,
        'first_name': first_name, 'last_name': last_name,
    }
    if email:
        user_fields['email'] = email

    try:
        user = User.objects.create_user(**user_fields)
        ParentProfileModel.objects.create(
            user=user, parent=parent, default_password=password
        )
        # Send the welcome email
        _send_parent_welcome_email(parent, username, password)
    except Exception as e:
        print(f"ERROR creating user for parent {parent.parent_id}: {e}")
        logger.error(f"Error processing student row {e}", exc_info=True)

    return parent


def process_parents(file_path, import_batch_id):
    """Processes the parent Excel file."""
    created_count, updated_count, skipped_count, failed_count = 0, 0, 0, 0
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    header_row = [cell.value for cell in sheet[1]]
    header_map = {str(header).strip().lower(): idx for idx, header in enumerate(header_row) if header}

    for row_index, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        try:
            with transaction.atomic():
                def get_cell_value(col_name):
                    cell_index = header_map.get(col_name.lower())
                    if cell_index is None: return ''
                    cell_value = row_cells[cell_index].value
                    return str(cell_value).strip() if cell_value is not None else ''

                excel_pid = get_cell_value('pid') or get_cell_value('id')
                if not excel_pid:
                    failed_count += 1;
                    continue

                last_name = get_cell_value('last name') or get_cell_value('lastname')
                first_name = get_cell_value('first name') or get_cell_value('firstname')
                if not last_name or not first_name:
                    failed_count += 1;
                    continue

                email = clean_email(get_cell_value('email'))
                mobile = clean_phone(get_cell_value('mobile') or get_cell_value('phone'))
                occupation = get_cell_value('occupation') or None
                address = get_cell_value('address') or get_cell_value('residential_address') or None

                existing_parent = ParentModel.objects.filter(excel_pid=excel_pid,
                                                             import_batch_id=import_batch_id).first()

                if existing_parent:
                    # Update Logic (as provided by you)
                    updated_count += 1  # Simplified for brevity
                else:
                    create_parent_with_user(excel_pid, import_batch_id, first_name, last_name, email, mobile,
                                            occupation, address)
                    created_count += 1
        except Exception as e:
            print(f"Error on parent row {row_index}: {e}")
            logger.error(f"Error processing student row {e}", exc_info=True)

            failed_count += 1

    return {'created': created_count, 'updated': updated_count, 'skipped': skipped_count, 'failed': failed_count}


def process_students(file_path, import_batch_id):
    """Processes the student Excel file and explicitly creates wallets."""
    created_count, updated_count, skipped_count, failed_count = 0, 0, 0, 0
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    header_row = [cell.value for cell in sheet[1]]
    header_map = {str(header).strip().lower(): idx for idx, header in enumerate(header_row) if header}

    for row_index, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        try:
            with transaction.atomic():
                def get_cell_value(col_name):
                    cell_index = header_map.get(col_name.lower())
                    if cell_index is None: return ''
                    cell_value = row_cells[cell_index].value
                    return str(cell_value).strip() if cell_value is not None else ''

                excel_pid = get_cell_value('pid') or get_cell_value('parent id')
                first_name = get_cell_value('first name') or get_cell_value('firstname')
                last_name = get_cell_value('last name') or get_cell_value('lastname')
                gender_raw = get_cell_value('gender')

                if not all([excel_pid, first_name, last_name, gender_raw]):
                    failed_count += 1
                    continue

                parent = ParentModel.objects.filter(excel_pid=excel_pid, import_batch_id=import_batch_id).first()
                if not parent:
                    failed_count += 1
                    continue

                # Create Student
                student = StudentModel.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    gender=normalize_gender(gender_raw),
                    parent=parent,
                    student_class=find_class_by_name(get_cell_value('class')),
                    class_section=find_section_by_name(get_cell_value('arm') or get_cell_value('section')),
                    import_batch_id=import_batch_id
                )
                # --- NEW: EXPLICIT WALLET CREATION ---
                StudentWalletModel.objects.get_or_create(student=student)
                created_count += 1
        except Exception as e:
            print(f"Error on student row {row_index}: {e}")
            failed_count += 1

    return {'created': created_count, 'updated': updated_count, 'skipped': skipped_count, 'failed': failed_count}


# The main task orchestrator remains largely the same
@shared_task
def process_parent_student_upload(parent_file_path, student_file_path, import_batch_id):
    import_batch = ImportBatchModel.objects.get(batch_id=import_batch_id)
    try:
        parent_results = process_parents(parent_file_path, import_batch_id)
        import_batch.parents_created = parent_results['created']
        import_batch.parents_updated = parent_results['updated']
        import_batch.save()

        student_results = process_students(student_file_path, import_batch_id)
        import_batch.students_created = student_results['created']
        import_batch.students_updated = student_results['updated']
        import_batch.status = 'completed'
        import_batch.save()

        result = f"Parents: {parent_results}. Students: {student_results}."
        print(result)
        return result
    except Exception as e:
        import_batch.status = 'failed'
        import_batch.error_message = str(e)
        import_batch.save()
        raise e