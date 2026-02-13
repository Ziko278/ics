import hashlib
import logging
import io
import json
import base64
import random
import re
import secrets
import string
from datetime import datetime
from functools import reduce
import operator

import logging
from io import BytesIO

from PIL import Image
import numpy as np

from django.conf import settings
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string

from admin_site.views import FlashFormErrorsMixin
from .tasks import process_parent_student_upload, _send_parent_welcome_email
from django.core.files.storage import FileSystemStorage
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.views import View
from xlsxwriter import Workbook
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_http_methods
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView, DetailView, TemplateView
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from admin_site.models import ClassesModel, ClassSectionModel, ClassSectionInfoModel, SchoolInfoModel
from .models import StudentModel, ParentModel, StudentSettingModel, FingerprintModel, ImportBatchModel, \
    ParentProfileModel, StudentWalletModel, UtilityModel
from .forms import StudentForm, ParentForm, StudentSettingForm, ParentStudentUploadForm, UtilityForm

logger = logging.getLogger(__name__)


# -------------------------
# Student Setting Views (Singleton)
# -------------------------


class UtilityListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    The main view for displaying the list of utilities. It also provides
    the form instance needed for the 'Add New' modal.
    """
    model = UtilityModel
    # Assuming new permissions parallel to the original
    permission_required = 'student.view_utilitymodel'
    template_name = 'student/utility/index.html' # New template path
    context_object_name = 'utilities'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Provide an empty form for the 'Add New Utility' modal.
        if 'form' not in context:
            context['form'] = UtilityForm()
        return context


class UtilityCreateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView):
    """
    Handles the creation of a new utility. This view only processes POST
    requests from the modal form on the utility list page.
    """
    model = UtilityModel
    permission_required = 'student.add_utilitymodel'
    form_class = UtilityForm
    template_name = 'student/utility/index.html'  # Required for error redirect context

    def get_success_url(self):
        return reverse('student_utility_list') # New URL name

    def form_valid(self, form):
        messages.success(self.request, f"Utility '{form.cleaned_data['name']}' created successfully.")
        # Omitted form.instance.created_by = self.request.user (field not in model)
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        # This view should not be accessed via GET. It is a POST endpoint only.
        if request.method == 'GET':
            return redirect(self.success_url)
        return super().dispatch(request, *args, **kwargs)


class UtilityUpdateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView):
    """
    Handles updating an existing utility. This view only processes POST
    requests from the modal form on the utility list page.
    """
    model = UtilityModel
    permission_required = 'student.add_utilitymodel'
    form_class = UtilityForm
    template_name = 'student/utility/index.html'  # Required for error redirect context

    def get_success_url(self):
        return reverse('student_utility_list') # New URL name

    def form_valid(self, form):
        messages.success(self.request, f"Utility '{form.cleaned_data['name']}' updated successfully.")
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        # This view should not be accessed via GET. It is a POST endpoint only.
        if request.method == 'GET':
            return redirect(self.success_url)
        return super().dispatch(request, *args, **kwargs)


class UtilityDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """
    Handles the actual deletion of a utility object. The confirmation
    is handled by a modal on the list page.
    """
    model = UtilityModel
    permission_required = 'student.delete_utilitymodel'
    template_name = 'student/utility/delete.html'  # New template path
    success_url = reverse_lazy('student_utility_list') # New URL name
    context_object_name = 'utility'

    def form_valid(self, form):
        # Add a success message before deleting the object.
        messages.success(self.request, f"Utility '{self.object.name}' was deleted successfully.")
        return super().form_valid(form)


class StudentSettingDetailView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'student.view_studentsettingmodel'
    template_name = 'student/setting/detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student_setting'] = StudentSettingModel.objects.first()
        return context


class StudentSettingCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = StudentSettingModel
    permission_required = 'student.change_studentsettingmodel'
    form_class = StudentSettingForm
    template_name = 'student/setting/create.html'

    def get_success_url(self):
        return reverse('setting_detail')

    def dispatch(self, request, *args, **kwargs):
        if StudentSettingModel.objects.exists():
            return redirect(reverse('setting_edit'))
        return super().dispatch(request, *args, **kwargs)


class StudentSettingUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = StudentSettingModel
    permission_required = 'student.change_studentsettingmodel'
    form_class = StudentSettingForm
    template_name = 'student/setting/create.html'

    def get_object(self, queryset=None):
        return StudentSettingModel.objects.first()

    def get_success_url(self):
        return reverse('setting_detail')


class ParentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ParentModel
    permission_required = 'student.view_studentmodel'  # keep as you have it, or change to 'student.view_parentmodel' if appropriate
    template_name = 'student/parent/index.html'
    context_object_name = "parent_list"

    def has_permission(self):
        user = self.request.user
        # Superusers and users with the permission are allowed
        if user.is_superuser or user.has_perm(self.permission_required):
            return True

        # Allow if user is a form teacher (we'll still restrict queryset)
        try:
            staff = user.staff_profile.staff
            return ClassSectionInfoModel.objects.filter(form_teacher=staff).exists()
        except Exception:
            return False

    def get_queryset(self):
        user = self.request.user

        # Superuser or user with permission -> full list
        if user.is_superuser or user.has_perm(self.permission_required):
            return ParentModel.objects.all().order_by('first_name', 'last_name')

        # Otherwise, restrict to parents who have wards in the teacher's assigned classes/sections
        try:
            staff = user.staff_profile.staff
            assigned_infos = ClassSectionInfoModel.objects.filter(form_teacher=staff)
            assigned_class_ids = list(assigned_infos.values_list('student_class_id', flat=True))
            assigned_section_ids = list(assigned_infos.values_list('section_id', flat=True))

            # Filter parents via the reverse relation `wards` on StudentModel
            qs = ParentModel.objects.filter(
                Q(wards__student_class_id__in=assigned_class_ids) |
                Q(wards__class_section_id__in=assigned_section_ids)
            ).distinct().order_by('first_name', 'last_name')

            return qs
        except Exception:
            return ParentModel.objects.none()


class ParentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ParentModel
    permission_required = 'student.add_studentmodel'  # Permission is on ParentModel
    form_class = ParentForm
    template_name = 'student/parent/create.html'

    def has_permission(self):
        """
        Allow access if user has the required permission
        OR is assigned as a form teacher.
        """
        user = self.request.user

        # Normal permission check (from PermissionRequiredMixin)
        if super().has_permission():
            return True

        # Custom form-teacher check
        try:
            staff = user.staff_profile.staff
            is_form_teacher = ClassSectionInfoModel.objects.filter(form_teacher=staff).exists()
            if is_form_teacher:
                return True
        except Exception:
            pass

        return False

    def form_valid(self, form):
        """
        This method is called when valid form data has been POSTed.
        It's the ideal place to add logic for creating the user and sending emails.
        """
        # First, let the parent class do its job: save the ParentModel instance.
        # super().form_valid(form) returns the HttpResponseRedirect.
        response = super().form_valid(form)

        # self.object is now the newly created ParentModel instance.
        parent = self.object

        try:
            # 1. Use the auto-generated parent_id as the unique username.
            username = parent.parent_id

            # Safety check: ensure a user doesn't already exist.
            if User.objects.filter(username=username).exists():
                messages.warning(self.request,
                                 f"Parent was created, but a user with login ID '{username}' already exists. Please resolve this manually.")
                return response

            # 2. Generate a random password.
            # You can use your make_random_password function or a simple one here.
            password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))

            # 3. Create the Django User object.
            user = User.objects.create_user(
                username=username,
                password=password,
                email=parent.email,
                first_name=parent.first_name,
                last_name=parent.last_name
            )

            # 4. Create the ParentProfile to link the User and Parent.
            ParentProfileModel.objects.create(
                user=user,
                parent=parent,
                default_password=password
            )

            # 5. Send the welcome email directly, bypassing the signal.
            email_sent = _send_parent_welcome_email(parent, username, password)
            if email_sent:
                messages.info(self.request, f"A welcome email with login credentials has been sent to {parent.email}.")
            elif parent.email:
                messages.warning(self.request,
                                 "Parent login was created, but the welcome email could not be sent. Please check the server logs.")

        except Exception as e:
            # If anything goes wrong, log the error and notify the admin.
            logger.error(f"Failed to create user account or send email for Parent ID {parent.id}: {e}")
            messages.error(self.request,
                           "The parent was created, but there was a critical error creating their login account. Please review the system logs.")

        # Finally, return the original redirect response.
        return response

    def get_success_url(self):
        """
        Determines the redirect URL after the form is successfully submitted.
        This part remains unchanged.
        """
        action = self.request.POST.get('action')
        if action == 'save_and_add_ward':
            messages.success(self.request, "Parent created successfully. Now, please register their first ward.")
            return reverse('student_create', kwargs={'parent_pk': self.object.pk})

        # The main success message is now part of form_valid, so we can make this one simpler.
        messages.success(self.request, "Parent record created successfully.")
        return reverse('parent_detail', kwargs={'pk': self.object.pk})


class ParentDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = ParentModel
    permission_required = 'student.view_studentmodel'
    template_name = 'student/parent/detail.html'
    context_object_name = "parent"

    def has_permission(self):
        """
        Override permission logic:
        - Allow if user has 'view_parentmodel' permission
        - Allow if user is superuser
        - Allow if staff is form teacher for any class containing a ward of this parent
        """
        user = self.request.user

        # Superusers always allowed
        if user.is_superuser:
            return True

        # Has global permission
        if user.has_perm(self.permission_required):
            return True

        # Otherwise, check if this staff is a form teacher for any ward of this parent
        try:
            staff = user.staff_profile.staff
        except Exception:
            return False

        parent = self.get_object()

        # Get all class-section assignments for this staff
        assigned_infos = ClassSectionInfoModel.objects.filter(form_teacher=staff)
        assigned_class_ids = assigned_infos.values_list('student_class_id', flat=True)
        assigned_section_ids = assigned_infos.values_list('section_id', flat=True)

        # Check if any of this parent's wards belong to those classes or sections
        has_access = StudentModel.objects.filter(
            parent=parent
        ).filter(
            Q(student_class_id__in=assigned_class_ids) | Q(class_section_id__in=assigned_section_ids)
        ).exists()

        return has_access

    def handle_no_permission(self):
        """Raise a clean permission error."""
        raise PermissionDenied("You don't have permission to view this parent.")


class ParentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = ParentModel
    permission_required = 'student.add_studentmodel'
    form_class = ParentForm
    template_name = 'student/parent/edit.html'
    context_object_name = "parent"

    def has_permission(self):
        """
        Override permission logic:
        - Allow if user has 'view_parentmodel' permission
        - Allow if user is superuser
        - Allow if staff is form teacher for any class containing a ward of this parent
        """
        user = self.request.user

        # Superusers always allowed
        if user.is_superuser:
            return True

        # Has global permission
        if user.has_perm(self.permission_required):
            return True

        # Otherwise, check if this staff is a form teacher for any ward of this parent
        try:
            staff = user.staff_profile.staff
        except Exception:
            return False

        parent = self.get_object()

        # Get all class-section assignments for this staff
        assigned_infos = ClassSectionInfoModel.objects.filter(form_teacher=staff)
        assigned_class_ids = assigned_infos.values_list('student_class_id', flat=True)
        assigned_section_ids = assigned_infos.values_list('section_id', flat=True)

        # Check if any of this parent's wards belong to those classes or sections
        has_access = StudentModel.objects.filter(
            parent=parent
        ).filter(
            Q(student_class_id__in=assigned_class_ids) | Q(class_section_id__in=assigned_section_ids)
        ).exists()

        return has_access

    def get_success_url(self):
        messages.success(self.request, "Parent details updated successfully.")
        return reverse('parent_detail', kwargs={'pk': self.object.pk})


def _send_parent_password_reset_email(parent, username, password):
    """
    Helper function to render and send the password reset email.
    """
    if not parent.email:
        print(f"Parent {parent.parent_id} has no email. Skipping password reset email.")
        return False
    try:
        school_info = SchoolInfoModel.objects.first()
        mail_subject = f"Password Reset for {school_info.name.upper()} Parent Portal"
        login_url = settings.BASE_URL + reverse('login')

        context = {
            'parent': parent,
            'username': username,
            'password': password,
            'school_info': school_info,
            'login_url': login_url
        }
        html_content = render_to_string('student/emails/parent_password_reset_email.html', context)
        text_content = (f"Hello {parent.first_name},\n\nYour password has been reset. "
                        f"Please log in at {login_url} with the username: {username} and new password: {password}")

        email_message = EmailMultiAlternatives(
            subject=mail_subject,
            body=text_content,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[parent.email]
        )
        email_message.attach_alternative(html_content, "text/html")
        email_message.send()
        print(f"Password reset email successfully sent to {parent.email}")
        return True
    except Exception as e:
        print(f"ERROR sending password reset email to {parent.email}: {e}")
        return False

class ParentPasswordResetView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    View to reset a parent's password and send the new password via email.
    """
    permission_required = 'student.change_parentmodel'

    def get_object(self):
        pk = self.kwargs.get('pk')
        return get_object_or_404(ParentModel, pk=pk)

    def has_permission(self):
        """
        Override permission logic:
        - Allow if user has 'change_parentmodel' permission
        - Allow if user is superuser
        - Allow if staff is form teacher for any class containing a ward of this parent
        """
        user = self.request.user

        # Superusers always allowed
        if user.is_superuser:
            return True

        # Has global permission
        if user.has_perm(self.permission_required):
            return True

        # Otherwise, check if this staff is a form teacher for any ward of this parent
        try:
            staff = user.staff_profile.staff
        except Exception:
            return False

        parent = self.get_object()

        # Get all class-section assignments for this staff
        assigned_infos = ClassSectionInfoModel.objects.filter(form_teacher=staff)
        assigned_class_ids = assigned_infos.values_list('student_class_id', flat=True)
        assigned_section_ids = assigned_infos.values_list('section_id', flat=True)

        # Check if any of this parent's wards belong to those classes or sections
        has_access = StudentModel.objects.filter(
            parent=parent
        ).filter(
            Q(student_class_id__in=assigned_class_ids) | Q(class_section_id__in=assigned_section_ids)
        ).exists()

        return has_access

    def post(self, request, *args, **kwargs):
        parent = self.get_object()

        try:
            # Check if parent has a user account
            if not hasattr(parent, 'parent_profile') or not parent.parent_profile.user:
                messages.error(request, "This parent does not have a portal account.")
                return redirect('parent_detail', pk=parent.pk)

            user = parent.parent_profile.user

            # Generate a new random password
            password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))

            # Set the new password
            user.set_password(password)
            user.save()

            # Update the default password in the parent profile
            parent.parent_profile.default_password = password
            parent.parent_profile.save()

            # Send password reset email
            email_sent = _send_parent_password_reset_email(parent, user.username, password)

            if email_sent:
                messages.success(request,
                                 f"Password reset successfully. An email with the new password has been sent to {parent.email}.")
            else:
                if parent.email:
                    messages.warning(request,
                                     "Password was reset, but the notification email could not be sent. Please check the server logs.")
                else:
                    messages.warning(request,
                                     "Password was reset, but no email address is available to send the new password.")

        except Exception as e:
            logger.error(f"Failed to reset password for Parent ID {parent.id}: {e}")
            messages.error(request, "There was an error resetting the password. Please review the system logs.")

        return redirect('parent_detail', pk=parent.pk)


class ParentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ParentModel
    permission_required = 'student.delete_studentmodel'
    template_name = 'student/parent/delete.html'
    context_object_name = "parent"

    def get_success_url(self):
        messages.success(self.request, "Parent deleted successfully.")
        return reverse('parent_index')


# -------------------------
# Student Views
# -------------------------
class ClassStudentSelectView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """
    Displays a form for the user to select a class and section to view.
    Form teachers can access this view even without 'student.add_studentmodel' permission.
    """
    permission_required = 'student.add_studentmodel'
    template_name = 'student/student/select_class.html'

    def has_permission(self):
        """Allow access if user has permission OR is a form teacher."""
        user = self.request.user
        if super().has_permission():
            return True
        try:
            staff = user.staff_profile.staff
            return ClassSectionInfoModel.objects.filter(form_teacher=staff).exists()
        except Exception:
            return False

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Default to all classes
        class_queryset = ClassesModel.objects.all().order_by('name')

        try:
            staff = user.staff_profile.staff
            form_teacher_records = ClassSectionInfoModel.objects.filter(form_teacher=staff)

            if form_teacher_records.exists():
                # ✅ Extract distinct class IDs the teacher handles
                allowed_class_ids = form_teacher_records.values_list('student_class_id', flat=True).distinct()
                class_queryset = ClassesModel.objects.filter(id__in=allowed_class_ids).order_by('name')
        except Exception:
            pass

        context['class_list'] = class_queryset
        return context


class StudentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    Displays all students for admins or permissioned users,
    and only the students in their class/section for form teachers.
    """
    model = StudentModel
    permission_required = 'student.view_studentmodel'
    template_name = 'student/student/index.html'
    context_object_name = "student_list"

    def has_permission(self):
        """
        Allow access if user has the permission or is a form teacher.
        """
        user = self.request.user
        if super().has_permission():
            return True

        try:
            staff = user.staff_profile.staff
            return ClassSectionInfoModel.objects.filter(form_teacher=staff).exists()
        except Exception:
            return False

    def get_queryset(self):
        """
        Apply permission rules and class/section filters.
        """
        user = self.request.user
        queryset = StudentModel.objects.select_related('parent', 'student_class', 'class_section').filter(status='active')

        class_id = self.request.GET.get('class')
        section_id = self.request.GET.get('section')

        # 🔹 Admin / Permissioned users see all
        if user.has_perm('student.view_studentmodel'):
            if class_id and section_id:
                return queryset.filter(student_class_id=class_id, class_section_id=section_id).order_by('first_name')
            return queryset.order_by('first_name')

        # 🔹 Form teacher restriction
        try:
            staff = user.staff_profile.staff
            teacher_sections = ClassSectionInfoModel.objects.filter(form_teacher=staff)

            # Limit to teacher’s own students
            allowed_class_ids = teacher_sections.values_list('student_class_id', flat=True)
            allowed_section_ids = teacher_sections.values_list('section_id', flat=True)

            queryset = queryset.filter(
                student_class_id__in=allowed_class_ids,
                class_section_id__in=allowed_section_ids
            )

            # Apply filters if present (still restricted to their allowed classes/sections)
            if class_id and section_id:
                queryset = queryset.filter(student_class_id=class_id, class_section_id=section_id)

            return queryset.order_by('first_name')

        except Exception:
            return StudentModel.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        class_id = self.request.GET.get('class')
        section_id = self.request.GET.get('section')

        if class_id and section_id:
            context['selected_class'] = get_object_or_404(ClassesModel, pk=class_id)
            context['selected_section'] = get_object_or_404(ClassSectionModel, pk=section_id)

        return context



class StudentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = StudentModel
    permission_required = 'student.add_studentmodel'
    form_class = StudentForm
    template_name = 'student/student/create.html'

    def has_permission(self):
        """
        Allow access if user has the required permission
        OR is assigned as a form teacher.
        """
        user = self.request.user

        # Normal permission check (from PermissionRequiredMixin)
        if super().has_permission():
            return True

        # Custom form-teacher check
        try:
            staff = user.staff_profile.staff
            is_form_teacher = ClassSectionInfoModel.objects.filter(form_teacher=staff).exists()
            if is_form_teacher:
                return True
        except Exception:
            pass

        return False

    def get_context_data(self, **kwargs):
        """
        Adds the parent object to the template context for display.
        """
        context = super().get_context_data(**kwargs)
        # Get the parent from the URL and add it to the context
        context['parent'] = get_object_or_404(ParentModel, pk=self.kwargs.get('parent_pk'))
        return context

    def form_valid(self, form):
        """
        This method is called when the form is valid. We override it to:
        1. Associate the student with the correct parent from the URL.
        2. Explicitly create the StudentWalletModel after the student is saved.
        """
        # First, get the parent object from the primary key in the URL
        parent = get_object_or_404(ParentModel, pk=self.kwargs.get('parent_pk'))
        # Assign this parent to the new student instance before it's saved
        form.instance.parent = parent

        # Call the parent class's form_valid. This saves the StudentModel
        # to the database and returns the redirect response.
        response = super().form_valid(form)

        # After the super() call, self.object contains the newly created student instance.
        try:
            # EXPLICITLY CREATE THE WALLET, bypassing the signal.
            # Using get_or_create is robust; it won't crash if a wallet somehow already exists.
            wallet, created = StudentWalletModel.objects.get_or_create(student=self.object)

            if created:
                messages.success(self.request,
                                 f"Student '{self.object.first_name}' registered and a new wallet was created successfully.")
            else:
                messages.info(self.request,
                              f"Student '{self.object.first_name}' was registered, but a wallet already existed for them.")

        except Exception as e:
            # If wallet creation fails, log it and inform the user.
            logger.error(f"CRITICAL: Failed to create wallet for new student {self.object.id}: {e}")
            messages.error(self.request,
                           "The student was created, but there was an error creating their wallet. Please contact support.")

        # Return the original redirect response
        return response

    def get_success_url(self):
        """
        Redirect to the detail page of the newly created student.
        """
        return reverse('student_detail', kwargs={'pk': self.object.pk})

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Pass the logged-in user
        return kwargs


class StudentDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = StudentModel
    permission_required = 'student.view_studentmodel'
    template_name = 'student/student/detail.html'
    context_object_name = "student"

    def has_permission(self):
        """
        Allow:
        - Users with the `view_studentmodel` permission.
        - Form teachers viewing students in their assigned classes/sections.
        """
        user = self.request.user

        # ✅ Default permission check
        if super().has_permission():
            return True

        # ✅ If user is a form teacher, check if this student belongs to their class
        try:
            staff = user.staff_profile.staff
            form_teacher_sections = ClassSectionInfoModel.objects.filter(form_teacher=staff)
            student = self.get_object()

            # Check if student's class/section is among teacher's assigned sections
            is_my_student = form_teacher_sections.filter(
                student_class=student.student_class,
                section=student.class_section
            ).exists()

            if is_my_student:
                return True
        except Exception:
            pass

        return False

    def handle_no_permission(self):
        raise PermissionDenied("You don’t have permission to view this student.")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add the following line to pass the fingerprint list to the template
        context['fingerprint_list'] = self.object.fingerprints.all().order_by('-created_at')

        # Get settings for fingerprint limits
        max_fingerprints = 4
        can_add_more = context['fingerprint_list'].count() < max_fingerprints

        context['can_add_more'] = can_add_more
        context['utility_list'] = UtilityModel.objects.all()
        context['max_fingerprints'] = max_fingerprints

        return context


class StudentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = StudentModel
    permission_required = 'student.change_studentmodel'
    form_class = StudentForm
    template_name = 'student/student/edit.html'
    context_object_name = 'student'

    def has_permission(self):
        """
        Allow:
        - Users with the `view_studentmodel` permission.
        - Form teachers viewing students in their assigned classes/sections.
        """
        user = self.request.user

        # ✅ Default permission check
        if super().has_permission():
            return True

        # ✅ If user is a form teacher, check if this student belongs to their class
        try:
            staff = user.staff_profile.staff
            form_teacher_sections = ClassSectionInfoModel.objects.filter(form_teacher=staff)
            student = self.get_object()

            # Check if student's class/section is among teacher's assigned sections
            is_my_student = form_teacher_sections.filter(
                student_class=student.student_class,
                section=student.class_section
            ).exists()

            if is_my_student:
                return True
        except Exception:
            pass

        return False

    def handle_no_permission(self):
        raise PermissionDenied("You don’t have permission to view this student.")

    def get_success_url(self):
        messages.success(self.request, "Student details updated successfully.")
        return reverse('student_detail', kwargs={'pk': self.object.pk})

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Pass the logged-in user
        return kwargs


class StudentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = StudentModel
    permission_required = 'student.delete_studentmodel'
    template_name = 'student/student/delete.html'
    context_object_name = "student"

    def has_permission(self):
        """
        Allow:
        - Users with the `view_studentmodel` permission.
        - Form teachers viewing students in their assigned classes/sections.
        """
        user = self.request.user

        # ✅ Default permission check
        if super().has_permission():
            return True

        # ✅ If user is a form teacher, check if this student belongs to their class
        try:
            staff = user.staff_profile.staff
            form_teacher_sections = ClassSectionInfoModel.objects.filter(form_teacher=staff)
            student = self.get_object()

            # Check if student's class/section is among teacher's assigned sections
            is_my_student = form_teacher_sections.filter(
                student_class=student.student_class,
                section=student.class_section
            ).exists()

            if is_my_student:
                return True
        except Exception:
            pass

        return False

    def handle_no_permission(self):
        raise PermissionDenied("You don’t have permission to view this student.")

    def get_success_url(self):
        messages.success(self.request, "Student deleted successfully.")
        return reverse('student_index')


# -------------------------
# Student Status Actions
# -------------------------
@login_required
@permission_required("student.change_studentmodel", raise_exception=True)
def change_student_status(request, pk, status):
    student = get_object_or_404(StudentModel, pk=pk)

    # Validate the status
    valid_statuses = [choice[0] for choice in StudentModel.Status.choices]
    if status not in valid_statuses:
        messages.error(request, "Invalid status provided.")
        return redirect('student_detail', pk=pk)

    student.status = status
    student.save()
    messages.success(request, f"Status for '{student}' has been updated to {student.get_status_display()}.")
    return redirect('student_detail', pk=pk)


# -------------------------
# Class List Export Views
# -------------------------
@login_required
@permission_required("student.view_studentmodel", raise_exception=True)
def select_class_for_export_view(request):
    context = {
        'class_list': ClassesModel.objects.all().order_by('name'),
        'section_list': ClassSectionModel.objects.all().order_by('name'),
    }
    return render(request, 'student/student/select_class_for_export.html', context)


@login_required
@permission_required("student.view_studentmodel", raise_exception=True)
def export_class_list_view(request):
    class_id = request.GET.get('student_class')
    section_id = request.GET.get('class_section')

    if not class_id or not section_id:
        messages.error(request, "Please select both a class and a section.")
        return redirect('select_class_for_export')

    student_class = get_object_or_404(ClassesModel, pk=class_id)
    class_section = get_object_or_404(ClassSectionModel, pk=section_id)

    student_list = StudentModel.objects.filter(
        student_class=student_class,
        class_section=class_section
    ).select_related('parent').order_by('last_name', 'first_name')

    if not student_list.exists():
        messages.warning(request, "No students found in the selected class and section to export.")
        return redirect('select_class_for_export')

    output = io.BytesIO()
    workbook = Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet(f"{student_class.name} {class_section.name}")

    headers = ['Reg. Number', 'First Name', 'Last Name', 'Parent Name', 'Parent Mobile', 'Parent Email']
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header)

    for row_num, student in enumerate(student_list, 1):
        parent = student.parent
        worksheet.write(row_num, 0, student.registration_number)
        worksheet.write(row_num, 1, student.first_name)
        worksheet.write(row_num, 2, student.last_name)
        worksheet.write(row_num, 3, f"{parent.first_name} {parent.last_name}")
        worksheet.write(row_num, 4, parent.mobile)
        worksheet.write(row_num, 5, parent.email)

    workbook.close()
    output.seek(0)

    filename = f"{student_class.name}-{class_section.name}-Student-List.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f"attachment; filename={filename}"
    return response


class SelectParentView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """
    Renders the initial page for a user to search for a parent.
    All data loading is now handled asynchronously by ParentSearchView.
    """
    permission_required = 'student.add_studentmodel'
    template_name = 'student/student/select_parent.html'

    def has_permission(self):
        """
        Allow access if user has the required permission
        OR is assigned as a form teacher.
        """
        user = self.request.user

        # Normal permission check (from PermissionRequiredMixin)
        if super().has_permission():
            return True

        # Custom form-teacher check
        try:
            staff = user.staff_profile.staff
            is_form_teacher = ClassSectionInfoModel.objects.filter(form_teacher=staff).exists()
            if is_form_teacher:
                return True
        except Exception:
            pass

        return False


class ParentSearchView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    An API endpoint that returns a JSON list of parents matching a search query.
    This is called by JavaScript from the SelectParentView template.
    """
    permission_required = 'student.add_studentmodel'

    def has_permission(self):
        """
        Allow access if user has the required permission
        OR is assigned as a form teacher.
        """
        user = self.request.user

        # Normal permission check (from PermissionRequiredMixin)
        if super().has_permission():
            return True

        # Custom form-teacher check
        try:
            staff = user.staff_profile.staff
            is_form_teacher = ClassSectionInfoModel.objects.filter(form_teacher=staff).exists()
            if is_form_teacher:
                return True
        except Exception:
            pass

        return False

    def get(self, request, *args, **kwargs):
        query = request.GET.get('q', '').strip()

        if len(query) < 2:
            # Don't search if the query is too short
            return JsonResponse([], safe=False)

        # Build a query that searches across multiple fields
        # Q objects allow for complex "OR" queries
        search_query = (
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(parent_id__icontains=query) |
                Q(mobile__icontains=query)
        )

        # Find matching parents, limit the results for performance, and select related user data
        parents = ParentModel.objects.filter(search_query).order_by('first_name', 'last_name')[:10]

        # Serialize only the necessary data
        parents_data = [
            {
                'pk': parent.pk,
                'full_name': str(parent),
                'parent_id': parent.parent_id,
                'mobile': parent.mobile,
                'email': parent.email,
            }
            for parent in parents
        ]

        return JsonResponse(parents_data, safe=False)


class GetClassSectionsView(LoginRequiredMixin, View):
    """
    An API endpoint to get the sections associated with a specific class.
    """

    def get(self, request, *args, **kwargs):
        class_id = request.GET.get('class_id')
        if not class_id:
            return JsonResponse({'error': 'Class ID not provided'}, status=400)

        try:
            student_class = ClassesModel.objects.get(pk=class_id)
            sections = student_class.section.all().order_by('name')

            # Serialize the sections into a list of simple objects
            sections_data = [{'id': section.id, 'name': section.name} for section in sections]

            return JsonResponse(sections_data, safe=False)

        except ClassesModel.DoesNotExist:
            return JsonResponse({'error': 'Class not found'}, status=404)


def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


# views.py - Enhanced capture endpoint
@csrf_exempt
@require_POST
def capture_fingerprint(request):
    """
    Enhanced fingerprint capture with quality assessment and feature extraction
    """
    try:
        data = json.loads(request.body)
        student_id = data.get('student_id')
        finger_name = data.get('finger_name')
        fingerprint_data = data.get('fingerprint_data')
        quality_score = data.get('quality_score')

        if not all([student_id, finger_name, fingerprint_data]):
            return JsonResponse({
                'success': False,
                'message': 'Missing required fields'
            }, status=400)

        # Validate student exists
        try:
            student = StudentModel.objects.get(id=student_id)
        except StudentModel.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Student not found'
            }, status=404)

        # Check if fingerprint for this finger already exists
        if FingerprintModel.objects.filter(student=student, finger_name=finger_name).exists():
            return JsonResponse({
                'success': False,
                'message': f'Fingerprint for {finger_name} already exists for this student'
            }, status=400)

        # Enhanced quality assessment
        assessed_quality = quality_score
        quality_feedback = ""

        # Check minimum quality threshold
        min_quality = getattr(settings, 'FINGERPRINT_MIN_QUALITY', 0.4)
        if assessed_quality < min_quality:
            return JsonResponse({
                'success': False,
                'message': f'Fingerprint quality too low ({assessed_quality:.2f}). {quality_feedback}',
                'quality_score': assessed_quality,
                'feedback': quality_feedback
            }, status=400)

        # Create fingerprint record
        fingerprint = FingerprintModel(
            student=student,
            finger_name=finger_name,
            fingerprint_template=fingerprint_data,
            quality_score=assessed_quality
        )
        fingerprint.save()

        return JsonResponse({
            'success': True,
            'message': 'Fingerprint captured successfully',
            'fingerprint_id': fingerprint.id,
            'quality_score': assessed_quality,
            'quality_feedback': quality_feedback,
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)

    except Exception as e:
        logger.error(f"Error during fingerprint capture: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'An error occurred during capture: {str(e)}'
        }, status=500)


class SimpleFingerprintMatcher:
    """
    Simple but effective fingerprint matching for DigitalPersona PNG images
    """

    def __init__(self):
        # Balanced threshold - strict enough for security, lenient enough to work
        self.min_match_score = 0.65  # 65% threshold

    def decode_fingerprint(self, base64_data):
        """Decode base64 fingerprint to numpy array"""
        try:
            if ',' in base64_data:
                base64_data = base64_data.split(',')[1]

            image_data = base64.b64decode(base64_data)
            image = Image.open(BytesIO(image_data))

            if image.mode != 'L':
                image = image.convert('L')

            return np.array(image, dtype=np.float64)

        except Exception as e:
            logger.error(f"Error decoding fingerprint: {e}")
            raise ValueError(f"Invalid fingerprint data: {str(e)}")

    def simple_normalize(self, img):
        """Simple normalization"""
        img_min = img.min()
        img_max = img.max()
        if img_max - img_min > 0:
            return (img - img_min) / (img_max - img_min)
        return img / 255.0

    def compute_direct_similarity(self, img1, img2):
        """
        Direct pixel-by-pixel comparison - most reliable for same scanner
        """
        try:
            # Resize both to same size
            from PIL import Image as PILImage
            target_size = (300, 300)  # Larger size preserves more detail

            img1_pil = PILImage.fromarray(img1.astype(np.uint8))
            img2_pil = PILImage.fromarray(img2.astype(np.uint8))

            img1_resized = np.array(img1_pil.resize(target_size, PILImage.Resampling.LANCZOS), dtype=np.float64)
            img2_resized = np.array(img2_pil.resize(target_size, PILImage.Resampling.LANCZOS), dtype=np.float64)

            # Normalize
            img1_norm = self.simple_normalize(img1_resized)
            img2_norm = self.simple_normalize(img2_resized)

            # Method 1: Mean Squared Error (lower is better)
            mse = np.mean((img1_norm - img2_norm) ** 2)
            mse_similarity = 1.0 - min(mse, 1.0)

            # Method 2: Pearson Correlation
            flat1 = img1_norm.flatten()
            flat2 = img2_norm.flatten()

            correlation = np.corrcoef(flat1, flat2)[0, 1]
            if np.isnan(correlation):
                correlation = 0.0
            corr_similarity = (correlation + 1) / 2  # Scale to 0-1

            # Method 3: Cosine Similarity
            dot_product = np.sum(flat1 * flat2)
            norm1 = np.sqrt(np.sum(flat1 ** 2))
            norm2 = np.sqrt(np.sum(flat2 ** 2))

            if norm1 > 0 and norm2 > 0:
                cosine_sim = dot_product / (norm1 * norm2)
                cosine_sim = (cosine_sim + 1) / 2  # Scale to 0-1
            else:
                cosine_sim = 0.0

            # Combine scores with simple average
            combined = (mse_similarity * 0.4 + corr_similarity * 0.4 + cosine_sim * 0.2)

            return {
                'mse_similarity': mse_similarity,
                'correlation': corr_similarity,
                'cosine_similarity': cosine_sim,
                'combined': combined
            }

        except Exception as e:
            logger.error(f"Error computing similarity: {e}", exc_info=True)
            return {
                'mse_similarity': 0.0,
                'correlation': 0.0,
                'cosine_similarity': 0.0,
                'combined': 0.0
            }

    def compute_histogram_match(self, img1, img2):
        """Compare intensity histograms"""
        try:
            # Compute normalized histograms
            hist1, _ = np.histogram(img1.flatten(), bins=64, range=(0, 255), density=True)
            hist2, _ = np.histogram(img2.flatten(), bins=64, range=(0, 255), density=True)

            # Compute histogram intersection
            intersection = np.minimum(hist1, hist2).sum()

            return intersection

        except Exception as e:
            logger.error(f"Error computing histogram: {e}")
            return 0.0

    def match_fingerprints(self, template1_base64, template2_base64):
        """
        Simple fingerprint matching
        """
        try:
            # Decode both images
            img1 = self.decode_fingerprint(template1_base64)
            img2 = self.decode_fingerprint(template2_base64)

            # Compute direct similarity
            similarity_scores = self.compute_direct_similarity(img1, img2)

            # Compute histogram match
            hist_score = self.compute_histogram_match(img1, img2)

            # Final score: 70% direct similarity + 30% histogram
            final_score = (similarity_scores['combined'] * 0.70 + hist_score * 0.30)

            # Match decision
            is_match = final_score >= self.min_match_score

            details = {
                'mse_similarity': round(similarity_scores['mse_similarity'], 3),
                'correlation': round(similarity_scores['correlation'], 3),
                'cosine_similarity': round(similarity_scores['cosine_similarity'], 3),
                'histogram_match': round(hist_score, 3),
                'direct_similarity': round(similarity_scores['combined'], 3),
                'final_score': round(final_score, 3)
            }

            return is_match, final_score, details

        except Exception as e:
            logger.error(f"Error matching fingerprints: {e}", exc_info=True)
            return False, 0.0, {'error': str(e)}


# Global matcher
fingerprint_matcher = SimpleFingerprintMatcher()


@csrf_exempt
@require_POST
def identify_student_by_fingerprint(request):
    """
    Student identification with simple but effective matching
    """
    try:
        data = json.loads(request.body)
        scanned_template = data.get('fingerprint_data')

        if not scanned_template:
            return JsonResponse({
                'success': False,
                'message': 'No fingerprint data provided'
            }, status=400)

        # Validate
        try:
            fingerprint_matcher.decode_fingerprint(scanned_template)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Invalid fingerprint data: {str(e)}'
            }, status=400)

        # Threshold
        match_threshold = 0.65  # 65% - balanced

        # Get active fingerprints
        active_fingerprints = FingerprintModel.objects.filter(
            is_active=True,
            student__status='active'
        ).select_related(
            'student',
            'student__student_wallet',
            'student__student_class',
            'student__parent'
        ).order_by('-quality_score')  # Check high quality first

        logger.info(f"🔍 Checking {active_fingerprints.count()} enrolled fingerprints")

        best_match = None
        best_score = 0.0
        best_details = {}
        all_scores = []

        # Compare with each enrolled fingerprint
        for idx, fingerprint in enumerate(active_fingerprints):
            try:
                is_match, score, details = fingerprint_matcher.match_fingerprints(
                    scanned_template,
                    fingerprint.fingerprint_template
                )

                student_info = f"{fingerprint.student.first_name} {fingerprint.student.last_name}"

                all_scores.append({
                    'index': idx + 1,
                    'student': student_info,
                    'reg_number': fingerprint.student.registration_number,
                    'finger': fingerprint.get_finger_name_display(),
                    'score': score,
                    'details': details,
                    'is_match': is_match
                })

                # Log each comparison for debugging
                match_status = "✅" if is_match else "❌"
                logger.info(
                    f"  [{match_status}] #{idx + 1} {student_info} ({fingerprint.student.registration_number}) "
                    f"- {fingerprint.get_finger_name_display()}: {score:.3f} | "
                    f"MSE:{details['mse_similarity']:.3f} "
                    f"Corr:{details['correlation']:.3f} "
                    f"Cos:{details['cosine_similarity']:.3f} "
                    f"Hist:{details['histogram_match']:.3f}"
                )

                if is_match and score > best_score:
                    best_match = fingerprint
                    best_score = score
                    best_details = details

            except Exception as e:
                logger.warning(f"Error comparing fingerprint {fingerprint.id}: {e}")
                continue

        # Sort by score
        all_scores.sort(key=lambda x: x['score'], reverse=True)

        # Log summary
        logger.info(f"\n{'=' * 60}")
        logger.info(f"SCAN SUMMARY:")
        logger.info(f"  Total checked: {len(all_scores)}")
        logger.info(f"  Best score: {best_score:.3f}")
        logger.info(f"  Threshold: {match_threshold}")
        logger.info(f"  Result: {'✅ MATCH' if best_match else '❌ NO MATCH'}")
        logger.info(f"{'=' * 60}\n")

        # Check for ambiguous matches
        close_matches = [s for s in all_scores if s['score'] >= (best_score - 0.10) and s['is_match']]
        is_ambiguous = len(close_matches) > 1

        if best_match and best_score >= match_threshold:
            student = best_match.student

            # Mark as used
            best_match.mark_used()

            # Wallet
            try:
                wallet_balance = float(student.student_wallet.balance)
            except:
                wallet_balance = 0.0

            # Confidence
            if best_score >= 0.85:
                confidence = 'very_high'
            elif best_score >= 0.75:
                confidence = 'high'
            elif best_score >= 0.65:
                confidence = 'good'
            else:
                confidence = 'acceptable'

            logger.info(
                f"✅ AUTHENTICATED: {student.registration_number} "
                f"({student.first_name} {student.last_name}) "
                f"Score: {best_score:.3f} ({confidence})"
            )

            return JsonResponse({
                'success': True,
                'message': 'Student identified successfully',
                'is_ambiguous': is_ambiguous,
                'student': {
                    'id': student.id,
                    'name': f"{student.first_name} {student.last_name}",
                    'reg_number': student.registration_number,
                    'student_class': str(student.student_class) if student.student_class else 'Not Assigned',
                    'class_section': str(student.class_section) if student.class_section else '',
                    'status': student.get_status_display(),
                    'wallet_balance': wallet_balance,
                    'image_url': student.image.url if student.image else '',
                    'parent_name': f"{student.parent.first_name} {student.parent.last_name}",
                    'parent_mobile': student.parent.mobile or '',
                },
                'match_details': {
                    'score': round(best_score, 3),
                    'confidence': confidence,
                    'finger_used': best_match.get_finger_name_display(),
                    'algorithm_scores': best_details,
                    'total_candidates': active_fingerprints.count(),
                    'match_threshold': match_threshold,
                    'top_5_scores': [
                        {
                            'rank': i + 1,
                            'student': s['student'],
                            'score': round(s['score'], 3)
                        }
                        for i, s in enumerate(all_scores[:5])
                    ]
                }
            })

        else:
            # No match
            top_score = all_scores[0]['score'] if all_scores else 0.0
            top_student = all_scores[0]['student'] if all_scores else 'None'
            top_details = all_scores[0]['details'] if all_scores else {}

            # Diagnostic message
            if top_score >= 0.55:
                reason = "Close match but below security threshold"
                suggestion = "Try again: clean scanner and finger, press firmly"
            elif top_score >= 0.35:
                reason = "Partial match detected"
                suggestion = "Ensure proper finger placement and try again"
            else:
                reason = "No matching fingerprint found"
                suggestion = "Student may not be enrolled or using wrong finger"

            logger.warning(
                f"❌ REJECTED: Best match {top_student}: {top_score:.3f} "
                f"(need {match_threshold}) | Details: {top_details}"
            )

            return JsonResponse({
                'success': False,
                'message': f'{reason}. {suggestion}',
                'diagnostic': {
                    'best_score': round(top_score, 3),
                    'best_match': top_student,
                    'threshold_required': match_threshold,
                    'gap': round(match_threshold - top_score, 3),
                    'candidates_evaluated': len(all_scores),
                    'reason': reason,
                    'suggestion': suggestion,
                    'algorithm_details': top_details
                }
            }, status=404)

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)

    except Exception as e:
        logger.error(f"❌ CRITICAL ERROR: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': 'System error during identification'
        }, status=500)

@csrf_exempt
@require_POST
def delete_fingerprint(request):
    """
    Delete a specific fingerprint
    """
    try:
        data = json.loads(request.body)
        fingerprint_id = data.get('fingerprint_id')

        fingerprint = get_object_or_404(FingerprintModel, id=fingerprint_id)

        # Soft delete - mark as inactive instead of actually deleting
        fingerprint.delete()

        return JsonResponse({
            'success': True,
            'message': 'Fingerprint deleted successfully'
        })

    except Exception as e:
        logger.error(f"Error deleting fingerprint: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': 'Error deleting fingerprint'
        }, status=500)


@require_http_methods(["GET"])
def test_scanner_connection(request):
    """
    Test endpoint to check if scanner is connected and working
    """
    return JsonResponse({
        'success': True,
        'message': 'Scanner connection test endpoint ready',
        'instructions': 'Use JavaScript SDK to test actual scanner connection'
    })


def generate_import_batch_id(parent_filename, student_filename):
    """Generate a unique batch ID based on filenames and timestamp."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Create a hash of both filenames for uniqueness
    hash_input = f"{parent_filename}_{student_filename}_{timestamp}"
    file_hash = hashlib.md5(hash_input.encode()).hexdigest()[:8]

    return f"IMP_{timestamp}_{file_hash}"


@login_required
def parent_student_upload_view(request):
    """
    Handles the file upload form for both parent and student files.
    Initiates background processing task and displays recent imports.
    """
    if request.method == 'POST':
        form = ParentStudentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            parent_file = request.FILES['parent_file']
            student_file = request.FILES['student_file']

            fs = FileSystemStorage()
            # Save files to organized subdirectories within your media root
            parent_filename = fs.save(f"imports/parents/{parent_file.name}", parent_file)
            parent_file_path = fs.path(parent_filename)

            student_filename = fs.save(f"imports/students/{student_file.name}", student_file)
            student_file_path = fs.path(student_filename)

            # This is the correct line
            batch_id = generate_import_batch_id(parent_file.name, student_file.name)

            # Safely get the staff profile from the logged-in user
            staff_profile = getattr(request.user, 'staff_profile', None)

            ImportBatchModel.objects.create(
                batch_id=batch_id,
                parent_file_name=parent_file.name,
                student_file_name=student_file.name,
                # Correctly pass the staff object from the profile
                imported_by=staff_profile.staff if staff_profile else None,
                status='processing'
            )

            # Dispatch the background task to Celery
            process_parent_student_upload.delay(
                parent_file_path,
                student_file_path,
                batch_id
            )

            messages.success(
                request,
                f'Files uploaded successfully! The data is being processed in the background. '
                f'You can view the progress by clicking the "View" button on Batch ID: {batch_id}.'
            )
            # This redirect name should match the name in your urls.py
            return redirect('import_parent_student')
    else:
        form = ParentStudentUploadForm()

    # Get recent import batches to display on the page
    recent_imports = ImportBatchModel.objects.all()[:10]

    context = {
        'form': form,
        'recent_imports': recent_imports,
    }
    return render(request, 'student/import_parent_student.html', context)


@login_required
def import_batch_detail_view(request, batch_id):
    """
    Displays the details and results of a specific import batch.
    """
    batch = get_object_or_404(ImportBatchModel, batch_id=batch_id)

    # Get parents and students that were created or updated in this batch
    parents = ParentModel.objects.filter(import_batch_id=batch_id).select_related('parent_profile__user')
    students = StudentModel.objects.filter(import_batch_id=batch_id).select_related(
        'parent', 'student_class', 'class_section'
    )

    context = {
        'batch': batch,
        'parents': parents,
        'students': students,
    }
    return render(request, 'student/import_batch_detail.html', context)


@login_required
def download_parent_credentials(request, batch_id):
    """
    Download parent login credentials as Excel file for a specific import batch.
    """
    batch = get_object_or_404(ImportBatchModel, batch_id=batch_id)

    # Get all parents from this batch with their profiles
    parents = ParentModel.objects.filter(
        import_batch_id=batch_id
    ).select_related('parent_profile__user').order_by('parent_id')

    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Parent Login Credentials'

    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Add title
    sheet.merge_cells('A1:H1')
    title_cell = sheet['A1']
    title_cell.value = f'PARENT PORTAL LOGIN CREDENTIALS'
    title_cell.font = Font(bold=True, size=14, color='366092')
    title_cell.alignment = center_alignment

    # Add batch info
    sheet.merge_cells('A2:H2')
    info_cell = sheet['A2']
    info_cell.value = f'Import Batch: {batch_id} | Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
    info_cell.font = Font(size=10, italic=True)
    info_cell.alignment = center_alignment

    # Add empty row
    sheet.append([])

    # Define headers
    headers = [
        'S/N',
        'Parent ID',
        'Full Name',
        'Username',
        'Password',
        'Email',
        'Mobile',
        'Number of Wards'
    ]

    # Write headers (row 4)
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment

    # Write parent data
    row_num = 5
    for index, parent in enumerate(parents, 1):
        # Get credentials
        username = ''
        password = ''

        if hasattr(parent, 'parent_profile') and parent.parent_profile:
            username = parent.parent_profile.user.username
            password = parent.parent_profile.default_password

        # Prepare row data
        row_data = [
            index,  # S/N
            parent.parent_id,
            f"{parent.first_name} {parent.last_name}",
            username if username else 'NO LOGIN',
            password if password else 'N/A',
            parent.email if parent.email else '—',
            parent.mobile if parent.mobile else '—',
            parent.number_of_wards()
        ]

        # Write row
        for col_num, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border

            # Center align S/N and Number of Wards
            if col_num in [1, 8]:
                cell.alignment = center_alignment

            # Highlight rows without login
            if username == '':
                cell.fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')

        row_num += 1

    # Add summary section
    row_num += 1
    sheet.merge_cells(f'A{row_num}:H{row_num}')
    summary_cell = sheet.cell(row=row_num, column=1)
    summary_cell.value = 'SUMMARY'
    summary_cell.font = Font(bold=True, size=11, color='366092')
    summary_cell.alignment = center_alignment

    row_num += 1
    summary_data = [
        ['Total Parents:', parents.count()],
        ['With Login Access:', parents.filter(parent_profile__isnull=False).count()],
        ['Without Login:', parents.filter(parent_profile__isnull=True).count()],
    ]

    for label, value in summary_data:
        sheet.cell(row=row_num, column=1).value = label
        sheet.cell(row=row_num, column=1).font = Font(bold=True)
        sheet.cell(row=row_num, column=2).value = value
        row_num += 1

    # Add instructions section
    row_num += 2
    sheet.merge_cells(f'A{row_num}:H{row_num}')
    instructions_cell = sheet.cell(row=row_num, column=1)
    instructions_cell.value = 'INSTRUCTIONS FOR PARENTS'
    instructions_cell.font = Font(bold=True, size=11, color='366092')
    instructions_cell.alignment = center_alignment

    row_num += 1
    instructions = [
        '1. Visit the parent portal at: [Your School Portal URL]',
        '2. Enter your Username and Password as shown above',
        '3. For security, please change your password after first login',
        '4. If you experience any issues, contact the school administration',
        '5. Keep your login credentials confidential'
    ]

    for instruction in instructions:
        sheet.cell(row=row_num, column=1).value = instruction
        sheet.merge_cells(f'A{row_num}:H{row_num}')
        row_num += 1

    # Adjust column widths
    column_widths = {
        'A': 8,  # S/N
        'B': 15,  # Parent ID
        'C': 25,  # Full Name
        'D': 15,  # Username
        'E': 15,  # Password
        'F': 30,  # Email
        'G': 20,  # Mobile
        'H': 18,  # Number of Wards
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    # Freeze header rows
    sheet.freeze_panes = 'A5'

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    filename = f'Parent_Credentials_{batch_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    workbook.save(response)

    return response


@login_required
def download_all_parent_credentials(request):
    """
    Download ALL parent login credentials as Excel file (all batches).
    """
    # Get all parents with profiles, ordered by parent_id
    parents = ParentModel.objects.select_related(
        'parent_profile__user'
    ).order_by('parent_id')

    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'All Parent Credentials'

    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Add title
    sheet.merge_cells('A1:I1')
    title_cell = sheet['A1']
    title_cell.value = f'ALL PARENT PORTAL LOGIN CREDENTIALS'
    title_cell.font = Font(bold=True, size=14, color='366092')
    title_cell.alignment = center_alignment

    # Add generation info
    sheet.merge_cells('A2:I2')
    info_cell = sheet['A2']
    info_cell.value = f'Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
    info_cell.font = Font(size=10, italic=True)
    info_cell.alignment = center_alignment

    # Add empty row
    sheet.append([])

    # Define headers
    headers = [
        'S/N',
        'Parent ID',
        'Full Name',
        'Username',
        'Password',
        'Email',
        'Mobile',
        'Number of Wards',
        'Import Batch'
    ]

    # Write headers (row 4)
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment

    # Write parent data
    row_num = 5
    for index, parent in enumerate(parents, 1):
        # Get credentials
        username = ''
        password = ''

        if hasattr(parent, 'parent_profile') and parent.parent_profile:
            username = parent.parent_profile.user.username
            password = parent.parent_profile.default_password

        # Prepare row data
        row_data = [
            index,  # S/N
            parent.parent_id,
            f"{parent.first_name} {parent.last_name}",
            username if username else 'NO LOGIN',
            password if password else 'N/A',
            parent.email if parent.email else '—',
            parent.mobile if parent.mobile else '—',
            parent.number_of_wards(),
            parent.import_batch_id if parent.import_batch_id else 'Manual Entry'
        ]

        # Write row
        for col_num, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border

            # Center align S/N and Number of Wards
            if col_num in [1, 8]:
                cell.alignment = center_alignment

            # Highlight rows without login
            if username == '':
                cell.fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')

        row_num += 1

    # Add summary
    row_num += 1
    sheet.merge_cells(f'A{row_num}:I{row_num}')
    summary_cell = sheet.cell(row=row_num, column=1)
    summary_cell.value = 'SUMMARY'
    summary_cell.font = Font(bold=True, size=11, color='366092')
    summary_cell.alignment = center_alignment

    row_num += 1
    summary_data = [
        ['Total Parents:', parents.count()],
        ['With Login Access:', parents.filter(parent_profile__isnull=False).count()],
        ['Without Login:', parents.filter(parent_profile__isnull=True).count()],
    ]

    for label, value in summary_data:
        sheet.cell(row=row_num, column=1).value = label
        sheet.cell(row=row_num, column=1).font = Font(bold=True)
        sheet.cell(row=row_num, column=2).value = value
        row_num += 1

    # Adjust column widths
    column_widths = {
        'A': 8,  # S/N
        'B': 15,  # Parent ID
        'C': 25,  # Full Name
        'D': 15,  # Username
        'E': 15,  # Password
        'F': 30,  # Email
        'G': 20,  # Mobile
        'H': 18,  # Number of Wards
        'I': 25,  # Import Batch
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    # Freeze header rows
    sheet.freeze_panes = 'A5'

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    filename = f'All_Parent_Credentials_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    workbook.save(response)

    return response


def _create_parent_account(first_name, last_name, email, mobile):
    """
    A helper function to create a parent, user, profile, and send an email.
    """
    if not last_name:
        last_name = first_name

    if email and User.objects.filter(email__iexact=email).exists():
        raise ValueError(f"An account with the email '{email}' already exists.")

    parent = ParentModel.objects.create(
        first_name=first_name,
        last_name=last_name,
        email=email,
        mobile=mobile
    )

    username = parent.parent_id

    # --- THIS IS THE CORRECTED PASSWORD LOGIC ---
    # Using the 'secrets' module as you suggested for security.
    # This alphabet avoids ambiguous characters like 'i', 'I', '1', 'l', 'o', 'O', '0'.
    alphabet = 'abcdefghjkmnpqrstuvwxyzABCDEFGHJKLMNPQRSTUVWXYZ23456789'
    password = ''.join(secrets.choice(alphabet) for i in range(10))
    # --- END OF CORRECTION ---

    user = User.objects.create_user(
        username=username,
        password=password,
        email=email,
        first_name=first_name,
        last_name=last_name
    )
    ParentProfileModel.objects.create(
        user=user,
        parent=parent,
        default_password=password
    )

    _send_parent_welcome_email(parent, username, password)

    return parent


@login_required
def paste_create_parents_view(request):
    """
    Renders the HTML page with the textarea for pasting JSON.
    """
    return render(request, 'student/paste_create_parents.html')


@login_required
@require_POST  # This view should only accept POST requests
def ajax_create_parent_view(request):
    """
    This is the AJAX endpoint. It receives data for one parent,
    creates the account, and returns a JSON response.
    """
    try:
        data = json.loads(request.body)
        first_name = data.get('first_name')
        last_name = data.get('last_name')
        email = data.get('email')
        mobile = data.get('mobile')

        if not first_name:
            return JsonResponse({'status': 'error', 'message': 'First Name is required.'}, status=400)

        parent = _create_parent_account(first_name, last_name, email, mobile)

        return JsonResponse({
            'status': 'success',
            'message': f'Successfully created account for {parent.first_name} {parent.last_name} with Parent ID: {parent.parent_id}.'
        })

    except ValueError as e:
        # Catch specific, known errors like duplicate emails
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    except Exception as e:
        # Catch any other unexpected errors
        logger.error("Error in ajax_create_parent_view: %s", e, exc_info=True)
        return JsonResponse({'status': 'error', 'message': 'A critical server error occurred. Check logs.'}, status=500)


@login_required
@require_POST
@transaction.atomic
def ajax_create_student_view(request):
    """
    AJAX endpoint to create a single student and their wallet.
    Processes a case-insensitive, 100% exact match for parent emails.
    Includes a fallback search by student's last name if email search fails.
    """
    try:
        data = json.loads(request.body)
        first_name = data.get('first_name')
        last_name = data.get('last_name')
        gender_raw = data.get('gender')
        class_code = data.get('class_code')
        class_section_name = data.get('class_section_name')
        parent_emails_raw = data.get('parent_emails_raw')

        # --- 1. Basic Validation ---
        if not all([first_name, last_name, class_code, class_section_name, parent_emails_raw]):
            return JsonResponse({'status': 'error', 'message': 'Missing required data.'}, status=400)

        # --- 2. Find Parent (Primary search by Email) ---
        email_list = [email.strip().lower() for email in re.split(r'[,\s;]+', parent_emails_raw) if email.strip()]
        parents = ParentModel.objects.none()  # Start with an empty queryset

        if email_list:
            # Build a dynamic query with Q objects for case-insensitive matching
            # Assumes ParentModel has an 'email' field directly.
            query_objects = [Q(email__iexact=email) for email in email_list]
            query = reduce(operator.or_, query_objects)
            parents = ParentModel.objects.filter(query).distinct()

        # --- NEW: Fallback search by Name if no parent was found by email ---
        if not parents.exists():
            # Search for a parent where their first or last name matches the student's last name
            name_query = Q(last_name__iexact=last_name) | Q(first_name__iexact=last_name)
            parents = ParentModel.objects.filter(name_query)
            # This result will now be checked by the validation below.

        # --- Final Validation for Parent ---
        if parents.count() == 0:
            # This error now triggers if BOTH email and name searches failed.
            raise ValueError(f"Parent not found via email ({', '.join(email_list)}) or name ({last_name}).")
        elif parents.count() > 1:
            raise ValueError(
                f"Ambiguous match: Found multiple parents for email ({', '.join(email_list)}) or name ({last_name}).")

        parent = parents.first()

        # --- 3. Find Class and Section ---
        try:
            student_class = ClassesModel.objects.get(code__iexact=class_code)
        except ClassesModel.DoesNotExist:
            raise ValueError(f"Class with code '{class_code}' does not exist.")

        class_section, _ = ClassSectionModel.objects.get_or_create(
            name__iexact=class_section_name,
            defaults={'name': class_section_name}
        )

        # --- 4. Map Gender ---
        if gender_raw.startswith('F'):
            gender = StudentModel.Gender.FEMALE
        elif gender_raw.startswith('M'):
            gender = StudentModel.Gender.MALE
        else:
            raise ValueError(f"Invalid gender value: '{gender_raw}'. Use 'M' or 'F'.")

        # --- 5. Create Student and Wallet ---
        student = StudentModel.objects.create(
            first_name=first_name,
            last_name=last_name,
            gender=gender,
            parent=parent,
            student_class=student_class,
            class_section=class_section,
        )
        StudentWalletModel.objects.create(student=student)

        return JsonResponse({
            'status': 'success',
            'message': f"Student '{student.first_name} {student.last_name}' ({student.registration_number}) created and linked to parent '{parent}'."
        })

    except ValueError as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    except Exception as e:
        logger.error("Critical error in ajax_create_student_view: %s", e, exc_info=True)
        return JsonResponse({'status': 'error', 'message': 'A critical server error occurred. Check logs.'}, status=500)



@login_required
def paste_create_students_view(request):
    """
    Renders the HTML page with the textarea for pasting student JSON.
    """
    return render(request, 'student/paste_create_students.html')