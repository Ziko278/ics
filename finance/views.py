import calendar
import json
import traceback
from datetime import date, datetime, timedelta
from decimal import Decimal
from reportlab.lib.pagesizes import landscape, A4
from io import BytesIO
import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core import serializers
from django.core.exceptions import ObjectDoesNotExist
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.core.serializers.json import DjangoJSONEncoder
from django.db import transaction
from django.db.models import Q, Sum, Avg, F, DecimalField, Value, Count, Prefetch
from django.db.models.functions import TruncMonth, Coalesce, Concat
from django.forms import modelformset_factory
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect, get_object_or_404, render
from django.urls import reverse, reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.utils import timezone
from django.utils.timezone import now
from django.views import View
from django.views.decorators.http import require_http_methods, require_POST
from django.views.generic import TemplateView, CreateView, UpdateView, ListView, DetailView, DeleteView, FormView
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
from admin_site.models import SessionModel, TermModel, SchoolSettingModel, ClassesModel, ActivityLogModel, \
    SchoolInfoModel
from admin_site.views import FlashFormErrorsMixin
from human_resource.models import StaffModel, StaffProfileModel, StaffWalletModel
from inventory.models import PurchaseOrderModel, PurchaseAdvanceModel, SaleModel, SaleItemModel
from student.models import StudentModel, StudentWalletModel
from student.signals import get_day_ordinal_suffix
from .models import FinanceSettingModel, SupplierPaymentModel, PurchaseAdvancePaymentModel, FeeModel, FeeGroupModel, \
    FeeMasterModel, InvoiceGenerationJob, InvoiceModel, FeePaymentModel, ExpenseCategoryModel, ExpenseModel, \
    IncomeCategoryModel, IncomeModel, TermlyFeeAmountModel, StaffBankDetail, SalaryRecord, SalaryAdvance, \
    SalaryStructure, StudentFundingModel, InvoiceItemModel, AdvanceSettlementModel, \
    SchoolBankDetail, StaffLoan, StaffLoanRepayment, StaffFundingModel, DiscountModel, DiscountApplicationModel, \
    StudentDiscountModel, OtherPaymentClearanceModel, OtherPaymentModel, SalarySetting, Bonus
from .forms import FinanceSettingForm, SupplierPaymentForm, PurchaseAdvancePaymentForm, FeeForm, FeeGroupForm, \
    InvoiceGenerationForm, FeePaymentForm, ExpenseCategoryForm, ExpenseForm, IncomeCategoryForm, \
    IncomeForm, TermlyFeeAmountFormSet, FeeMasterCreateForm, BulkPaymentForm, StaffBankDetailForm, PaysheetRowForm, \
    SalaryAdvanceForm, SalaryStructureForm, StudentFundingForm, SchoolBankDetailForm, \
    StaffLoanForm, StaffLoanRepaymentForm, StaffFundingForm, DiscountForm, DiscountApplicationForm, \
    StudentDiscountAssignForm, OtherPaymentClearanceForm, OtherPaymentCreateForm, SalarySettingForm, BonusForm, \
    BonusFilterForm
from finance.tasks import generate_invoices_task
from pytz import timezone as pytz_timezone

from .utility import *


# ===================================================================
# Finance Settings Views (Singleton Pattern)
# ===================================================================


class FinanceSettingDetailView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """
    Displays the current finance settings. If no settings exist,
    it redirects the user to the create page.
    """
    permission_required = 'finance.view_financesettingmodel'
    template_name = 'finance/setting/detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['finance_setting'] = FinanceSettingModel.objects.first()
        return context

    def dispatch(self, request, *args, **kwargs):
        # If no settings object exists, redirect to the create view
        if not FinanceSettingModel.objects.exists():
            return redirect(reverse('finance_setting_create'))
        return super().dispatch(request, *args, **kwargs)


class FinanceSettingCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    """
    Handles the initial creation of the finance settings.
    This view will only be accessible if no settings object exists.
    """
    model = FinanceSettingModel
    permission_required = 'finance.change_financesettingmodel'
    form_class = FinanceSettingForm
    template_name = 'finance/setting/create.html'
    success_message = 'Finance Settings Created Successfully'
    success_url = reverse_lazy('finance_setting_detail')

    def dispatch(self, request, *args, **kwargs):
        # If a settings object already exists, redirect to the edit view
        if FinanceSettingModel.objects.exists():
            return redirect(reverse('finance_setting_update'))
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        # Set the 'updated_by' field to the current user upon creation
        form.instance.updated_by = self.request.user
        return super().form_valid(form)


class FinanceSettingUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    """
    Handles updating the existing finance settings object.
    """
    model = FinanceSettingModel
    permission_required = 'finance.change_financesettingmodel'
    form_class = FinanceSettingForm
    template_name = 'finance/setting/create.html'
    success_message = 'Finance Settings Updated Successfully'
    success_url = reverse_lazy('finance_setting_detail')

    def get_object(self, queryset=None):
        # This view will always edit the first (and only) settings object
        return FinanceSettingModel.objects.first()

    def form_valid(self, form):
        # Update the 'updated_by' field to the current user upon update
        form.instance.updated_by = self.request.user
        return super().form_valid(form)


class SupplierAccountsListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    PRIMARY VIEW ("Accounts Payable"): Lists all submitted Purchase Orders
    that require payment, showing their financial status.
    """
    model = PurchaseOrderModel
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/supplier_payment/accounts_payable_list.html'
    context_object_name = 'purchase_orders'
    paginate_by = 20

    def get_queryset(self):
        # We start with submitted POs and prefetch supplier data for efficiency
        queryset = super().get_queryset().filter(
            status__in=['submitted', 'partially_received', 'received']
        ).select_related('supplier', 'session', 'term')

        session_id = self.request.GET.get('session')
        term_id = self.request.GET.get('term')
        query = self.request.GET.get('q')

        if session_id:
            queryset = queryset.filter(session_id=session_id)
        if term_id:
            queryset = queryset.filter(term_id=term_id)
        if query:
            queryset = queryset.filter(
                Q(supplier__name__icontains=query) | Q(order_number__icontains=query)
            ).distinct()

        return queryset.order_by('-order_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sessions'] = SessionModel.objects.all().order_by('-start_year')
        context['terms'] = TermModel.objects.all().order_by('order')
        context['search_query'] = self.request.GET.get('q', '')
        # ... (add selected session/term logic here for the template filters) ...
        return context


class SupplierAccountDetailView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    DETAIL VIEW: Manages payments for a single Purchase Order.
    This view displays the PO's financial details, lists existing payments,
    and includes an inline form to add new payments.
    """
    model = SupplierPaymentModel
    form_class = SupplierPaymentForm
    permission_required = 'finance.add_expensemodel'
    template_name = 'finance/supplier_payment/po_payment_detail.html'

    def get_form_kwargs(self):
        """
        This method is overridden to pass the purchase_order instance
        to the form, so it knows the maximum payable amount.
        """
        kwargs = super().get_form_kwargs()
        self.purchase_order = get_object_or_404(PurchaseOrderModel, pk=self.kwargs['po_pk'])
        kwargs['purchase_order'] = self.purchase_order
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        self.purchase_order = get_object_or_404(PurchaseOrderModel, pk=self.kwargs['po_pk'])
        context['po'] = self.purchase_order
        # Get all payments related to this PO, including reverted ones for history
        context['payments'] = self.purchase_order.supplierpaymentmodel_set.all().order_by('-payment_date')
        return context

    def form_valid(self, form):
        self.purchase_order = get_object_or_404(PurchaseOrderModel, pk=self.kwargs['po_pk'])
        payment = form.save(commit=False)
        payment.supplier = self.purchase_order.supplier
        payment.created_by = self.request.user
        payment.save()
        # M2M fields must be saved after the initial save
        payment.purchase_orders.add(self.purchase_order)
        messages.success(self.request, "Payment recorded successfully against the Purchase Order.")
        return redirect('finance_po_payment_detail', po_pk=self.purchase_order.pk)


class SupplierPaymentRevertView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Handles reverting a 'Completed' payment to 'Reverted' status."""
    permission_required = 'finance.add_expensemodel'

    def post(self, request, *args, **kwargs):
        payment = get_object_or_404(SupplierPaymentModel, pk=self.kwargs['pk'])
        # Find the PO this payment is linked to, to redirect back correctly
        po_pk = payment.purchase_orders.first().pk if payment.purchase_orders.exists() else None

        if payment.status == 'completed':
            payment.status = 'reverted'
            payment.save()
            messages.warning(request, f"Payment {payment.receipt_number} has been successfully reverted.")
        else:
            messages.error(request, "Only 'Completed' payments can be reverted.")

        if po_pk:
            return redirect('finance_po_payment_detail', po_pk=po_pk)
        return redirect('finance_all_payments_list')  # Fallback redirect


class AllSupplierPaymentsListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    SECONDARY VIEW: Provides a comprehensive, filterable log of all
    individual supplier payment transactions for auditing purposes.
    """
    model = SupplierPaymentModel
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/supplier_payment/all_payments_list.html'
    context_object_name = 'payments'
    paginate_by = 25

    def get_queryset(self):
        queryset = super().get_queryset().select_related('supplier', 'session', 'term')

        # Get filter parameters from the request
        session_id = self.request.GET.get('session')
        term_id = self.request.GET.get('term')
        query = self.request.GET.get('q')

        # Apply filters if provided
        if session_id:
            queryset = queryset.filter(session_id=session_id)
        if term_id:
            queryset = queryset.filter(term_id=term_id)
        if query:
            queryset = queryset.filter(
                Q(supplier__name__icontains=query) |
                Q(purchase_orders__order_number__icontains=query) |
                Q(receipt_number__icontains=query) |
                Q(reference__icontains=query)
            ).distinct()  # Use distinct to avoid duplicates from M2M join

        return queryset.order_by('-payment_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        school_setting = SchoolSettingModel.objects.first()

        # Add filter options and selections to the context
        context['sessions'] = SessionModel.objects.all().order_by('-start_year')
        context['terms'] = TermModel.objects.all().order_by('order')
        context['search_query'] = self.request.GET.get('q', '')

        selected_session_id = self.request.GET.get('session')
        if selected_session_id:
            context['selected_session'] = get_object_or_404(SessionModel, pk=selected_session_id)
        elif school_setting:
            context['selected_session'] = school_setting.session

        selected_term_id = self.request.GET.get('term')
        if selected_term_id:
            context['selected_term'] = get_object_or_404(TermModel, pk=selected_term_id)
        elif school_setting:
            context['selected_term'] = school_setting.term

        return context


class SupplierPaymentReceiptView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """
    Displays a printable receipt for a single supplier payment.
    """
    model = SupplierPaymentModel
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/supplier_payment/receipt.html'
    context_object_name = 'payment'


class PurchaseAdvanceAccountsListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Lists approved advances that need payments"""
    model = PurchaseAdvanceModel
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/purchase_advance/accounts_list.html'
    context_object_name = 'purchase_advances'
    paginate_by = 20

    def get_queryset(self):
        from inventory.models import PurchaseAdvanceModel
        queryset = PurchaseAdvanceModel.objects.filter(
            status__in=['approved', 'disbursed']
        ).select_related('staff', 'session', 'term')

        session_id = self.request.GET.get('session')
        term_id = self.request.GET.get('term')
        query = self.request.GET.get('q')

        if session_id:
            queryset = queryset.filter(session_id=session_id)
        if term_id:
            queryset = queryset.filter(term_id=term_id)
        if query:
            queryset = queryset.filter(
                Q(staff__first_name__icontains=query) | Q(staff__last_name__icontains=query) |
                Q(advance_number__icontains=query)
            ).distinct()

        return queryset.order_by('-request_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sessions'] = SessionModel.objects.all().order_by('-start_year')
        context['terms'] = TermModel.objects.all().order_by('order')
        context['search_query'] = self.request.GET.get('q', '')
        return context


class PurchaseAdvancePaymentDetailView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Manages payments for a single Purchase Advance"""
    model = PurchaseAdvancePaymentModel
    form_class = PurchaseAdvancePaymentForm
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/purchase_advance/payment_detail.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        from inventory.models import PurchaseAdvanceModel
        self.advance = get_object_or_404(PurchaseAdvanceModel, pk=self.kwargs['advance_pk'])
        kwargs['advance'] = self.advance
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from inventory.models import PurchaseAdvanceModel
        self.advance = get_object_or_404(PurchaseAdvanceModel, pk=self.kwargs['advance_pk'])
        context['advance'] = self.advance
        context['payments'] = self.advance.payments.all().order_by('-payment_date')
        return context

    def form_valid(self, form):
        from inventory.models import PurchaseAdvanceModel
        self.advance = get_object_or_404(PurchaseAdvanceModel, pk=self.kwargs['advance_pk'])
        payment = form.save(commit=False)
        payment.advance = self.advance
        payment.created_by = self.request.user
        payment.save()
        messages.success(self.request, "Payment recorded successfully against the Purchase Advance.")
        return redirect('finance_advance_payment_detail', advance_pk=self.advance.pk)


# ===================================================================
# Fee Type Views (Modal Interface)
# ===================================================================
class FeeListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = FeeModel
    permission_required = 'finance.view_feemodel'
    template_name = 'finance/fee/index.html'
    context_object_name = 'fees'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'form' not in context:
            context['form'] = FeeForm()
        return context


class FeeCreateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView):
    model = FeeModel
    permission_required = 'finance.add_feemodel'
    form_class = FeeForm

    def get_success_url(self):
        return reverse('finance_fee_list')

    def form_valid(self, form):
        messages.success(self.request, "Fee Type created successfully.")
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET': return redirect(self.success_url)
        return super().dispatch(request, *args, **kwargs)


class FeeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView):
    model = FeeModel
    permission_required = 'finance.add_feemodel'
    form_class = FeeForm

    def get_success_url(self):
        return reverse('finance_fee_list')

    def form_valid(self, form):
        messages.success(self.request, "Fee Type updated successfully.")
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET': return redirect(self.success_url)
        return super().dispatch(request, *args, **kwargs)


class FeeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = FeeModel
    permission_required = 'finance.add_feemodel'
    template_name = 'finance/fee/delete.html'
    success_url = reverse_lazy('finance_fee_list')

    def form_valid(self, form):
        messages.success(self.request, f"Fee Type '{self.object.name}' deleted successfully.")
        return super().form_valid(form)


# ===================================================================
# Fee Group Views (Modal Interface)
# ===================================================================
class FeeGroupListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = FeeGroupModel
    permission_required = 'finance.view_feemodel'
    template_name = 'finance/fee_group/index.html'
    context_object_name = 'fee_groups'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'form' not in context:
            context['form'] = FeeGroupForm()
        return context


class FeeGroupCreateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView):
    model = FeeGroupModel
    permission_required = 'finance.add_feemodel'
    form_class = FeeGroupForm

    def get_success_url(self):
        return reverse('finance_fee_group_list')

    def form_valid(self, form):
        messages.success(self.request, "Fee Group created successfully.")
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET': return redirect(self.success_url)
        return super().dispatch(request, *args, **kwargs)


class FeeGroupUpdateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView):
    model = FeeGroupModel
    permission_required = 'finance.add_feemodel'
    form_class = FeeGroupForm

    def get_success_url(self):
        return reverse('finance_fee_group_list')

    def form_valid(self, form):
        messages.success(self.request, "Fee Group updated successfully.")
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET': return redirect(self.success_url)
        return super().dispatch(request, *args, **kwargs)


class FeeGroupDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = FeeGroupModel
    permission_required = 'finance.add_feemodel'
    template_name = 'finance/fee_group/delete.html'
    success_url = reverse_lazy('finance_fee_group_list')
    context_object_name = 'fee_group'

    def form_valid(self, form):
        messages.success(self.request, f"Fee Group '{self.object.name}' deleted successfully.")
        return super().form_valid(form)


# ===================================================================
# Fee Master (Structure)
# ===================================================================
class FeeMasterListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Displays a list of all created fee structures."""
    model = FeeMasterModel
    permission_required = 'finance.view_feemodel'
    template_name = 'finance/fee_master/index.html'
    context_object_name = 'fee_structures'
    paginate_by = 15


class FeeMasterCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Handles the creation of the main FeeMasterModel header (Step 1)."""
    model = FeeMasterModel
    permission_required = 'finance.add_feemodel'
    form_class = FeeMasterCreateForm
    template_name = 'finance/fee_master/create.html'

    def form_valid(self, form):
        messages.success(self.request, "Fee structure created. Now, set the price for each term.")
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        # Redirect to the detail page to set the termly prices
        return reverse('finance_fee_master_detail', kwargs={'pk': self.object.pk})


# Updated view
class FeeMasterDetailView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.view_feemodel'
    template_name = 'finance/fee_master/detail.html'

    def get(self, request, *args, **kwargs):
        fee_structure = get_object_or_404(FeeMasterModel, pk=self.kwargs.get('pk'))

        # Determine which terms to show
        if fee_structure.fee.occurrence == FeeModel.FeeOccurrence.TERMLY:
            terms = TermModel.objects.all().order_by('order')
        else:
            if fee_structure.fee.payment_term:
                terms = [fee_structure.fee.payment_term]
            else:
                terms = []

        # Get existing amounts
        term_amounts = {}
        for amount in fee_structure.termly_amounts.filter(term__in=terms):
            term_amounts[amount.term.id] = amount.amount

        # Prepare display data
        display_terms = []
        for term in terms:
            display_terms.append({
                'term': term,
                'amount': term_amounts.get(term.id, Decimal('0.00'))
            })

        context = {
            'fee_structure': fee_structure,
            'display_terms': display_terms,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        fee_structure = get_object_or_404(FeeMasterModel, pk=self.kwargs.get('pk'))

        # Determine which terms to process
        if fee_structure.fee.occurrence == FeeModel.FeeOccurrence.TERMLY:
            terms = TermModel.objects.all().order_by('order')
        else:
            if fee_structure.fee.payment_term:
                terms = [fee_structure.fee.payment_term]
            else:
                terms = []

        # Save amounts
        for term in terms:
            field_name = f'term_{term.id}_amount'
            if field_name in request.POST:
                amount = Decimal(request.POST[field_name] or '0')
                TermlyFeeAmountModel.objects.update_or_create(
                    fee_structure=fee_structure,
                    term=term,
                    defaults={'amount': amount}
                )

        messages.success(request, "Fee amounts saved successfully!")
        return redirect('finance_fee_master_detail', pk=fee_structure.pk)


class FeeMasterUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Handles updating the core details of a FeeMasterModel, such as the
    group, fee, and class assignments.
    """
    model = FeeMasterModel
    permission_required = 'finance.add_feemodel'
    form_class = FeeMasterCreateForm  # We can reuse the create form for updating
    template_name = 'finance/fee_master/update.html'
    context_object_name = 'fee_structure'

    def form_valid(self, form):
        messages.success(self.request, "Fee structure details updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        # Redirect back to the detail page after updating
        return reverse('finance_fee_master_detail', kwargs={'pk': self.object.pk})


class FeeMasterDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Handles the deletion of a FeeMasterModel."""
    model = FeeMasterModel
    permission_required = 'finance.add_feemodel'
    template_name = 'finance/fee_master/delete.html'
    success_url = reverse_lazy('finance_fee_master_list')

    def form_valid(self, form):
        messages.success(self.request, f"Fee structure '{self.object}' deleted successfully.")
        return super().form_valid(form)


# ===================================================================
# Invoice Generation Views (Asynchronous Workflow)
# ===================================================================
class InvoiceGenerationView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    form_class = InvoiceGenerationForm
    template_name = 'finance/invoice/generate.html'
    permission_required = 'finance.add_feemodel'

    def form_valid(self, form):
        job = form.save(commit=False)
        job.created_by = self.request.user
        job.save()
        form.save_m2m()  # Important for ManyToManyField ('classes_to_invoice')

        generate_invoices_task.delay(str(job.job_id))
        messages.info(self.request,
                      "Invoice generation has been started in the background. You will be redirected to the status page.")
        return redirect('finance_invoice_job_status', pk=job.job_id)


class InvoiceJobStatusView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = InvoiceGenerationJob
    template_name = 'finance/invoice/status.html'
    context_object_name = 'job'
    permission_required = 'finance.view_feemodel'


def invoice_job_status_api(request, pk):
    job = get_object_or_404(InvoiceGenerationJob, pk=pk)
    return JsonResponse({
        'status': job.get_status_display(),
        'total_students': job.total_students,
        'processed_students': job.processed_students,
        'error_message': job.error_message,
        'is_complete': job.status in ['success', 'failure']
    })


# ===================================================================
# Invoice and Payment Management Views
# ===================================================================

class InvoiceListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = InvoiceModel
    permission_required = 'finance.view_feemodel'
    template_name = 'finance/invoice/index.html'
    context_object_name = 'invoices'
    paginate_by = 30

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'student', 'student__student_class', 'student__class_section', 'session', 'term'
        )

        search_query = self.request.GET.get('q', '').strip()
        session_id = self.request.GET.get('session')
        term_id = self.request.GET.get('term')
        status = self.request.GET.get('status', '')

        # Get school setting for defaults
        school_setting = SchoolSettingModel.objects.first()

        # Apply session filter - use current if not specified
        if session_id:
            queryset = queryset.filter(session_id=session_id)
        elif school_setting and school_setting.session:
            # THIS IS KEY: Actually filter, don't just set context
            queryset = queryset.filter(session=school_setting.session)

        # Apply term filter - use current if not specified
        if term_id:
            queryset = queryset.filter(term_id=term_id)
        elif school_setting and school_setting.term:
            # THIS IS KEY: Actually filter, don't just set context
            queryset = queryset.filter(term=school_setting.term)

        # Apply status filter
        if status:
            queryset = queryset.filter(status=status)

        # Apply search query
        if search_query:
            queryset = queryset.annotate(
                student_full_name=Concat(
                    'student__first_name', Value(' '), 'student__last_name'
                )
            ).filter(
                Q(student_full_name__icontains=search_query) |
                Q(student__registration_number__icontains=search_query) |
                Q(invoice_number__icontains=search_query)
            )

        # Order by parent, then student
        return queryset.order_by(
            'student__parent__last_name',
            'student__parent__first_name',
            'student__parent__id',
            'student__last_name',
            'student__first_name',
            '-issue_date'
        )

    def get_context_data(self, **kwargs):
        """
        This method adds the necessary data for the filter dropdowns and
        to remember the user's current selections.
        """
        context = super().get_context_data(**kwargs)
        school_setting = SchoolSettingModel.objects.first()

        # Data for the filter dropdowns
        context['sessions'] = SessionModel.objects.all().order_by('-start_year')
        context['terms'] = TermModel.objects.all().order_by('order')
        context['status_choices'] = InvoiceModel.Status.choices

        # Get the currently selected filter values to keep them selected in the form
        selected_session_id = self.request.GET.get('session')
        selected_term_id = self.request.GET.get('term')

        # Pass the selected objects or defaults to the template
        if selected_session_id:
            context['selected_session'] = SessionModel.objects.get(pk=selected_session_id)
        elif school_setting:
            context['selected_session'] = school_setting.session

        if selected_term_id:
            context['selected_term'] = TermModel.objects.get(pk=selected_term_id)
        elif school_setting:
            context['selected_term'] = school_setting.term

        context['selected_status'] = self.request.GET.get('status', '')
        context['search_query'] = self.request.GET.get('q', '')
        context['group_by_parent'] = True

        return context


class InvoiceDetailView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = FeePaymentModel
    form_class = FeePaymentForm
    permission_required = 'finance.view_feemodel'
    template_name = 'finance/invoice/detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        self.invoice = get_object_or_404(InvoiceModel, pk=self.kwargs['pk'])
        context['invoice'] = self.invoice
        context['payments'] = self.invoice.payments.all().order_by('-date')
        return context

    def form_valid(self, form):
        invoice = get_object_or_404(InvoiceModel, pk=self.kwargs['pk'])
        payment = form.save(commit=False)
        payment.invoice = invoice
        payment.status = FeePaymentModel.PaymentStatus.CONFIRMED  # Or based on your workflow
        payment.confirmed_by = self.request.user
        payment.save()

        # Update invoice status
        if invoice.balance <= 0:
            invoice.status = InvoiceModel.Status.PAID
        else:
            invoice.status = InvoiceModel.Status.PARTIALLY_PAID
        invoice.save()

        messages.success(self.request, "Payment recorded successfully.")
        return redirect('finance_invoice_detail', pk=invoice.pk)


class InvoiceItemDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = InvoiceItemModel
    permission_required = 'finance.delete_invoicemodel'
    template_name = 'finance/invoice/delete_item.html'

    def get_success_url(self):
        return reverse('finance_invoice_detail', kwargs={'pk': self.object.invoice.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        invoice_item = self.get_object()
        invoice = invoice_item.invoice

        # Check if there are any confirmed payments for this invoice
        has_confirmed_payments = invoice.payments.filter(status='confirmed').exists()

        context['invoice'] = invoice
        context['has_confirmed_payments'] = has_confirmed_payments
        return context

    def delete(self, request, *args, **kwargs):
        invoice_item = self.get_object()
        invoice = invoice_item.invoice

        # Check if there are any confirmed payments for this invoice
        has_confirmed_payments = invoice.payments.filter(status='confirmed').exists()

        if has_confirmed_payments:
            messages.error(request, "Cannot delete invoice item because the invoice has confirmed payments.")
            return redirect('finance_invoice_detail', pk=invoice.pk)

        messages.success(request, f"Invoice item '{invoice_item.description}' has been deleted successfully.")
        return super().delete(request, *args, **kwargs)


class InvoiceDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = InvoiceModel
    permission_required = 'finance.delete_invoicemodel'
    template_name = 'finance/invoice/delete.html'
    success_url = reverse_lazy('finance_invoice_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        invoice = self.get_object()

        # Check if there are any confirmed payments for this invoice
        has_confirmed_payments = invoice.payments.filter(status='confirmed').exists()

        # Get all items for display in the confirmation page
        items = invoice.items.all()
        total_amount = items.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        context['has_confirmed_payments'] = has_confirmed_payments
        context['items'] = items
        context['total_amount'] = total_amount
        return context

    def delete(self, request, *args, **kwargs):
        invoice = self.get_object()

        # Check if there are any confirmed payments for this invoice
        has_confirmed_payments = invoice.payments.filter(status='confirmed').exists()

        if has_confirmed_payments:
            messages.error(request, "Cannot delete invoice because it has confirmed payments.")
            return redirect('finance_invoice_list')

        messages.success(request, f"Invoice '{invoice.invoice_number}' has been deleted successfully.")
        return super().delete(request, *args, **kwargs)


class StudentFeeSearchView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'finance.add_feemodel'
    template_name = 'finance/payment/select_student.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get classes with sections as JSON-serializable data
        class_list = []
        for cls in ClassesModel.objects.prefetch_related('section').all().order_by('name'):
            class_list.append({
                'id': cls.id,
                'name': cls.name,
                'sections': [{'id': s.id, 'name': s.name} for s in cls.section.all()]
            })

        context['class_list'] = ClassesModel.objects.prefetch_related('section').all().order_by('name')
        context['class_list_json'] = json.dumps(class_list)

        # Pre-load students with related names
        all_students = StudentModel.objects.filter(status='active').select_related('student_class', 'class_section')

        # Create custom student data with class/section names
        student_data = []
        for student in all_students:
            student_data.append({
                'pk': student.id,
                'fields': {
                    'first_name': student.first_name,
                    'last_name': student.last_name,
                    'registration_number': student.registration_number,
                    'gender': student.gender,
                    'image': student.image.url if student.image else '',
                    'student_class_id': student.student_class.id if student.student_class else '',
                    'student_class_name': student.student_class.name if student.student_class else '',
                    'class_section_id': student.class_section.id if student.class_section else '',
                    'class_section_name': student.class_section.name if student.class_section else '',
                }
            })

        context['student_list_json'] = json.dumps(student_data)
        return context


def get_students_by_class_ajax(request):
    """AJAX endpoint to fetch students for a given class and section."""
    class_pk = request.GET.get('class_pk')
    section_pk = request.GET.get('section_pk')
    students = StudentModel.objects.filter(student_class_id=class_pk, class_section_id=section_pk, status='active')
    return render(request, 'finance/payment/partials/student_search_results.html', {'students': students})


def get_students_by_reg_no_ajax(request):
    """AJAX endpoint to fetch students by registration number."""
    reg_no = request.GET.get('reg_no', '').strip()
    students = StudentModel.objects.filter(registration_number__icontains=reg_no, status='active')
    return render(request, 'finance/payment/partials/student_search_results.html', {'students': students})


class StudentFinancialDashboardView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'finance.view_feemodel'
    template_name = 'finance/payment/student_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = get_object_or_404(StudentModel, pk=self.kwargs['pk'])
        context['student'] = student

        school_setting = SchoolSettingModel.objects.first()
        other_payments = OtherPaymentModel.objects.filter(
            student=student
        ).exclude(status='paid')

        context['other_payments'] = other_payments
        context['total_other_payment_balance'] = sum(op.balance for op in other_payments)

        # --- NEW LOGIC: Load a specific invoice or the current one ---
        invoice_id = self.request.GET.get('invoice_id')
        if invoice_id:
            # Load a specific invoice from the history
            current_invoice = get_object_or_404(student.invoices, pk=invoice_id)
        else:
            # Default to the invoice for the current term
            current_invoice = student.invoices.filter(
                session=school_setting.session, term=school_setting.term
            ).first()

        context['current_invoice'] = current_invoice
        context['invoice_history'] = student.invoices.order_by('-session__start_year', '-term__order')
        context['all_payments'] = FeePaymentModel.objects.filter(invoice__student=student).order_by('-date')
        # Add discount information
        if current_invoice:
            context['invoice_discounts'] = StudentDiscountModel.objects.filter(
                invoice_item__invoice=current_invoice
            ).select_related('discount_application__discount')

        # Pass the form for payment details, bound to the current invoice
        if current_invoice:
            context['payment_form'] = FeePaymentForm()
        else:
            context['payment_form'] = FeePaymentForm()

        return context

    def post(self, request, *args, **kwargs):
        student = get_object_or_404(StudentModel, pk=self.kwargs.get('pk'))
        invoice_id = request.POST.get('invoice_id')
        invoice = get_object_or_404(InvoiceModel, pk=invoice_id)

        # 1. Instantiate the form that holds the payment details (mode, date, bank)
        payment_form = FeePaymentForm(request.POST)

        # 2. Loop through item payments to calculate total and prep data
        total_paid_in_transaction = Decimal('0.00')
        item_payment_data = {}  # Stores {item_instance: amount_to_pay}

        for key, value in request.POST.items():
            if key.startswith('item_') and value:
                try:
                    item_id = int(key.split('_')[1])
                    amount_for_item = Decimal(value)

                    if amount_for_item > 0:
                        item = get_object_or_404(InvoiceItemModel, pk=item_id, invoice=invoice)

                        # Don't allow overpayment on a single item
                        payable_amount = min(amount_for_item, item.balance)

                        # Store this for the atomic transaction
                        item_payment_data[item] = payable_amount
                        total_paid_in_transaction += payable_amount

                except (ValueError, TypeError, InvoiceItemModel.DoesNotExist):
                    # Malicious or bad data, just skip it
                    continue

        # 3. Check if any payment was made AND if the payment details are valid
        if total_paid_in_transaction <= 0:
            messages.warning(request, "No payment amount was entered.")
            return redirect('finance_student_dashboard', pk=student.pk)

        if payment_form.is_valid():
            payment_mode = payment_form.cleaned_data['payment_mode']

            # Validate wallet if payment mode is wallet
            if payment_mode == 'wallet':
                wallet = getattr(student, 'student_wallet', None)
                if not wallet or wallet.fee_balance < total_paid_in_transaction:
                    messages.error(request,
                                   f"Insufficient fee wallet balance. Available: ₦{wallet.fee_balance if wallet else 0:,.2f}")
                    return redirect('finance_student_dashboard', pk=student.pk)

            try:
                with transaction.atomic():
                    # First, apply the payments to the individual items
                    for item, amount in item_payment_data.items():
                        item.amount_paid += amount
                        item.save(update_fields=['amount_paid'])

                        # Handle parent-bound fees
                        if item.fee_master.fee.parent_bound:
                            siblings = student.parent.wards.exclude(pk=student.pk)
                            for sibling in siblings:
                                try:
                                    sibling_invoice = InvoiceModel.objects.get(
                                        student=sibling,
                                        session=invoice.session,
                                        term=invoice.term
                                    )
                                    sibling_item = InvoiceItemModel.objects.get(
                                        invoice=sibling_invoice,
                                        fee_master=item.fee_master
                                    )
                                    sibling_item.paid_by_sibling = student
                                    sibling_item.amount_paid = sibling_item.amount
                                    sibling_item.save(update_fields=['paid_by_sibling', 'amount_paid'])
                                except (InvoiceModel.DoesNotExist, InvoiceItemModel.DoesNotExist):
                                    continue

                    # Build item_breakdown for JSON storage
                    # Format: {"item_id": "amount", "item_id": "amount"}
                    item_breakdown = {
                        str(item.pk): str(amount)
                        for item, amount in item_payment_data.items()
                    }

                    # Now, create the single FeePaymentModel with item_breakdown
                    FeePaymentModel.objects.create(
                        invoice=invoice,
                        amount=total_paid_in_transaction,
                        payment_mode=payment_form.cleaned_data['payment_mode'],
                        description=payment_form.cleaned_data['description'],
                        currency=payment_form.cleaned_data['currency'],
                        date=payment_form.cleaned_data['date'],
                        bank_account=payment_form.cleaned_data['bank_account'],
                        reference=payment_form.cleaned_data['reference'],
                        notes=payment_form.cleaned_data['notes'],
                        status=FeePaymentModel.PaymentStatus.CONFIRMED,
                        confirmed_by=request.user,
                        item_breakdown=item_breakdown  # NEW: Save the breakdown
                    )

                    if payment_mode == 'wallet':
                        wallet = StudentWalletModel.objects.select_for_update().get(student=student)
                        wallet.fee_balance -= total_paid_in_transaction
                        wallet.save(update_fields=['fee_balance'])

                    # Finally, update the parent invoice's status
                    invoice.refresh_from_db()
                    if invoice.balance <= Decimal('0.01'):
                        invoice.status = InvoiceModel.Status.PAID
                    else:
                        invoice.status = InvoiceModel.Status.PARTIALLY_PAID
                    invoice.save(update_fields=['status'])

                    messages.success(request,
                                     f"Payment of ₦{total_paid_in_transaction:,.2f} was applied successfully.")

            except Exception as e:
                messages.error(request, f"An error occurred while saving the payment: {e}")

        else:
            messages.error(request, "Payment failed: The payment details (mode, date, or bank) were invalid.")
            for field, errors in payment_form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.title()}: {error}")

        return redirect('finance_student_dashboard', pk=student.pk)


class InvoiceReceiptView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = InvoiceModel
    permission_required = 'finance.view_feemodel'
    template_name = 'finance/payment/invoice_receipt.html'
    context_object_name = 'invoice'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # You can add school settings to the context if needed for the header
        # context['school_setting'] = SchoolSettingModel.objects.first()
        return context


class FeePaymentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = FeePaymentModel
    permission_required = 'finance.view_feemodel'  # Assumes default permission
    template_name = 'finance/payment/payment_index.html'
    context_object_name = 'payment_list'
    paginate_by = 25

    def get_queryset(self):
        # Start with a base queryset, pre-fetching related data for efficiency
        queryset = FeePaymentModel.objects.select_related(
            'invoice__student',
            'invoice__session',
            'invoice__term'
        ).order_by('-date', '-created_at')

        # Get filter parameters from the URL
        session_id = self.request.GET.get('session', '')
        term_id = self.request.GET.get('term', '')
        search_query = self.request.GET.get('search', '').strip()

        # Apply filters if they exist
        if session_id:
            queryset = queryset.filter(invoice__session_id=session_id)

        if term_id:
            queryset = queryset.filter(invoice__term_id=term_id)

        # Apply search query if it exists
        if search_query:
            # Annotate the student's full name to make it searchable
            queryset = queryset.annotate(
                student_full_name=Concat(
                    'invoice__student__first_name', Value(' '), 'invoice__student__last_name'
                )
            ).filter(
                Q(student_full_name__icontains=search_query) |
                Q(invoice__student__registration_number__icontains=search_query) |
                Q(invoice__invoice_number__icontains=search_query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Pass data for the filter dropdowns
        context['session_list'] = SessionModel.objects.all().order_by('-start_year')
        context['term_list'] = TermModel.objects.all().order_by('order')

        # Pass current filter values back to the template to maintain state
        context['current_session_id'] = self.request.GET.get('session', '')
        context['current_term_id'] = self.request.GET.get('term', '')
        context['search_query'] = self.request.GET.get('search', '')

        return context


class FeePendingPaymentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = FeePaymentModel
    permission_required = 'finance.view_feemodel'  # Assumes default permission
    template_name = 'finance/payment/pending_payment_index.html'
    context_object_name = 'payment_list'
    paginate_by = 25

    def get_queryset(self):
        # Start with a base queryset, pre-fetching related data for efficiency
        queryset = FeePaymentModel.objects.filter(status='pending').select_related(
            'invoice__student',
            'invoice__session',
            'invoice__term',
        ).order_by('-date', '-created_at')

        # Get filter parameters from the URL
        session_id = self.request.GET.get('session', '')
        term_id = self.request.GET.get('term', '')
        search_query = self.request.GET.get('search', '').strip()

        # Apply filters if they exist
        if session_id:
            queryset = queryset.filter(invoice__session_id=session_id)

        if term_id:
            queryset = queryset.filter(invoice__term_id=term_id)

        # Apply search query if it exists
        if search_query:
            # Annotate the student's full name to make it searchable
            queryset = queryset.annotate(
                student_full_name=Concat(
                    'invoice__student__first_name', Value(' '), 'invoice__student__last_name'
                )
            ).filter(
                Q(student_full_name__icontains=search_query) |
                Q(invoice__student__registration_number__icontains=search_query) |
                Q(invoice__invoice_number__icontains=search_query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Pass data for the filter dropdowns
        context['session_list'] = SessionModel.objects.all().order_by('-start_year')
        context['term_list'] = TermModel.objects.all().order_by('order')

        # Pass current filter values back to the template to maintain state
        context['current_session_id'] = self.request.GET.get('session', '')
        context['current_term_id'] = self.request.GET.get('term', '')
        context['search_query'] = self.request.GET.get('search', '')

        return context


class BulkFeePaymentView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """
    Handles a single "bulk" payment that is intelligently allocated
    across multiple outstanding invoices for a student, oldest first.
    """
    form_class = BulkPaymentForm
    permission_required = 'finance.add_feemodel'
    template_name = 'finance/payment/bulk_payment_form.html'

    def get_form_kwargs(self):
        """Pass student to form for validation"""
        kwargs = super().get_form_kwargs()
        kwargs['student'] = get_object_or_404(StudentModel, pk=self.kwargs['pk'])
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        self.student = get_object_or_404(StudentModel, pk=self.kwargs['pk'])
        context['student'] = self.student
        context['outstanding_invoices'] = self.student.invoices.exclude(status=InvoiceModel.Status.PAID).order_by(
            'issue_date')

        # Add total balance to context
        context['total_balance'] = sum(invoice.balance for invoice in context['outstanding_invoices'])

        return context

    def form_valid(self, form):
        student = get_object_or_404(StudentModel, pk=self.kwargs['pk'])
        total_amount_paid = form.cleaned_data['amount']
        payment_mode = form.cleaned_data['payment_mode']
        amount_to_allocate = total_amount_paid

        # Get all unpaid or partially paid invoices, oldest first
        outstanding_invoices = student.invoices.exclude(status=InvoiceModel.Status.PAID).order_by('issue_date')
        total_outstanding = sum(invoice.balance for invoice in outstanding_invoices)

        with transaction.atomic():
            # Handle wallet payment - validate and lock wallet
            if payment_mode == 'wallet':
                wallet = StudentWalletModel.objects.select_for_update().get(student=student)
                if wallet.fee_balance < total_amount_paid:
                    messages.error(
                        self.request,
                        f"Insufficient fee wallet balance. Available: ₦{wallet.fee_balance:,.2f}, Required: ₦{total_amount_paid:,.2f}"
                    )
                    return redirect('finance_student_dashboard', pk=student.pk)

            # Track amounts for messaging
            total_applied_to_fees = Decimal('0.00')
            invoices_paid_count = 0

            # Allocate to invoices
            for invoice in outstanding_invoices:
                if amount_to_allocate <= 0:
                    break

                payment_for_this_invoice = min(invoice.balance, amount_to_allocate)

                if payment_for_this_invoice > 0:
                    # Track how much is allocated to each item in this invoice
                    item_breakdown = {}
                    remaining_for_invoice = payment_for_this_invoice

                    items_with_balance = [item for item in invoice.items.order_by('id') if item.balance > 0]

                    # Distribute payment across items in this invoice
                    for item in items_with_balance:
                        if remaining_for_invoice <= 0:
                            break

                        payable = min(item.balance, remaining_for_invoice)
                        item.amount_paid += payable
                        item.save(update_fields=['amount_paid'])

                        # Track this allocation
                        item_breakdown[str(item.pk)] = str(payable)
                        remaining_for_invoice -= payable

                        # Handle parent-bound fees
                        if item.fee_master.fee.parent_bound and item.amount_paid >= item.amount_after_discount:
                            siblings = student.parent.wards.exclude(pk=student.pk)
                            for sibling in siblings:
                                try:
                                    sibling_invoice = InvoiceModel.objects.get(
                                        student=sibling,
                                        session=invoice.session,
                                        term=invoice.term
                                    )
                                    sibling_item = InvoiceItemModel.objects.get(
                                        invoice=sibling_invoice,
                                        fee_master=item.fee_master
                                    )
                                    sibling_item.paid_by_sibling = student
                                    sibling_item.amount_paid = sibling_item.amount_after_discount
                                    sibling_item.save(update_fields=['paid_by_sibling', 'amount_paid'])
                                except (InvoiceModel.DoesNotExist, InvoiceItemModel.DoesNotExist):
                                    continue

                    # Create payment record with item breakdown
                    FeePaymentModel.objects.create(
                        invoice=invoice,
                        amount=payment_for_this_invoice,
                        payment_mode=payment_mode,
                        date=form.cleaned_data['date'],
                        description=form.cleaned_data.get('description', ''),
                        currency=form.cleaned_data['currency'],
                        bank_account=form.cleaned_data.get('bank_account'),
                        reference=form.cleaned_data.get('reference') or f"bulk-pmt-{invoice.invoice_number}",
                        status=FeePaymentModel.PaymentStatus.CONFIRMED,
                        confirmed_by=self.request.user,
                        item_breakdown=item_breakdown
                    )

                    # Refresh invoice from DB before checking balance
                    invoice.refresh_from_db()
                    if invoice.balance <= Decimal('0.01'):
                        invoice.status = InvoiceModel.Status.PAID
                        invoices_paid_count += 1
                    else:
                        invoice.status = InvoiceModel.Status.PARTIALLY_PAID
                    invoice.save()

                    total_applied_to_fees += payment_for_this_invoice
                    amount_to_allocate -= payment_for_this_invoice

            # Deduct from wallet if payment mode is wallet
            if payment_mode == 'wallet':
                wallet.fee_balance -= total_amount_paid
                wallet.save(update_fields=['fee_balance'])

            # Handle excess payment - fund fee wallet
            if amount_to_allocate > 0:
                wallet_to_fund, _ = StudentWalletModel.objects.get_or_create(student=student)
                wallet_to_fund.fee_balance += amount_to_allocate  # Fund fee wallet, not canteen
                wallet_to_fund.save(update_fields=['fee_balance'])

                messages.success(
                    self.request,
                    f"Payment processed successfully! ₦{total_applied_to_fees:,.2f} applied to {invoices_paid_count} invoice(s). "
                    f"Excess amount of ₦{amount_to_allocate:,.2f} has been credited to the student's fee wallet for future use."
                )
            else:
                messages.success(
                    self.request,
                    f"Bulk payment of ₦{total_amount_paid:,.2f} allocated successfully across invoice(s)."
                )

        return redirect('finance_student_dashboard', pk=student.pk)

@login_required
@permission_required('finance.change_feepaymentmodel', raise_exception=True)
def confirm_fee_payment_view(request, payment_id):
    """
    Confirms a pending fee payment uploaded by parent.
    Supports both quick payment (auto-distribute) and itemized payment (parent's allocation).
    Staff can override parent's allocation if needed.
    """
    payment = get_object_or_404(FeePaymentModel, pk=payment_id)

    if payment.status != FeePaymentModel.PaymentStatus.PENDING:
        messages.warning(request, "This payment has already been processed.")
        return redirect('pending_fee_payment_list')

    # Check if this is a confirmation with override
    override_allocation = request.POST.get('override_allocation') == 'true'

    with transaction.atomic():
        # Track the actual allocations made
        item_breakdown = {}

        # Confirm the payment
        payment.status = FeePaymentModel.PaymentStatus.CONFIRMED
        payment.confirmed_by = request.user
        payment.save(update_fields=['status', 'confirmed_by'])

        invoice = payment.invoice
        amount_to_allocate = payment.amount

        # Check if parent specified item allocations
        import json
        import re
        parent_allocations = {}

        if payment.notes:
            # Try to extract JSON allocation from notes
            match = re.search(r'Item Allocations:\s*(\{[^}]+\}|\[[^\]]+\])', payment.notes, re.DOTALL)
            if match:
                try:
                    allocations_str = match.group(1)
                    # Handle potential multi-line JSON
                    full_json_match = re.search(r'Item Allocations:\s*(\{.*?\n.*?\})', payment.notes, re.DOTALL)
                    if full_json_match:
                        allocations_str = full_json_match.group(1)
                    parent_allocations = json.loads(allocations_str)
                except json.JSONDecodeError:
                    pass

        # Determine allocation strategy
        if parent_allocations and not override_allocation:
            # Use parent's specified allocations
            for item_id_str, allocation_data in parent_allocations.items():
                try:
                    item_id = int(item_id_str)
                    allocated_amount = Decimal(str(allocation_data['amount']))

                    item = InvoiceItemModel.objects.get(pk=item_id, invoice=invoice)

                    # Apply payment to this specific item
                    payable = min(item.balance, allocated_amount)
                    item.amount_paid += payable
                    item.save(update_fields=['amount_paid'])
                    amount_to_allocate -= payable

                    # Track this allocation
                    item_breakdown[str(item.pk)] = str(payable)

                    # Handle parent-bound fees
                    if item.fee_master.fee.parent_bound and item.amount_paid >= item.amount_after_discount:
                        student = invoice.student
                        if student.parent:
                            siblings = student.parent.wards.exclude(pk=student.pk)
                            for sibling in siblings:
                                try:
                                    sibling_invoice = InvoiceModel.objects.get(
                                        student=sibling,
                                        session=invoice.session,
                                        term=invoice.term
                                    )
                                    sibling_item = InvoiceItemModel.objects.get(
                                        invoice=sibling_invoice,
                                        fee_master=item.fee_master
                                    )
                                    sibling_item.paid_by_sibling = student
                                    sibling_item.amount_paid = sibling_item.amount_after_discount
                                    sibling_item.save(update_fields=['paid_by_sibling', 'amount_paid'])
                                except (InvoiceModel.DoesNotExist, InvoiceItemModel.DoesNotExist):
                                    continue

                except (ValueError, InvoiceItemModel.DoesNotExist):
                    continue

            # If there's leftover amount (due to balance changes), distribute it
            if amount_to_allocate > Decimal('0.01'):
                for item in invoice.items.filter(amount_paid__lt=F('amount_after_discount')).order_by('id'):
                    if amount_to_allocate <= 0:
                        break
                    payable = min(item.balance, amount_to_allocate)
                    item.amount_paid += payable
                    item.save(update_fields=['amount_paid'])
                    amount_to_allocate -= payable

                    # Track this allocation (add to existing or create new)
                    current = Decimal(item_breakdown.get(str(item.pk), '0'))
                    item_breakdown[str(item.pk)] = str(current + payable)

        else:
            # Auto-distribute payment across items (original behavior or staff override)
            for item in invoice.items.filter(amount_paid__lt=F('amount_after_discount')).order_by('id'):
                if amount_to_allocate <= 0:
                    break

                payable = min(item.balance, amount_to_allocate)
                item.amount_paid += payable
                item.save(update_fields=['amount_paid'])
                amount_to_allocate -= payable

                # Track this allocation
                item_breakdown[str(item.pk)] = str(payable)

                # Handle parent-bound fees
                if item.fee_master.fee.parent_bound and item.amount_paid >= item.amount_after_discount:
                    student = invoice.student
                    if student.parent:
                        siblings = student.parent.wards.exclude(pk=student.pk)
                        for sibling in siblings:
                            try:
                                sibling_invoice = InvoiceModel.objects.get(
                                    student=sibling,
                                    session=invoice.session,
                                    term=invoice.term
                                )
                                sibling_item = InvoiceItemModel.objects.get(
                                    invoice=sibling_invoice,
                                    fee_master=item.fee_master
                                )
                                sibling_item.paid_by_sibling = student
                                sibling_item.amount_paid = sibling_item.amount_after_discount
                                sibling_item.save(update_flags=['paid_by_sibling', 'amount_paid'])
                            except (InvoiceModel.DoesNotExist, InvoiceItemModel.DoesNotExist):
                                continue

        # Save the actual breakdown to the payment record
        payment.item_breakdown = item_breakdown
        payment.save(update_fields=['item_breakdown'])

        # Update invoice status
        invoice.refresh_from_db()
        if invoice.balance <= Decimal('0.01'):
            invoice.status = InvoiceModel.Status.PAID
        else:
            invoice.status = InvoiceModel.Status.PARTIALLY_PAID
        invoice.save(update_fields=['status'])

    allocation_method = "auto-distributed" if (
            not parent_allocations or override_allocation) else "parent's specified items"
    messages.success(request, f"Payment of ₦{payment.amount:,.2f} confirmed successfully ({allocation_method}).")
    return redirect('pending_fee_payment_list')


@login_required
@permission_required('finance.change_feepaymentmodel', raise_exception=True)
def payment_review_view(request, payment_id):
    """
    Confirms a pending fee payment uploaded by parent.
    Supports both quick payment (auto-distribute) and itemized payment (parent's allocation).
    Staff can override parent's allocation if needed.
    """
    payment = get_object_or_404(FeePaymentModel, pk=payment_id)

    # Check if this is a confirmation with override
    context = {
        'payment': payment
    }
    return render(request, 'finance/payment/review.html', context)


class FeePaymentRevertView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Handles reverting a confirmed student fee payment to safely correct errors."""
    permission_required = 'finance.add_feemodel'

    def post(self, request, *args, **kwargs):
        payment = get_object_or_404(FeePaymentModel, pk=self.kwargs['pk'])
        invoice = payment.invoice
        student = invoice.student

        # Only revert confirmed payments
        if payment.status != FeePaymentModel.PaymentStatus.CONFIRMED:
            messages.warning(request, "Only confirmed payments can be reverted.")
            return redirect('finance_student_dashboard', pk=student.pk)

        with transaction.atomic():
            # Check if payment has item_breakdown
            if payment.item_breakdown:
                # Use the saved breakdown to reverse allocations
                for item_id_str, amount_str in payment.item_breakdown.items():
                    try:
                        item_id = int(item_id_str)
                        amount = Decimal(amount_str)

                        item = InvoiceItemModel.objects.get(pk=item_id, invoice=invoice)

                        # Reverse the payment
                        item.amount_paid -= amount
                        # Ensure amount_paid doesn't go negative
                        if item.amount_paid < 0:
                            item.amount_paid = Decimal('0.00')
                        item.save(update_fields=['amount_paid'])

                        # Handle parent-bound fees - reverse sibling payments
                        if item.fee_master.fee.parent_bound:
                            # If this item was fully paid and triggered sibling payments, reverse them
                            if student.parent:
                                siblings = student.parent.wards.exclude(pk=student.pk)
                                for sibling in siblings:
                                    try:
                                        sibling_invoice = InvoiceModel.objects.get(
                                            student=sibling,
                                            session=invoice.session,
                                            term=invoice.term
                                        )
                                        sibling_item = InvoiceItemModel.objects.get(
                                            invoice=sibling_invoice,
                                            fee_master=item.fee_master
                                        )

                                        # Only reverse if this student paid for the sibling
                                        if sibling_item.paid_by_sibling == student:
                                            sibling_item.paid_by_sibling = None
                                            sibling_item.amount_paid = Decimal('0.00')
                                            sibling_item.save(update_fields=['paid_by_sibling', 'amount_paid'])

                                    except (InvoiceModel.DoesNotExist, InvoiceItemModel.DoesNotExist):
                                        continue

                    except (ValueError, InvoiceItemModel.DoesNotExist) as e:
                        # Log the error but continue reverting other items
                        messages.warning(request, f"Could not reverse item {item_id_str}: {str(e)}")
                        continue
            else:
                # Fallback: No breakdown saved (old payment records)
                # We can't accurately reverse, so just warn the user
                messages.error(
                    request,
                    f"This payment (Reference: {payment.reference}) was created before the item tracking feature. "
                    f"Manual adjustment may be required. Please review the invoice items carefully."
                )
                return redirect('finance_student_dashboard', pk=student.pk)

                # Still mark as reverted but don't touch item amounts

            # Mark payment as reverted
            payment.status = FeePaymentModel.PaymentStatus.REVERTED
            payment.save(update_fields=['status'])

            if payment.payment_mode == FeePaymentModel.PaymentMode.WALLET:
                try:
                    wallet = StudentWalletModel.objects.select_for_update().get(student=student)
                    wallet.fee_balance += payment.amount
                    wallet.save(update_fields=['fee_balance'])

                    messages.info(
                        request,
                        f"₦{payment.amount:,.2f} has been refunded to {student}'s fee wallet."
                    )
                except StudentWalletModel.DoesNotExist:
                    messages.warning(
                        request,
                        f"Payment reverted but student wallet not found. Manual refund of ₦{payment.amount:,.2f} may be required."
                    )
                    
            # Update invoice status based on new balance
            invoice.refresh_from_db()
            if invoice.amount_paid <= 0:
                invoice.status = InvoiceModel.Status.UNPAID
            elif invoice.balance <= Decimal('0.01'):
                invoice.status = InvoiceModel.Status.PAID
            else:
                invoice.status = InvoiceModel.Status.PARTIALLY_PAID
            invoice.save(update_fields=['status'])

        messages.warning(request, f"Payment {payment.reference or payment.pk} has been reverted successfully.")
        return redirect('finance_student_dashboard', pk=student.pk)


class FeePaymentReceiptView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """Displays a printable receipt for a single student fee payment."""
    model = FeePaymentModel
    permission_required = 'finance.view_feemodel'
    template_name = 'finance/payment/receipt.html'
    context_object_name = 'payment'


# -------------------------
# Expense Category Views
# -------------------------
class ExpenseCategoryCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView
):
    model = ExpenseCategoryModel
    permission_required = 'finance.add_expensemodel'
    form_class = ExpenseCategoryForm
    template_name = 'finance/expense_category/index.html'
    success_message = 'Expense Category Successfully Created'

    def get_success_url(self):
        return reverse('expense_category_index')

    def dispatch(self, request, *args, **kwargs):
        # original UX: POST-only create endpoint; GET redirects back to index
        if request.method == 'GET':
            return redirect(reverse('expense_category_index'))
        return super().dispatch(request, *args, **kwargs)


class ExpenseCategoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ExpenseCategoryModel
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/expense_category/index.html'
    context_object_name = "category_list"

    def get_queryset(self):
        return ExpenseCategoryModel.objects.all().order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ExpenseCategoryForm()
        return context


class ExpenseCategoryUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = ExpenseCategoryModel
    permission_required = 'finance.add_expensemodel'
    form_class = ExpenseCategoryForm
    template_name = 'finance/expense_category/index.html'
    success_message = 'Expense Category Successfully Updated'

    def get_success_url(self):
        return reverse('expense_category_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('expense_category_index'))
        return super().dispatch(request, *args, **kwargs)


class ExpenseCategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ExpenseCategoryModel
    permission_required = 'finance.add_expensemodel'
    template_name = 'finance/expense_category/delete.html'
    context_object_name = "category"
    success_message = 'Expense Category Successfully Deleted'

    def get_success_url(self):
        return reverse('expense_category_index')


# -------------------------
# Expense Views
# -------------------------
class ExpenseListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = ExpenseModel
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/expense/index.html'
    context_object_name = "expense_list"
    paginate_by = 20

    def get_queryset(self):
        # Select related fields that exist on your model
        queryset = ExpenseModel.objects.select_related(
            'category', 'session', 'term', 'created_by', 'bank_account',
            'prepared_by', 'authorised_by', 'collected_by'
        ).order_by('-expense_date')

        # Filter by category, session, term
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        session = self.request.GET.get('session')
        if session:
            queryset = queryset.filter(session_id=session)

        term = self.request.GET.get('term')
        if term:
            queryset = queryset.filter(term_id=term)

        # Search over description, reference, name, and voucher_number
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(description__icontains=search) |
                Q(reference__icontains=search) |
                Q(name__icontains=search) |
                Q(voucher_number__icontains=search) |
                Q(notes__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = ExpenseCategoryModel.objects.all().order_by('name')
        context['total_amount'] = self.get_queryset().aggregate(Sum('amount'))['amount__sum'] or 0

        # Pass search query back to template
        context['search_query'] = self.request.GET.get('search', '')

        return context


class ExpenseCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ExpenseModel
    permission_required = 'finance.add_expensemodel'
    form_class = ExpenseForm
    template_name = 'finance/expense/create.html'
    success_message = 'Expense Successfully Created'

    def get_success_url(self):
        return reverse('expense_detail', kwargs={'pk': self.object.pk})


class ExpenseUpdateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView):
    model = ExpenseModel
    permission_required = 'finance.add_expensemodel'
    form_class = ExpenseForm
    template_name = 'finance/expense/edit.html'
    success_message = 'Expense Successfully Updated'
    context_object_name = "expense"

    def get_success_url(self):
        return reverse('expense_detail', kwargs={'pk': self.object.pk})


class ExpenseDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = ExpenseModel
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/expense/detail.html'
    context_object_name = "expense"


class ExpensePrintVoucherView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """
    View for printing payment voucher
    """
    model = ExpenseModel
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/expense/print_voucher.html'
    context_object_name = "expense"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get school settings if available
        try:
            from admin_site.models import SchoolInfoModel
            context['school_setting'] = SchoolInfoModel.objects.first()
        except:
            context['school_setting'] = None

        return context


# -------------------------
# Income Category Views
# -------------------------
class IncomeCategoryCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView
):
    model = IncomeCategoryModel
    permission_required = 'finance.add_expensemodel'
    form_class = IncomeCategoryForm
    template_name = 'finance/income_category/index.html'
    success_message = 'Income Category Successfully Created'

    def get_success_url(self):
        return reverse('income_category_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('income_category_index'))
        return super().dispatch(request, *args, **kwargs)


class IncomeCategoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = IncomeCategoryModel
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/income_category/index.html'
    context_object_name = "category_list"

    def get_queryset(self):
        return IncomeCategoryModel.objects.all().order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = IncomeCategoryForm()
        return context


class IncomeCategoryUpdateView(
    LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView
):
    model = IncomeCategoryModel
    permission_required = 'finance.add_expensemodel'
    form_class = IncomeCategoryForm
    template_name = 'finance/income_category/index.html'
    success_message = 'Income Category Successfully Updated'

    def get_success_url(self):
        return reverse('income_category_index')

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(reverse('income_category_index'))
        return super().dispatch(request, *args, **kwargs)


class IncomeCategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = IncomeCategoryModel
    permission_required = 'finance.add_expensemodel'
    template_name = 'finance/income_category/delete.html'
    context_object_name = "category"
    success_message = 'Income Category Successfully Deleted'

    def get_success_url(self):
        return reverse('income_category_index')


# -------------------------
# Income Views
# -------------------------
class IncomeListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = IncomeModel
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/income/index.html'
    context_object_name = "income_list"
    paginate_by = 20

    def get_queryset(self):
        queryset = IncomeModel.objects.select_related(
            'category', 'session', 'term', 'created_by'
        ).order_by('-income_date')

        # Filter by category, session, term
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        session = self.request.GET.get('session')
        if session:
            queryset = queryset.filter(session_id=session)

        term = self.request.GET.get('term')
        if term:
            queryset = queryset.filter(term_id=term)

        # Search over description, reference, source
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(description__icontains=search) |
                Q(reference__icontains=search) |
                Q(source__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = IncomeCategoryModel.objects.all().order_by('name')
        # removed departments
        context['total_amount'] = self.get_queryset().aggregate(Sum('amount'))['amount__sum'] or 0
        return context


class IncomeCreateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView):
    model = IncomeModel
    permission_required = 'finance.add_expensemodel'
    form_class = IncomeForm
    template_name = 'finance/income/create.html'
    success_message = 'Income Successfully Created'

    def get_success_url(self):
        return reverse('income_detail', kwargs={'pk': self.object.pk})


class IncomeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView):
    model = IncomeModel
    permission_required = 'finance.add_expensemodel'
    form_class = IncomeForm
    template_name = 'finance/income/edit.html'
    success_message = 'Income Successfully Updated'

    def get_success_url(self):
        return reverse('income_detail', kwargs={'pk': self.object.pk})


class IncomeDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = IncomeModel
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/income/detail.html'
    context_object_name = "income"


# ===================================================================
# Staff Bank Detail Views (Modal Interface)
# ===================================================================
class StaffBankDetailListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = StaffBankDetail
    permission_required = 'finance.view_salaryrecord'
    template_name = 'finance/staff_bank/index.html'
    context_object_name = "bank_detail_list"

    def get_queryset(self):
        return StaffBankDetail.objects.select_related('staff__staff_profile__user').order_by(
            'staff__staff_profile__user__first_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = StaffBankDetailForm()
        return context


class StaffBankDetailCreateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView):
    model = StaffBankDetail
    permission_required = 'finance.add_salaryrecord'
    form_class = StaffBankDetailForm

    def get_success_url(self):
        return reverse('finance_staff_bank_detail_list')

    def form_valid(self, form):
        messages.success(self.request, "Bank Detail Created Successfully.")
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET': return redirect(self.success_url)
        return super().dispatch(request, *args, **kwargs)


class StaffBankDetailDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = StaffBankDetail
    permission_required = 'finance.add_salaryrecord'
    template_name = 'finance/staff_bank/delete.html'
    context_object_name = "bank_detail"
    success_url = reverse_lazy('finance_staff_bank_detail_list')

    def form_valid(self, form):
        messages.success(self.request, "Bank Detail Deleted Successfully.")
        return super().form_valid(form)



class StaffBankDetailUpdateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView):
    model = StaffBankDetail
    permission_required = 'finance.add_salaryrecord'
    form_class = StaffBankDetailForm
    success_url = reverse_lazy('finance_staff_bank_detail_list')

    def form_valid(self, form):
        messages.success(self.request, "Bank Detail Updated Successfully.")
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET': return redirect(self.success_url)
        return super().dispatch(request, *args, **kwargs)


class StaffBankDetailDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = StaffBankDetail
    permission_required = 'finance.add_salaryrecord'
    template_name = 'finance/staff_bank/delete.html'
    context_object_name = "bank_detail"
    success_url = reverse_lazy('finance_staff_bank_detail_list')

    def form_valid(self, form):
        messages.success(self.request, "Bank Detail Deleted Successfully.")
        return super().form_valid(form)


# ===================================================================
# School Bank Detail Views (Modal Interface)
# ===================================================================
class SchoolBankDetailListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = SchoolBankDetail
    permission_required = 'finance.view_expensemodel'
    template_name = 'finance/school_bank/index.html'
    context_object_name = "school_bank_detail_list"

    def get_queryset(self):
        return SchoolBankDetail.objects.order_by('bank_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Assuming you will create a SchoolBankDetailForm
        context['form'] = SchoolBankDetailForm()
        return context


class SchoolBankDetailCreateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView):
    model = SchoolBankDetail
    permission_required = 'finance.add_expensemodel'
    form_class = SchoolBankDetailForm

    def get_success_url(self):
        return reverse('finance_school_bank_detail_list')

    def form_valid(self, form):
        messages.success(self.request, "School Bank Detail Created Successfully.")
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(self.success_url)
        return super().dispatch(request, *args, **kwargs)


class SchoolBankDetailUpdateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView):
    model = SchoolBankDetail
    permission_required = 'finance.add_expensemodel'
    form_class = SchoolBankDetailForm
    success_url = reverse_lazy('finance_school_bank_detail_list')

    def form_valid(self, form):
        messages.success(self.request, "School Bank Detail Updated Successfully.")
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        if request.method == 'GET':
            return redirect(self.success_url)
        return super().dispatch(request, *args, **kwargs)


class SchoolBankDetailDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = SchoolBankDetail
    permission_required = 'finance.add_expensemodel'
    template_name = 'finance/school_bank/delete.html'
    context_object_name = "school_bank_detail"
    success_url = reverse_lazy('finance_school_bank_detail_list')

    def form_valid(self, form):
        messages.success(self.request, "School Bank Detail Deleted Successfully.")
        return super().form_valid(form)


# ===================================================================
# Salary Advance Views (NEW - Multi-page Interface)
# ===================================================================
# ============================================
# SALARY ADVANCE VIEWS
# ============================================

def salary_advance_list_view(request):
    """List view for salary advances with filtering and export."""
    session_id = request.GET.get('session', None)
    term_id = request.GET.get('term', None)
    status_filter = request.GET.get('status', '').strip()
    search_query = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    page = request.GET.get('page', 1)

    school_setting = SchoolSettingModel.objects.first()
    if not session_id:
        session = school_setting.session
    else:
        session = SessionModel.objects.get(id=session_id)
    if not term_id:
        term = school_setting.term
    else:
        term = TermModel.objects.get(id=term_id)

    session_list = SessionModel.objects.all()
    term_list = TermModel.objects.all()

    queryset = SalaryAdvance.objects.filter(
        session=session,
        term=term
    ).select_related('staff__staff_profile__user').order_by('-request_date')

    if status_filter:
        queryset = queryset.filter(status=status_filter)

    if search_query:
        queryset = queryset.filter(
            Q(staff__staff_profile__user__first_name__icontains=search_query) |
            Q(staff__staff_profile__user__last_name__icontains=search_query) |
            Q(staff__staff_id__icontains=search_query)
        )

    if date_from:
        queryset = queryset.filter(request_date__gte=date_from)
    if date_to:
        queryset = queryset.filter(request_date__lte=date_to)

    # Handle downloads
    if 'download' in request.GET:
        download_format = request.GET.get('download')
        if download_format == 'excel':
            return download_salary_advance_excel(queryset, session, term)
        elif download_format == 'pdf':
            return download_salary_advance_pdf(queryset, session, term)

    paginator = Paginator(queryset, 15)
    try:
        advances = paginator.page(page)
    except PageNotAnInteger:
        advances = paginator.page(1)
    except EmptyPage:
        advances = paginator.page(paginator.num_pages)

    context = {
        'advances': advances,
        'current_session': session,
        'current_term': term,
        'session_list': session_list,
        'term_list': term_list,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'status_filter': status_filter,
    }
    return render(request, 'finance/salary_advance/index.html', context)


def download_salary_advance_excel(queryset, session, term):
    """Export salary advances to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Advances"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border_side = Side(style='thin', color='000000')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = ['Staff Member', 'Staff ID', 'Request Date', 'Amount (₦)', 'Repaid (₦)', 'Balance (₦)', 'Status']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12

    for row_num, advance in enumerate(queryset, 2):
        cell = ws.cell(row=row_num, column=1, value=str(advance.staff))
        cell.border = border

        cell = ws.cell(row=row_num, column=2, value=advance.staff.staff_id)
        cell.border = border
        cell.alignment = center_align

        date_str = advance.request_date.strftime("%b %d, %Y") if advance.request_date else ""
        cell = ws.cell(row=row_num, column=3, value=date_str)
        cell.border = border
        cell.alignment = center_align

        cell = ws.cell(row=row_num, column=4, value=float(advance.amount))
        cell.number_format = '#,##0.00'
        cell.border = border
        cell.alignment = right_align

        cell = ws.cell(row=row_num, column=5, value=float(advance.repaid_amount))
        cell.number_format = '#,##0.00'
        cell.border = border
        cell.alignment = right_align

        cell = ws.cell(row=row_num, column=6, value=float(advance.balance))
        cell.number_format = '#,##0.00'
        cell.border = border
        cell.alignment = right_align

        cell = ws.cell(row=row_num, column=7, value=advance.get_status_display())
        cell.border = border
        cell.alignment = center_align

    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"salary_advances_{session.__str__()}_{term.name}_{timestamp}.xlsx"

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def download_salary_advance_pdf(queryset, session, term):
    """Export salary advances to PDF."""
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    title = Paragraph(f"Salary Advances - {session.__str__()} - {term.name}", title_style)
    elements.append(title)

    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        alignment=TA_CENTER
    )

    generation_date = datetime.now().strftime("%B %d, %Y at %H:%M")
    info = Paragraph(f"Generated on {generation_date}", info_style)
    elements.append(info)
    elements.append(Spacer(1, 20))

    headers = ['Staff Member', 'Staff ID', 'Request Date', 'Amount (₦)', 'Repaid (₦)', 'Balance (₦)', 'Status']
    col_widths = [1.8 * inch, 1 * inch, 1.1 * inch, 1.1 * inch, 1.1 * inch, 1.1 * inch, 0.9 * inch]

    data = [headers]

    for advance in queryset:
        staff_name = str(advance.staff)
        date_str = advance.request_date.strftime("%b %d, %Y") if advance.request_date else ""

        row = [
            staff_name,
            advance.staff.staff_id,
            date_str,
            f"₦{advance.amount:,.2f}",
            f"₦{advance.repaid_amount:,.2f}",
            f"₦{advance.balance:,.2f}",
            advance.get_status_display(),
        ]

        data.append(row)

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (3, 1), (5, -1), 'RIGHT'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ])

    table.setStyle(table_style)
    elements.append(table)

    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )

    total_amount = sum(a.amount for a in queryset)
    total_repaid = sum(a.repaid_amount for a in queryset)
    total_balance = sum(a.balance for a in queryset)
    footer_text = f"Total Records: {queryset.count()} | Total Amount: ₦{total_amount:,.2f} | Total Repaid: ₦{total_repaid:,.2f} | Total Balance: ₦{total_balance:,.2f}"
    footer = Paragraph(footer_text, footer_style)
    elements.append(footer)

    doc.build(elements)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"salary_advances_{session.__str__()}_{term.name}_{timestamp}.pdf"

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


class SalaryAdvanceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = SalaryAdvance
    permission_required = 'finance.add_salaryrecord'
    form_class = SalaryAdvanceForm
    template_name = 'finance/salary_advance/create.html'

    def get_success_url(self):
        messages.success(self.request, "Salary advance request submitted successfully.")
        return reverse('finance_salary_advance_detail', kwargs={'pk': self.object.pk})


class SalaryAdvanceDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = SalaryAdvance
    permission_required = 'finance.view_salaryrecord'
    template_name = 'finance/salary_advance/detail.html'
    context_object_name = 'advance'


class SalaryAdvanceActionView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """A single view to handle all status changes for a salary advance."""
    permission_required = 'finance.add_salaryrecord'

    def post(self, request, *args, **kwargs):
        advance = get_object_or_404(SalaryAdvance, pk=self.kwargs.get('pk'))
        action = request.POST.get('action')

        if action == 'approve' and advance.status == 'pending':
            advance.status = SalaryAdvance.Status.APPROVED
            advance.approved_by = request.user
            advance.approved_date = date.today()
            advance.save()
            messages.success(request, "Salary advance has been approved.")
        elif action == 'reject' and advance.status == 'pending':
            advance.status = SalaryAdvance.Status.REJECTED
            advance.save()
            messages.warning(request, "Salary advance has been rejected.")
        elif action == 'disburse' and advance.status == 'approved':
            advance.status = SalaryAdvance.Status.DISBURSED
            # In a real system, you might link this to an Expense record
            advance.save()
            messages.info(request, "Salary advance marked as disbursed.")
        else:
            messages.error(request, "Invalid action or status.")

        return redirect('finance_salary_advance_detail', pk=advance.pk)


# ===================================================================
# Staff Loan Views
# ===================================================================
def staff_loan_list_view(request):
    """List view for staff loans with filtering and export."""
    session_id = request.GET.get('session', None)
    term_id = request.GET.get('term', None)
    status_filter = request.GET.get('status', '').strip()
    search_query = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    page = request.GET.get('page', 1)

    school_setting = SchoolSettingModel.objects.first()
    if not session_id:
        session = school_setting.session
    else:
        session = SessionModel.objects.get(id=session_id)
    if not term_id:
        term = school_setting.term
    else:
        term = TermModel.objects.get(id=term_id)

    session_list = SessionModel.objects.all()
    term_list = TermModel.objects.all()

    queryset = StaffLoan.objects.filter(
        session=session,
        term=term
    ).select_related('staff__staff_profile__user').order_by('-request_date')

    if status_filter:
        queryset = queryset.filter(status=status_filter)

    if search_query:
        queryset = queryset.filter(
            Q(staff__staff_profile__user__first_name__icontains=search_query) |
            Q(staff__staff_profile__user__last_name__icontains=search_query) |
            Q(staff__staff_id__icontains=search_query)
        )

    if date_from:
        queryset = queryset.filter(request_date__gte=date_from)
    if date_to:
        queryset = queryset.filter(request_date__lte=date_to)

    # Handle downloads
    if 'download' in request.GET:
        download_format = request.GET.get('download')
        if download_format == 'excel':
            return download_staff_loan_excel(queryset, session, term)
        elif download_format == 'pdf':
            return download_staff_loan_pdf(queryset, session, term)

    paginator = Paginator(queryset, 15)
    try:
        loans = paginator.page(page)
    except PageNotAnInteger:
        loans = paginator.page(1)
    except EmptyPage:
        loans = paginator.page(paginator.num_pages)

    context = {
        'loans': loans,
        'current_session': session,
        'current_term': term,
        'session_list': session_list,
        'term_list': term_list,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'status_filter': status_filter,
    }
    return render(request, 'finance/staff_loan/index.html', context)


def download_staff_loan_excel(queryset, session, term):
    """Export staff loans to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Loans"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border_side = Side(style='thin', color='000000')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = ['Staff Member', 'Staff ID', 'Request Date', 'Amount (₦)', 'Repaid (₦)', 'Balance (₦)', 'Status']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12

    for row_num, loan in enumerate(queryset, 2):
        cell = ws.cell(row=row_num, column=1, value=str(loan.staff))
        cell.border = border

        cell = ws.cell(row=row_num, column=2, value=loan.staff.staff_id)
        cell.border = border
        cell.alignment = center_align

        date_str = loan.request_date.strftime("%b %d, %Y") if loan.request_date else ""
        cell = ws.cell(row=row_num, column=3, value=date_str)
        cell.border = border
        cell.alignment = center_align

        cell = ws.cell(row=row_num, column=4, value=float(loan.amount))
        cell.number_format = '#,##0.00'
        cell.border = border
        cell.alignment = right_align

        cell = ws.cell(row=row_num, column=5, value=float(loan.repaid_amount))
        cell.number_format = '#,##0.00'
        cell.border = border
        cell.alignment = right_align

        cell = ws.cell(row=row_num, column=6, value=float(loan.balance))
        cell.number_format = '#,##0.00'
        cell.border = border
        cell.alignment = right_align

        cell = ws.cell(row=row_num, column=7, value=loan.get_status_display())
        cell.border = border
        cell.alignment = center_align

    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"staff_loans_{session.__str__()}_{term.name}_{timestamp}.xlsx"

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def download_staff_loan_pdf(queryset, session, term):
    """Export staff loans to PDF."""
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    title = Paragraph(f"Staff Loans - {session.__str__()} - {term.name}", title_style)
    elements.append(title)

    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        alignment=TA_CENTER
    )

    generation_date = datetime.now().strftime("%B %d, %Y at %H:%M")
    info = Paragraph(f"Generated on {generation_date}", info_style)
    elements.append(info)
    elements.append(Spacer(1, 20))

    headers = ['Staff Member', 'Staff ID', 'Request Date', 'Amount (₦)', 'Repaid (₦)', 'Balance (₦)', 'Status']
    col_widths = [1.8 * inch, 1 * inch, 1.1 * inch, 1.1 * inch, 1.1 * inch, 1.1 * inch, 0.9 * inch]

    data = [headers]

    for loan in queryset:
        staff_name = str(loan.staff)
        date_str = loan.request_date.strftime("%b %d, %Y") if loan.request_date else ""

        row = [
            staff_name,
            loan.staff.staff_id,
            date_str,
            f"₦{loan.amount:,.2f}",
            f"₦{loan.repaid_amount:,.2f}",
            f"₦{loan.balance:,.2f}",
            loan.get_status_display(),
        ]

        data.append(row)

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (3, 1), (5, -1), 'RIGHT'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ])

    table.setStyle(table_style)
    elements.append(table)

    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )

    total_amount = sum(l.amount for l in queryset)
    total_repaid = sum(l.repaid_amount for l in queryset)
    total_balance = sum(l.balance for l in queryset)
    footer_text = f"Total Records: {queryset.count()} | Total Amount: ₦{total_amount:,.2f} | Total Repaid: ₦{total_repaid:,.2f} | Total Balance: ₦{total_balance:,.2f}"
    footer = Paragraph(footer_text, footer_style)
    elements.append(footer)

    doc.build(elements)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"staff_loans_{session.__str__()}_{term.name}_{timestamp}.pdf"

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


class StaffLoanCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = StaffLoan
    permission_required = 'finance.add_salaryrecord'
    form_class = StaffLoanForm
    template_name = 'finance/staff_loan/create.html'

    def get_success_url(self):
        messages.success(self.request, "Staff loan request submitted successfully.")
        return reverse('finance_staff_loan_detail', kwargs={'pk': self.object.pk})


class StaffLoanDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = StaffLoan
    permission_required = 'finance.view_salaryrecord'
    template_name = 'finance/staff_loan/detail.html'
    context_object_name = 'loan'


class StaffLoanActionView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.add_salaryrecord'

    def post(self, request, *args, **kwargs):
        loan = get_object_or_404(StaffLoan, pk=self.kwargs.get('pk'))
        action = request.POST.get('action')

        if action == 'approve' and loan.status == 'pending':
            loan.status = StaffLoan.Status.APPROVED
            loan.approved_by = request.user
            loan.approved_date = date.today()
            loan.save()
            messages.success(request, "Staff loan has been approved.")
        elif action == 'reject' and loan.status == 'pending':
            loan.status = StaffLoan.Status.REJECTED
            loan.save()
            messages.warning(request, "Staff loan has been rejected.")
        elif action == 'disburse' and loan.status == 'approved':
            loan.status = StaffLoan.Status.DISBURSED
            loan.save()
            messages.info(request, "Staff loan marked as disbursed.")
        else:
            messages.error(request, "Invalid action or status.")
        return redirect('finance_staff_loan_detail', pk=loan.pk)


class StaffLoanDebtorsListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = StaffModel
    permission_required = 'finance.view_salaryrecord'
    template_name = 'finance/staff_loan/debtors_list.html'
    context_object_name = 'debtors'

    def get_queryset(self):
        return StaffModel.objects.annotate(
            total_loan=Coalesce(Sum('staff_loans__amount', filter=Q(staff_loans__status__in=['disbursed', 'completed'])), Value(0), output_field=DecimalField()),
            total_repaid=Coalesce(Sum('staff_loans__repaid_amount', filter=Q(staff_loans__status__in=['disbursed', 'completed'])), Value(0), output_field=DecimalField())
        ).annotate(
            total_due=F('total_loan') - F('total_repaid')
        ).filter(total_due__gt=0).order_by('staff_profile__user__last_name')


class StaffLoanDebtDetailView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Shows a breakdown of a staff's debt, repayment history, and a repayment form."""
    permission_required = 'finance.view_salaryrecord'
    template_name = 'finance/staff_loan/staff_debt_detail.html'

    def get(self, request, *args, **kwargs):
        staff = get_object_or_404(StaffModel, pk=self.kwargs.get('staff_pk'))

        # Get all loans that are currently owing
        outstanding_loans = StaffLoan.objects.filter(
            staff=staff,
            status=StaffLoan.Status.DISBURSED
        ).order_by('request_date')

        # Calculate total amount due from outstanding loans
        total_due = sum(loan.balance for loan in outstanding_loans)

        # NEW: Get the history of all past repayment transactions
        repayment_history = StaffLoanRepayment.objects.filter(
            staff=staff
        ).select_related('created_by').order_by('-payment_date', '-created_at')

        # The form for making a new payment
        form = StaffLoanRepaymentForm()

        context = {
            'staff': staff,
            'outstanding_loans': outstanding_loans,
            'total_due': total_due,
            'repayment_history': repayment_history,  # Add history to the context
            'form': form
        }
        return render(request, self.template_name, context)


@login_required
@permission_required("finance.add_salaryrecord", raise_exception=True)
@transaction.atomic
def record_staff_loan_repayment(request, staff_pk):
    if request.method != 'POST':
        return redirect('finance_staff_loan_debtors')

    staff = get_object_or_404(StaffModel, pk=staff_pk)
    form = StaffLoanRepaymentForm(request.POST)

    if form.is_valid():
        amount_paid = form.cleaned_data['amount_paid']
        payment_to_apply = amount_paid

        repayment = form.save(commit=False)
        repayment.staff = staff
        repayment.created_by = request.user
        repayment.save()

        outstanding_loans = StaffLoan.objects.filter(staff=staff, status=StaffLoan.Status.DISBURSED).order_by('request_date')
        for loan in outstanding_loans:
            if payment_to_apply <= 0: break
            payment_for_this_loan = min(payment_to_apply, loan.balance)
            loan.repaid_amount += payment_for_this_loan
            payment_to_apply -= payment_for_this_loan
            if loan.balance <= 0:
                loan.status = StaffLoan.Status.COMPLETED
            loan.save()
        messages.success(request, f"Repayment of {amount_paid} recorded for {staff}.")
    else:
        messages.error(request, "Invalid data submitted. Please check the form.")
    return redirect('finance_staff_loan_debt_detail', staff_pk=staff.pk)



class DepositPaymentSelectStudentView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, TemplateView):
    template_name = 'finance/funding/select_student.html'
    permission_required = 'finance.add_studentfundingmodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['class_list'] = ClassesModel.objects.all().order_by('name')
        student_list = StudentModel.objects.all()
        context['student_list'] = serializers.serialize("json", student_list)
        return context


class DepositPaymentSelectStaffView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, TemplateView):
    template_name = 'finance/funding/select_staff.html'
    permission_required = 'finance.add_studentfundingmodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        staff_list = StaffModel.objects.all()
        context['staff_list'] = serializers.serialize("json", staff_list)
        return context


@login_required
def deposit_get_class_students(request):
    class_pk = request.GET.get('class_pk')
    section_pk = request.GET.get('section_pk')

    student_list = StudentModel.objects.filter(student_class=class_pk, class_section=section_pk)
    result = ''
    for student in student_list:
        full_name = "{} {}".format(student.first_name.title(), student.last_name.title())
        result += """<li class='list-group-item select_student d-flex justify-content-between align-items-center' student_id='{}'>
        {} </li>""".format(student.id, full_name)
    if result == '':
        result += """<li class='list-group-item  d-flex justify-content-between align-items-center bg-danger text-white'>
        No Student in Selected Class</li>"""
    return HttpResponse(result)


@login_required
def deposit_get_class_students_by_reg_number(request):
    reg_no = request.GET.get('reg_no')

    student_list = StudentModel.objects.filter(registration_number__contains=reg_no)
    result = ''
    for student in student_list:
        full_name = "{} {}".format(student.first_name.title(), student.last_name.title())
        result += """<li class='list-group-item select_student d-flex justify-content-between align-items-center' student_id={}>
        {} </li>""".format(student.id, full_name)
    if result == '':
        result += """<li class='list-group-item d-flex justify-content-between align-items-center bg-danger text-white'>
        No Student in with inputed Registration Number</li>"""
    return HttpResponse(result)


@login_required
@permission_required("finance.view_studentfundingmodel", raise_exception=True)
def deposit_payment_list_view(request):
    session_id = request.GET.get('session', None)
    term_id = request.GET.get('term', None)
    wallet_type_filter = request.GET.get('wallet_type', '').strip()
    search_query = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    page = request.GET.get('page', 1)

    school_setting = SchoolSettingModel.objects.first()
    if not session_id:
        session = school_setting.session
    else:
        session = SessionModel.objects.get(id=session_id)
    if not term_id:
        term = school_setting.term
    else:
        term = TermModel.objects.get(id=term_id)
    session_list = SessionModel.objects.all()
    term_list = TermModel.objects.all()

    queryset = StudentFundingModel.objects.filter(session=session, term=term).exclude(status='pending').order_by('-id')

    if wallet_type_filter:
        queryset = queryset.filter(wallet_type=wallet_type_filter)

    if search_query:
        queryset = queryset.filter(
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query)
        )

    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)

    # Handle downloads
    if 'download' in request.GET:
        download_format = request.GET.get('download')
        if download_format == 'excel':
            return download_funding_excel(queryset, session, term)
        elif download_format == 'pdf':
            return download_funding_pdf(queryset, session, term)

    paginator = Paginator(queryset, 50)
    try:
        fee_payment_list = paginator.page(page)
    except PageNotAnInteger:
        fee_payment_list = paginator.page(1)
    except EmptyPage:
        fee_payment_list = paginator.page(paginator.num_pages)

    context = {
        'fee_payment_list': fee_payment_list,
        'current_session': session,
        'current_term': term,
        'session_list': session_list,
        'term_list': term_list,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'wallet_type_filter': wallet_type_filter,
    }
    return render(request, 'finance/funding/index.html', context)


def download_funding_excel(queryset, session, term):
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Funding"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border_side = Side(style='thin', color='000000')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = ['Student', 'Class', 'Amount Paid (₦)', 'Date', 'Method', 'Status']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12

    for row_num, payment in enumerate(queryset, 2):
        cell = ws.cell(row=row_num, column=1, value=f"{payment.student.first_name} {payment.student.last_name}")
        cell.border = border

        student_class = ""
        if payment.student.student_class:
            student_class = f"{payment.student.student_class} {payment.student.class_section or ''}"
        cell = ws.cell(row=row_num, column=2, value=student_class)
        cell.border = border
        cell.alignment = center_align

        cell = ws.cell(row=row_num, column=3, value=float(payment.amount))
        cell.number_format = '#,##0.00'
        cell.border = border
        cell.alignment = right_align

        date_str = payment.created_at.strftime("%b %d, %Y") if payment.created_at else ""
        cell = ws.cell(row=row_num, column=4, value=date_str)
        cell.border = border
        cell.alignment = center_align

        method = payment.mode
        if payment.method:
            method += f" ({payment.method})"
        cell = ws.cell(row=row_num, column=5, value=method)
        cell.border = border

        cell = ws.cell(row=row_num, column=6, value=payment.status.title())
        cell.border = border
        cell.alignment = center_align

    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"student_funding_{session.__str__()}_{term.name}_{timestamp}.xlsx"

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def download_funding_pdf(queryset, session, term):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    title = Paragraph(f"Student Wallet Funding - {session.__str__()} - {term.name}", title_style)
    elements.append(title)

    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        alignment=TA_CENTER
    )

    generation_date = datetime.now().strftime("%B %d, %Y at %H:%M")
    info = Paragraph(f"Generated on {generation_date}", info_style)
    elements.append(info)
    elements.append(Spacer(1, 20))

    headers = ['Student', 'Class', 'Amount (₦)', 'Date', 'Method', 'Status']
    col_widths = [2 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch, 1.5 * inch, 1 * inch]

    data = [headers]

    for payment in queryset:
        student_name = f"{payment.student.first_name} {payment.student.last_name}"

        student_class = ""
        if payment.student.student_class:
            student_class = f"{payment.student.student_class} {payment.student.class_section or ''}"

        date_str = payment.created_at.strftime("%b %d, %Y") if payment.created_at else ""

        method = payment.mode
        if payment.method:
            method += f"\n({payment.method})"

        row = [
            student_name,
            student_class,
            f"₦{payment.amount:,.2f}",
            date_str,
            method,
            payment.status.title(),
        ]

        data.append(row)

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ])

    table.setStyle(table_style)
    elements.append(table)

    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )

    total_amount = sum(p.amount for p in queryset)
    footer_text = f"Total Records: {queryset.count()} | Total Amount: ₦{total_amount:,.2f}"
    footer = Paragraph(footer_text, footer_style)
    elements.append(footer)

    doc.build(elements)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"student_funding_{session.__str__()}_{term.name}_{timestamp}.pdf"

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
@permission_required("finance.view_studentfundingmodel", raise_exception=True)
def staff_deposit_payment_list_view(request):
    session_id = request.GET.get('session', None)
    term_id = request.GET.get('term', None)
    search_query = request.GET.get('search', '').strip()
    page = request.GET.get('page', 1)

    school_setting = SchoolSettingModel.objects.first()
    if not session_id:
        session = school_setting.session
    else:
        session = SessionModel.objects.get(id=session_id)
    if not term_id:
        term = school_setting.term
    else:
        term = TermModel.objects.get(id=term_id)
    session_list = SessionModel.objects.all()
    term_list = TermModel.objects.all()

    # Base queryset
    queryset = StaffFundingModel.objects.filter(session=session, term=term).exclude(status='pending').order_by('-id')

    # Apply search filter if search query is provided
    if search_query:
        queryset = queryset.filter(
            Q(staff__first_name__icontains=search_query) |
            Q(staff__last_name__icontains=search_query)
        )

    # Apply pagination
    paginator = Paginator(queryset, 50)
    try:
        fee_payment_list = paginator.page(page)
    except PageNotAnInteger:
        fee_payment_list = paginator.page(1)
    except EmptyPage:
        fee_payment_list = paginator.page(paginator.num_pages)

    context = {
        'fee_payment_list': fee_payment_list,
        'current_session': session,
        'current_term': term,
        'session_list': session_list,
        'term_list': term_list,
        'search_query': search_query  # Pass search query to template
    }
    return render(request, 'finance/funding/staff_index.html', context)


@login_required
@permission_required("finance.add_studentfundingmodel", raise_exception=True)
def deposit_create_view(request, student_pk):
    student = StudentModel.objects.get(pk=student_pk)
    setting = SchoolSettingModel.objects.first()

    if request.method == 'POST':
        form = StudentFundingForm(request.POST, request.FILES)  # Pass request.FILES for file uploads
        if form.is_valid():
            deposit = form.save(commit=False)  # Don't save yet, we need to set the student
            deposit.student = student  # Associate the funding with the student
            wallet_type = deposit.wallet_type

            try:
                profile = StaffProfileModel.objects.get(user=request.user)
                deposit.created_by = profile.staff
            except Exception:
                pass
            # Set session and term based on school setting if not provided by form
            if not deposit.session:
                deposit.session = setting.session
            if not deposit.term:
                deposit.term = setting.term

            amount = deposit.amount  # Get amount directly from the saved instance
            messages.success(request, f'Deposit of ₦{amount} successful!')

            # Update student wallet
            student_wallet, created = StudentWalletModel.objects.get_or_create(student=student)

            if wallet_type == StudentFundingModel.WalletType.CANTEEN:
                student_wallet.balance += amount

                if student_wallet.debt > 0:
                    if student_wallet.balance > student_wallet.debt:
                        student_wallet.balance -= student_wallet.debt
                        student_wallet.debt = 0
                    else:
                        student_wallet.debt -= student_wallet.balance
                        student_wallet.balance = 0
            else:  # FEE wallet
                student_wallet.fee_balance += amount

            student_wallet.save()

            deposit.balance = student_wallet.balance - student_wallet.debt
            deposit.save()  # Now save the deposit

            try:
                target_timezone = pytz_timezone('Africa/Lagos')

                localized_created_at = timezone.localtime(deposit.created_at, timezone=target_timezone)

                formatted_time = localized_created_at.strftime(
                    f"%B {localized_created_at.day}{get_day_ordinal_suffix(localized_created_at.day)} %Y %I:%M%p"
                )

                log = f"""
                           <div class='text-white bg-success' style='padding:5px;'>
                           <p class=''>Student Wallet Funding: <a href={reverse('deposit_detail', kwargs={'pk': deposit.id})}><b>₦{amount}</b></a> deposit to wallet of
                           <a href={reverse('student_detail', kwargs={'pk': deposit.student.id})}><b>{deposit.student.__str__().title()}</b></a>
                            by <a href={reverse('staff_detail', kwargs={'pk': deposit.created_by.id})}><b>{deposit.created_by.__str__().title()}</b></a>
                           <br><span style='float:right'>{formatted_time}</span>
                           </p>
    
                           </div>
                           """

                activity = ActivityLogModel.objects.create(log=log)
                activity.save()
            except Exception:
                pass

            return redirect('deposit_detail',
                        pk=deposit.pk)  # Redirect to prevent form resubmission on refresh
        else:
            # If form is not valid, messages.error can show form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = StudentFundingForm()  # Instantiate an empty form for GET request

    context = {
        'student': student,
        'form': form,
        'payment_list': StudentFundingModel.objects.filter(student=student, term=setting.term,
                                                           session=setting.session).order_by('-created_at'),
        'setting': setting
    }
    return render(request, 'finance/funding/create.html', context)


@login_required
@permission_required("finance.change_studentfundingmodel", raise_exception=True)
def deposit_revert_view(request, pk):
    """
    Revert a funding (refund). POST only. Expects 'reason' in POST.
    Conditions:
      - funding must be CONFIRMED (or you can adapt to other statuses),
      - student wallet balance must be >= funding.amount
    On success: deduct from wallet, mark funding as REVERTED, record who and why.
    """
    funding = get_object_or_404(StudentFundingModel, pk=pk)
    student = funding.student

    # Only allow revert for confirmed (customize if you allow other statuses)
    if funding.status != StudentFundingModel.PaymentStatus.CONFIRMED:
        messages.error(request, "Only confirmed funding records can be reverted.")
        return redirect('deposit_detail', pk=funding.pk)

    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('deposit_detail', pk=funding.pk)

    reason = request.POST.get('reason', '').strip()
    if not reason:
        messages.error(request, "Please provide a reason for the revert.")
        return redirect('deposit_detail', pk=funding.pk)

    try:
        profile = StaffProfileModel.objects.get(user=request.user)
        staff = profile.staff
    except Exception:
        staff = None

    # Perform atomic wallet and funding update
    with transaction.atomic():
        student_wallet, created = StudentWalletModel.objects.select_for_update().get_or_create(student=student)

        refund_amount = funding.amount
        wallet_type = funding.wallet_type

        # Deduct from wallet
        if wallet_type == StudentFundingModel.WalletType.CANTEEN:
            if student_wallet.balance < refund_amount:
                messages.error(request, "Student canteen wallet balance is insufficient to perform this revert.")
                return redirect('deposit_detail', pk=funding.pk)
            student_wallet.balance -= refund_amount
        else:  # FEE wallet
            if student_wallet.fee_balance < refund_amount:
                messages.error(request, "Student fee wallet balance is insufficient to perform this revert.")
                return redirect('deposit_detail', pk=funding.pk)
            student_wallet.fee_balance -= refund_amount

        student_wallet.save()

        # Mark funding reverted and store reason/who/when
        funding.mark_reverted(reason=reason, staff=staff)

        funding.save()

    messages.success(request, f"Funding of ₦{refund_amount} has been reverted successfully.")
    # Redirect to deposit detail page (you can change to index if you prefer)
    return redirect('deposit_detail', pk=funding.pk)


@login_required
@permission_required("finance.view_studentfundingmodel", raise_exception=True)
def deposit_detail_view(request, pk):
    deposit = get_object_or_404(StudentFundingModel, pk=pk)

    # Compute total profit here
    return render(request, 'finance/funding/detail.html', {
        'funding': deposit,
    })


@login_required
@permission_required("finance.add_studentfundingmodel", raise_exception=True)
def staff_deposit_create_view(request, staff_pk):
    staff = StaffModel.objects.get(pk=staff_pk)
    setting = SchoolSettingModel.objects.first()

    if request.method == 'POST':
        form = StaffFundingForm(request.POST, request.FILES)  # Pass request.FILES for file uploads
        if form.is_valid():
            deposit = form.save(commit=False)  # Don't save yet, we need to set the staff
            deposit.staff = staff  # Associate the funding with the staff

            try:
                profile = StaffProfileModel.objects.get(user=request.user)
                deposit.created_by = profile.staff
            except Exception:
                pass
            # Set session and term based on school setting if not provided by form
            if not deposit.session:
                deposit.session = setting.session
            if not deposit.term:
                deposit.term = setting.term

            amount = deposit.amount  # Get amount directly from the saved instance
            messages.success(request, f'Deposit of ₦{amount} successful!')

            # Update staff wallet
            staff_wallet, created = StaffWalletModel.objects.get_or_create(staff=staff)  # Get or create wallet

            staff_wallet.balance += amount

            staff_wallet.save()

            deposit.balance = staff_wallet.balance
            deposit.save()  # Now save the deposit

            try:
                target_timezone = pytz_timezone('Africa/Lagos')

                localized_created_at = timezone.localtime(deposit.created_at, timezone=target_timezone)

                formatted_time = localized_created_at.strftime(
                    f"%B {localized_created_at.day}{get_day_ordinal_suffix(localized_created_at.day)} %Y %I:%M%p"
                )

                log = f"""
                           <div class='text-white bg-success' style='padding:5px;'>
                           <p class=''>Staff Wallet Funding: <a href={reverse('staff_deposit_detail', kwargs={'pk': deposit.id})}><b>₦{amount}</b></a> deposit to wallet of
                           <a href={reverse('staff_detail', kwargs={'pk': deposit.staff.id})}><b>{deposit.staff.__str__().title()}</b></a>
                            by <a href={reverse('staff_detail', kwargs={'pk': deposit.created_by.id})}><b>{deposit.created_by.__str__().title()}</b></a>
                           <br><span style='float:right'>{formatted_time}</span>
                           </p>
    
                           </div>
                           """

                activity = ActivityLogModel.objects.create(log=log)
                activity.save()
            except Exception:
                pass

            return redirect('staff_deposit_detail',
                        pk=deposit.pk)  # Redirect to prevent form resubmission on refresh
        else:
            # If form is not valid, messages.error can show form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = StaffFundingForm()  # Instantiate an empty form for GET request

    staff_wallet, created = StaffWalletModel.objects.get_or_create(staff=staff)  # Get or create wallet

    context = {
        'staff': staff,
        'form': form,
        'payment_list': StaffFundingModel.objects.filter(staff=staff, term=setting.term,
                                                           session=setting.session).order_by('-created_at'),
        'setting': setting
    }
    return render(request, 'finance/funding/staff_create.html', context)


@login_required
@permission_required("finance.view_studentfundingmodel", raise_exception=True)
def staff_deposit_detail_view(request, pk):
    deposit = get_object_or_404(StaffFundingModel, pk=pk)

    # Compute total profit here
    return render(request, 'finance/funding/staff_detail.html', {
        'funding': deposit,
    })


@login_required
@permission_required("finance.view_studentfundingmodel", raise_exception=True)
def staff_pending_deposit_payment_list_view(request):
    session_id = request.GET.get('session', None)
    session = SessionModel.objects.get(id=session_id)
    term_id = request.GET.get('term', None)
    term = TermModel.objects.get(id=term_id)
    session_list = SessionModel.objects.all()
    term_list = TermModel.objects.all().order_by('order')
    fee_payment_list = StaffFundingModel.objects.filter(session=session, term=term, status='pending').order_by('-id')
    context = {
        'fee_payment_list': fee_payment_list,
        'session': session,
        'term': term,
        'session_list': session_list,
        'term_list': term_list,
    }
    return render(request, 'finance/funding/staff_pending.html', context)


@login_required
@permission_required("finance.change_studentfundingmodel", raise_exception=True)
@transaction.atomic
def staff_confirm_payment_view(request, payment_id):
    payment = get_object_or_404(StaffFundingModel, pk=payment_id)
    staff = payment.staff # Get the staff associated with this payment

    if request.method == 'POST':
        # Check if the payment is already confirmed or declined
        if payment.status != 'pending':
            messages.warning(request, f"Payment is already {payment.status.capitalize()}. Cannot confirm.")
            # Redirect to a list of payments or the payment detail page
            return redirect(reverse('pending_deposit_index')) # Replace with your actual URL name

        # Get or create staff wallet
        staff_wallet, created = StaffWalletModel.objects.get_or_create(staff=staff)

        # Apply the payment amount to the wallet balance
        # Keeping calculations as float as per original deposit_create_view
        staff_wallet.balance += payment.amount

        staff_wallet.save() # Save the updated wallet

        # Update the payment status and its internal balance field
        payment.status = 'confirmed'
        # Replicate the balance update from the original view
        payment.save() # Save the updated payment record

        # Log wallet confirmation
        from pytz import timezone as pytz_timezone
        localized_created_at = timezone.localtime(now(), timezone=pytz_timezone('Africa/Lagos'))
        formatted_time = localized_created_at.strftime(
            f"%B {localized_created_at.day}{get_day_ordinal_suffix(localized_created_at.day)} %Y %I:%M%p"
        )

        payment_url = reverse('staff_deposit_detail', kwargs={'pk': payment.pk})
        staff = StaffProfileModel.objects.get(user=request.user).staff
        staff_url = reverse('staff_detail', kwargs={'pk': staff.pk}) if staff else '#'

        log = f"""
        <div class='text-white bg-success p-2' style='border-radius:5px;'>
          <p>
            Payment of <a href="{payment_url}"><b>₦{payment.amount:.2f}</b></a> for
            <a href="{staff_url}"><b>{staff.__str__().title()}</b></a> was
            <b>confirmed</b> by
            <a href="{staff_url}"><b>{staff.__str__().title()}</b></a>.
            <br>
            <b>Status:</b> Confirmed &nbsp; | &nbsp;
            <b>Wallet Balance:</b> ₦{staff_wallet.balance:.2f}
            <span class='float-end'>{now().strftime('%Y-%m-%d %H:%M:%S')}</span>
          </p>
        </div>
        """

        ActivityLogModel.objects.create(
            log=log,
        )

        messages.success(request, f"Payment of ₦{payment.amount} for {staff.first_name} {staff.last_name} confirmed successfully.")
        return redirect(reverse('staff_deposit_index')) # Replace with your actual URL name

    else:
        # For GET requests to this URL, you might want to display a confirmation prompt
        # or just redirect with a message. Assuming redirect for simplicity.
        messages.info(request, "Please use a POST request to confirm this payment.")
        return redirect(reverse('staff_pending_deposit_index'))  # Replace with your actual URL name


# --- Decline Payment View ---
@login_required
@permission_required("finance.change_studentfundingmodel", raise_exception=True)
@transaction.atomic
def staff_decline_payment_view(request, payment_id):
    payment = get_object_or_404(StaffFundingModel, pk=payment_id)
    staff = payment.staff # Get the staff associated with this payment

    if request.method == 'POST':
        # Check if the payment is already confirmed or declined
        if payment.status != 'pending':
            messages.warning(request, f"Payment is already {payment.status.capitalize()}. Cannot decline.")
            # Redirect to a list of payments or the payment detail page
            return redirect(reverse('staff_pending_deposit_index')) # Replace with your actual URL name

        # Update the payment status to 'declined'
        payment.status = 'declined'
        payment.save()

        # Log wallet deposit decline
        from pytz import timezone as pytz_timezone
        localized_created_at = timezone.localtime(now(), timezone=pytz_timezone('Africa/Lagos'))
        formatted_time = localized_created_at.strftime(
            f"%B {localized_created_at.day}{get_day_ordinal_suffix(localized_created_at.day)} %Y %I:%M%p"
        )

        payment_url = reverse('staff_deposit_detail', kwargs={'pk': payment.pk})
        staff = StaffProfileModel.objects.get(user=request.user).staff
        staff_url = reverse('staff_detail', kwargs={'pk': staff.pk}) if staff else '#'

        log = f"""
        <div class='text-white bg-danger p-2' style='border-radius:5px;'>
          <p>
            Payment of <a href="{payment_url}"><b>₦{payment.amount:.2f}</b></a> for
            <a href="{staff_url}"><b>{staff.__str__().title()}</b></a> was
            <b>declined</b> by
            <a href="{staff_url}"><b>{staff.__str__().title()}</b></a>.
            <br>
            <b>Status:</b> Declined
            <span class='float-end'>{now().strftime('%Y-%m-%d %H:%M:%S')}</span>
          </p>
        </div>
        """

        ActivityLogModel.objects.create(
            log=log,
        )

        messages.success(request, f"Payment of ₦{payment.amount} for {staff.first_name} {staff.last_name} has been declined.")
        return redirect(reverse('staff_deposit_index'))  # Replace with your actual URL name
    else:
        # For GET requests to this URL, you might want to display a confirmation prompt
        # or just redirect with a message. Assuming redirect for simplicity.
        messages.info(request, "Method Not Supported.")
        return redirect(reverse('staff_pending_deposit_index'))  # Replace with your actual URL name


@login_required
@permission_required("finance.change_stafffundingmodel", raise_exception=True)
def staff_deposit_revert_view(request, pk):
    """
    Revert a staff funding (refund). POST only. Expects 'reason' in POST.
    Conditions:
      - funding.status must be CONFIRMED (adjust if you allow others)
      - staff wallet balance must be >= funding.amount
    """
    funding = get_object_or_404(StaffFundingModel, pk=pk)
    staff_person = funding.staff

    if funding.status != StaffFundingModel.PaymentStatus.CONFIRMED:
        messages.error(request, "Only confirmed funding records can be reverted.")
        return redirect('staff_deposit_detail', pk=funding.pk)

    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('staff_deposit_detail', pk=funding.pk)

    reason = request.POST.get('reason', '').strip()
    if not reason:
        messages.error(request, "Please provide a reason for the revert.")
        return redirect('staff_deposit_detail', pk=funding.pk)

    try:
        profile = StaffProfileModel.objects.get(user=request.user)
        acting_staff = profile.staff
    except Exception:
        acting_staff = None

    with transaction.atomic():
        staff_wallet = StaffWalletModel.objects.select_for_update().get_or_create(staff=staff_person)[0]
        refund_amount = funding.amount

        if staff_wallet.balance < refund_amount:
            messages.error(request, "Staff wallet balance is insufficient to perform this revert.")
            return redirect('staff_deposit_detail', pk=funding.pk)

        # Deduct from staff wallet
        staff_wallet.balance = staff_wallet.balance - refund_amount
        staff_wallet.save()

        # Mark funding as reverted (records who/when/reason)
        funding.mark_reverted(reason=reason, staff=acting_staff)

    messages.success(request, f"Funding of ₦{refund_amount} has been reverted successfully.")
    return redirect('staff_deposit_detail', pk=funding.pk)


class StaffUploadDepositView(LoginRequiredMixin, CreateView):
    """
    A view for a logged-in staff member to submit a deposit request (e.g., a teller).
    This creates a StaffFundingModel instance with a 'pending' status.
    It does NOT credit their wallet.
    """
    model = StaffFundingModel
    form_class = StaffFundingForm
    template_name = 'finance/funding/staff_upload_form.html'
    success_url = reverse_lazy('staff_deposit_history')  # Redirect to their history page

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Upload Deposit Teller"
        context['bank_detail'] = SchoolSettingModel.objects.first()
        return context

    def form_valid(self, form):
        try:
            # Get the logged-in user's Staff profile and Staff instance
            staff_member = self.request.user.staff_profile.staff
        except StaffProfileModel.DoesNotExist:
            messages.error(self.request, "Your user account is not linked to a staff profile. Cannot submit deposit.")
            return super().form_invalid(form)
        except AttributeError:
            messages.error(self.request, "Could not find a staff profile for your user.")
            return super().form_invalid(form)

        deposit = form.save(commit=False)
        deposit.staff = staff_member
        deposit.created_by = staff_member  # Record who submitted it

        # --- THIS IS THE KEY ---
        # Set status to PENDING and do NOT touch the wallet
        deposit.status = StaffFundingModel.PaymentStatus.PENDING
        # --- END KEY ---

        # Set session and term from settings
        setting = SchoolSettingModel.objects.first()
        if setting:
            if not deposit.session:
                deposit.session = setting.session
            if not deposit.term:
                deposit.term = setting.term

        # We must save the object before we can log about it
        super().form_valid(form)

        messages.success(self.request,
                         f"Your deposit request of ₦{deposit.amount:,.2f} has been submitted for review.")

        # Note: We are calling super().form_valid() which handles saving and redirection
        return redirect(self.get_success_url())


# --- 2. STAFF: VIEW MY DEPOSIT HISTORY (LIST VIEW) ---

class StaffDepositHistoryView(LoginRequiredMixin, ListView):
    """
    Shows a logged-in staff member a paginated list of their
    own deposit submissions and their current status.
    """
    model = StaffFundingModel
    template_name = 'finance/funding/staff_history_list.html'
    context_object_name = 'payment_list'
    paginate_by = 20

    def get_queryset(self):
        try:
            # Get the logged-in user's Staff profile
            staff_member = self.request.user.staff_profile.staff
            # Return ONLY this staff member's deposits, newest first
            return StaffFundingModel.objects.filter(staff=staff_member).order_by('-created_at')
        except (StaffProfileModel.DoesNotExist, AttributeError):
            messages.warning(self.request, "Your user account is not linked to a staff profile.")
            return StaffFundingModel.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "My Deposit History"
        return context


@login_required
@permission_required("finance.view_studentfundingmodel", raise_exception=True)
def pending_deposit_payment_list_view(request):
    session_id = request.GET.get('session', None)
    school_setting = SchoolSettingModel.objects.first()
    if not session_id:
        session = school_setting.session
    else:
        session = SessionModel.objects.get(id=session_id)
    term = request.GET.get('term', None)
    if not term:
        term = school_setting.term

    session_list = SessionModel.objects.all()
    term_list = TermModel.objects.all().order_by('order')
    fee_payment_list = StudentFundingModel.objects.filter(session=session, term=term, status='pending').order_by('-id')
    context = {
        'fee_payment_list': fee_payment_list,
        'session': session,
        'term': term,
        'session_list': session_list,
        'term_list': term_list,
    }
    return render(request, 'finance/funding/pending.html', context)


@login_required
@permission_required("finance.change_studentfundingmodel", raise_exception=True)
@transaction.atomic
def confirm_payment_view(request, payment_id):
    payment = get_object_or_404(StudentFundingModel, pk=payment_id)
    student = payment.student # Get the student associated with this payment

    if request.method == 'POST':
        # Check if the payment is already confirmed or declined
        if payment.status != 'pending':
            messages.warning(request, f"Payment is already {payment.status.capitalize()}. Cannot confirm.")
            # Redirect to a list of payments or the payment detail page
            return redirect(reverse('pending_deposit_index')) # Replace with your actual URL name

        # Get or create student wallet
        student_wallet, created = StudentWalletModel.objects.get_or_create(student=student)
        wallet_type = payment.wallet_type

        if wallet_type == StudentFundingModel.WalletType.CANTEEN:
            student_wallet.balance += payment.amount

            if student_wallet.debt > 0:
                if student_wallet.balance > student_wallet.debt:
                    student_wallet.balance -= student_wallet.debt
                    student_wallet.debt = 0.0
                else:
                    student_wallet.debt -= student_wallet.balance
                    student_wallet.balance = 0.0
        else:  # FEE wallet
            student_wallet.fee_balance += payment.amount

        student_wallet.save()

        # Update the payment status and its internal balance field
        payment.status = 'confirmed'
        # Replicate the balance update from the original view
        payment.balance = student_wallet.balance - student_wallet.debt
        payment.save() # Save the updated payment record

        # Log wallet confirmation
        from pytz import timezone as pytz_timezone
        localized_created_at = timezone.localtime(now(), timezone=pytz_timezone('Africa/Lagos'))
        formatted_time = localized_created_at.strftime(
            f"%B {localized_created_at.day}{get_day_ordinal_suffix(localized_created_at.day)} %Y %I:%M%p"
        )

        student_url = reverse('student_detail', kwargs={'pk': student.pk})
        payment_url = reverse('deposit_detail', kwargs={'pk': payment.pk})
        staff = StaffProfileModel.objects.get(user=request.user).staff
        staff_url = reverse('staff_detail', kwargs={'pk': staff.pk}) if staff else '#'

        log = f"""
        <div class='text-white bg-success p-2' style='border-radius:5px;'>
          <p>
            Payment of <a href="{payment_url}"><b>₦{payment.amount:.2f}</b></a> for
            <a href="{student_url}"><b>{student.__str__().title()}</b></a> was
            <b>confirmed</b> by
            <a href="{staff_url}"><b>{staff.__str__().title()}</b></a>.
            <br>
            <b>Status:</b> Confirmed &nbsp; | &nbsp;
            <b>Wallet Balance:</b> ₦{student_wallet.balance:.2f}
            <span class='float-end'>{now().strftime('%Y-%m-%d %H:%M:%S')}</span>
          </p>
        </div>
        """

        ActivityLogModel.objects.create(
            log=log,
        )

        messages.success(request, f"Payment of ₦{payment.amount} for {student.first_name} {student.last_name} confirmed successfully.")
        return redirect(reverse('deposit_index')) # Replace with your actual URL name

    else:
        # For GET requests to this URL, you might want to display a confirmation prompt
        # or just redirect with a message. Assuming redirect for simplicity.
        messages.info(request, "Please use a POST request to confirm this payment.")
        return redirect(reverse('pending_deposit_index'))  # Replace with your actual URL name


# --- Decline Payment View ---
@login_required
@permission_required("finance.change_studentfundingmodel", raise_exception=True)
@transaction.atomic
def decline_payment_view(request, payment_id):
    payment = get_object_or_404(StudentFundingModel, pk=payment_id)
    student = payment.student # Get the student associated with this payment

    if request.method == 'POST':
        # Check if the payment is already confirmed or declined
        if payment.status != 'pending':
            messages.warning(request, f"Payment is already {payment.status.capitalize()}. Cannot decline.")
            # Redirect to a list of payments or the payment detail page
            return redirect(reverse('pending_deposit_index')) # Replace with your actual URL name

        # Update the payment status to 'declined'
        payment.status = 'declined'
        payment.save()

        # Log wallet deposit decline
        from pytz import timezone as pytz_timezone
        localized_created_at = timezone.localtime(now(), timezone=pytz_timezone('Africa/Lagos'))
        formatted_time = localized_created_at.strftime(
            f"%B {localized_created_at.day}{get_day_ordinal_suffix(localized_created_at.day)} %Y %I:%M%p"
        )

        student_url = reverse('student_detail', kwargs={'pk': student.pk})
        payment_url = reverse('deposit_detail', kwargs={'pk': payment.pk})
        staff = StaffProfileModel.objects.get(user=request.user).staff
        staff_url = reverse('staff_detail', kwargs={'pk': staff.pk}) if staff else '#'

        log = f"""
        <div class='text-white bg-danger p-2' style='border-radius:5px;'>
          <p>
            Payment of <a href="{payment_url}"><b>₦{payment.amount:.2f}</b></a> for
            <a href="{student_url}"><b>{student.__str__().title()}</b></a> was
            <b>declined</b> by
            <a href="{staff_url}"><b>{staff.__str__().title()}</b></a>.
            <br>
            <b>Status:</b> Declined
            <span class='float-end'>{now().strftime('%Y-%m-%d %H:%M:%S')}</span>
          </p>
        </div>
        """

        ActivityLogModel.objects.create(
            log=log,
        )

        messages.success(request, f"Payment of ₦{payment.amount} for {student.first_name} {student.last_name} has been declined.")
        return redirect(reverse('deposit_index'))  # Replace with your actual URL name
    else:
        # For GET requests to this URL, you might want to display a confirmation prompt
        # or just redirect with a message. Assuming redirect for simplicity.
        messages.info(request, "Method Not Supported.")
        return redirect(reverse('pending_deposit_index'))  # Replace with your actual URL name


@login_required
def fee_dashboard(request):
    # Get current session and term or allow filtering
    current_setting = SchoolSettingModel.objects.first()
    selected_session_id = request.GET.get('session',
                                          current_setting.session.id if current_setting and current_setting.session else None)
    selected_term_id = request.GET.get('term',
                                       current_setting.term.id if current_setting and current_setting.term else None)

    if selected_session_id:
        selected_session = SessionModel.objects.get(id=selected_session_id)
    else:
        selected_session = None

    if selected_term_id:
        selected_term = TermModel.objects.get(id=selected_term_id)
    else:
        selected_term = None

    # Base queryset for invoices
    invoice_filter = Q()
    if selected_session:
        invoice_filter &= Q(session=selected_session)
    if selected_term:
        invoice_filter &= Q(term=selected_term)

    invoices = InvoiceModel.objects.filter(invoice_filter)

    # === KEY METRICS ===
    # Calculate total expected after discounts
    total_expected_before_discount = invoices.aggregate(
        total=Sum('items__amount')
    )['total'] or Decimal('0.00')

    # Calculate total discounts
    total_discounts = StudentDiscountModel.objects.filter(
        invoice_item__invoice__in=invoices
    ).aggregate(total=Sum('amount_discounted'))['total'] or Decimal('0.00')

    total_expected = total_expected_before_discount - total_discounts

    total_paid = FeePaymentModel.objects.filter(
        invoice__in=invoices,
        status='confirmed'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    total_pending = total_expected - total_paid

    # Student funding (additional payments)
    funding_filter = Q(status='confirmed')
    if selected_session:
        funding_filter &= Q(session=selected_session)
    if selected_term:
        funding_filter &= Q(term=selected_term)

    total_funding = StudentFundingModel.objects.filter(funding_filter).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    # Collection rate
    collection_rate = (total_paid / total_expected * 100) if total_expected > 0 else 0

    # === DISTRIBUTION ANALYSIS ===

    # Fee distribution by type
    fee_distribution = InvoiceItemModel.objects.filter(
        invoice__in=invoices
    ).values(
        'fee_master__fee__name'
    ).annotate(
        expected=Sum('amount'),
        paid=Sum('amount_paid'),
        pending=F('expected') - F('paid')
    ).order_by('-expected')

    # Distribution by class
    class_distribution = invoices.values(
        'student__student_class__name'
    ).annotate(
        expected=Sum('items__amount'),
        paid_amount=Sum('payments__amount', filter=Q(payments__status='confirmed')),
        student_count=Count('student', distinct=True)
    ).order_by('-expected')

    # === PAYMENT TRENDS (Last 30 days) ===
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)

    daily_payments = []
    current_date = start_date
    while current_date <= end_date:
        day_payments = FeePaymentModel.objects.filter(
            date=current_date,
            status='confirmed',
            invoice__in=invoices
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        daily_payments.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'amount': float(day_payments)
        })
        current_date += timedelta(days=1)

    # === DEFAULTERS & ALERTS ===

    # Students with high outstanding balances
    defaulters = []
    for invoice in invoices.filter(status__in=['unpaid', 'partially_paid']):
        balance = invoice.balance
        if balance > 0:
            defaulters.append({
                'student': invoice.student,
                'class': invoice.student.student_class.name if invoice.student.student_class else 'N/A',
                'balance': balance,
                'invoice': invoice
            })

    # Sort defaulters by balance (highest first)
    defaulters = sorted(defaulters, key=lambda x: x['balance'], reverse=True)[:20]

    # === MONTHLY TRENDS (Last 12 months) ===
    monthly_trends = []
    for i in range(12):
        month_date = end_date.replace(day=1) - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if month_date.month == 12:
            month_end = month_date.replace(year=month_date.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month_date.replace(month=month_date.month + 1, day=1) - timedelta(days=1)

        month_payments = FeePaymentModel.objects.filter(
            date__range=[month_start, month_end],
            status='confirmed'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        month_funding = StudentFundingModel.objects.filter(
            created_at__date__range=[month_start, month_end],
            status='confirmed'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        monthly_trends.append({
            'month': month_date.strftime('%b %Y'),
            'fee_payments': float(month_payments),
            'funding': float(month_funding),
            'total_income': float(month_payments + month_funding)
        })

    monthly_trends.reverse()  # Latest first

    # === PAYMENT METHOD ANALYSIS ===
    payment_methods = FeePaymentModel.objects.filter(
        invoice__in=invoices,
        status='confirmed'
    ).values('payment_mode').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')

    funding_methods = StudentFundingModel.objects.filter(
        funding_filter
    ).values('method').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')

    # === RECENT ACTIVITY ===
    recent_payments = FeePaymentModel.objects.filter(
        invoice__in=invoices,
        status='confirmed'
    ).select_related(
        'invoice__student', 'invoice__student__student_class'
    ).order_by('-date', '-created_at')[:10]

    recent_funding = StudentFundingModel.objects.filter(
        funding_filter
    ).select_related(
        'student', 'student__student_class'
    ).order_by('-created_at')[:10]

    context = {
        # Filters
        'sessions': SessionModel.objects.all(),
        'terms': TermModel.objects.all(),
        'selected_session': selected_session,
        'selected_term': selected_term,

        # Key metrics
        'total_expected': total_expected,
        'total_discounts': total_discounts,  # Add this line
        'total_expected_before_discount': total_expected_before_discount,  # Add this line
        'total_paid': total_paid,
        'total_pending': total_pending,
        'total_funding': total_funding,
        'collection_rate': round(collection_rate, 1),
        'total_students': invoices.values('student').distinct().count(),
        'total_invoices': invoices.count(),
        'pending_invoices': invoices.exclude(status='paid').count(),

        # Distributions
        'fee_distribution': fee_distribution,
        'class_distribution': class_distribution,
        'defaulters': defaulters,
        'payment_methods': payment_methods,
        'funding_methods': funding_methods,

        # Recent activity
        'recent_payments': recent_payments,
        'recent_funding': recent_funding,

        # Chart data (JSON)
        'daily_payments_data': json.dumps(daily_payments),
        'monthly_trends_data': json.dumps(monthly_trends),
        'fee_distribution_chart': json.dumps([
            {'name': item['fee_master__fee__name'], 'value': float(item['expected'])}
            for item in fee_distribution
        ]),
        'class_distribution_chart': json.dumps([
            {'name': item['student__student_class__name'] or 'No Class', 'value': float(item['expected'] or 0)}
            for item in class_distribution
        ]),
    }

    return render(request, 'finance/fee_dashboard.html', context)


@login_required
def finance_dashboard(request):
    # Get current session and term or allow filtering
    current_setting = SchoolSettingModel.objects.first()
    selected_session_id = request.GET.get('session',
                                          current_setting.session.id if current_setting and current_setting.session else None)
    selected_term_id = request.GET.get('term',
                                       current_setting.term.id if current_setting and current_setting.term else None)

    if selected_session_id:
        selected_session = SessionModel.objects.get(id=selected_session_id)
    else:
        selected_session = None

    if selected_term_id:
        selected_term = TermModel.objects.get(id=selected_term_id)
    else:
        selected_term = None

    # Base filters
    session_filter = Q()
    term_filter = Q()
    if selected_session:
        session_filter = Q(session=selected_session)
    if selected_term:
        term_filter = Q(term=selected_term)

    combined_filter = session_filter & term_filter

    # === INCOME CALCULATIONS ===

    # 1. Fee payments
    fee_payments = FeePaymentModel.objects.filter(
        status='confirmed'
    )
    if selected_session:
        fee_payments = fee_payments.filter(invoice__session=selected_session)
    if selected_term:
        fee_payments = fee_payments.filter(invoice__term=selected_term)

    total_fee_income = fee_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # 2. Student funding
    total_funding_income = StudentFundingModel.objects.filter(
        combined_filter,
        status='confirmed'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # 3. Other income
    total_other_income = IncomeModel.objects.filter(
        combined_filter
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # 4. Salary advance repayments
    total_loan_repayments = StaffLoanRepayment.objects.filter(
        combined_filter
    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')

    # 5. Purchase advance settlements (payments by staff)
    total_advance_settlements_income = AdvanceSettlementModel.objects.filter(
        settlement_type='payment',
        advance__session=selected_session if selected_session else None,
        advance__term=selected_term if selected_term else None
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    total_income = (total_fee_income + total_funding_income + total_other_income +
                    total_loan_repayments + total_advance_settlements_income)

    # === EXPENSE CALCULATIONS ===

    # 1. General expenses
    total_expenses = ExpenseModel.objects.filter(
        combined_filter
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # 2. Salary payments
    total_salary_payments = SalaryRecord.objects.filter(
        combined_filter
    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')

    # 3. Supplier payments
    total_supplier_payments = SupplierPaymentModel.objects.filter(
        combined_filter,
        status='completed'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # 4. Purchase advance payments
    total_purchase_advance_payments = PurchaseAdvancePaymentModel.objects.filter(
        advance__session=selected_session if selected_session else None,
        advance__term=selected_term if selected_term else None
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # 5. Salary advance disbursements (when status changed to disbursed)
    # Note: This tracks new advances given out, not repayments
    salary_advances_given = SalaryAdvance.objects.filter(
        combined_filter,
        status='disbursed'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # 6. Advance settlements (refunds to staff)
    total_advance_settlements_expense = AdvanceSettlementModel.objects.filter(
        settlement_type='refund'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    total_expenses_paid = (total_expenses + total_salary_payments + total_supplier_payments +
                           total_purchase_advance_payments + salary_advances_given +
                           total_advance_settlements_expense)

    # === OUTSTANDING RECEIVABLES (What we're owed) ===

    # 1. Unpaid fee balances
    invoice_filter = Q()
    if selected_session:
        invoice_filter &= Q(session=selected_session)
    if selected_term:
        invoice_filter &= Q(term=selected_term)

    invoices = InvoiceModel.objects.filter(invoice_filter)
    total_fee_receivables = sum(invoice.balance for invoice in invoices if invoice.balance > 0)

    # 2. Outstanding salary advances
    outstanding_salary_advances = SalaryAdvance.objects.filter(
        status='disbursed'
    ).aggregate(
        total=Sum(F('amount') - F('repaid_amount'))
    )['total'] or Decimal('0.00')

    # 3. Outstanding purchase advances
    outstanding_purchase_advances = Decimal('0.00')
    try:
        purchase_advances = PurchaseAdvanceModel.objects.filter(
            status='disbursed'
        )
        for advance in purchase_advances:
            outstanding_purchase_advances += (advance.approved_amount - advance.disbursed_amount)
    except:
        pass  # Handle if PurchaseAdvanceModel doesn't have these fields

    total_receivables = total_fee_receivables + outstanding_salary_advances + outstanding_purchase_advances

    # === OUTSTANDING OBLIGATIONS (What we owe) ===

    # 1. Unpaid staff salaries
    # Calculate unpaid salaries manually since net_salary is a property
    unpaid_salaries = Decimal('0.00')
    salary_records = SalaryRecord.objects.filter(combined_filter)

    for record in salary_records:
        if record.net_salary > record.amount_paid:
            unpaid_salaries += (record.net_salary - record.amount_paid)

    total_obligations = unpaid_salaries

    # === NET POSITION ===
    net_cash_position = total_income - total_expenses_paid
    net_overall_position = (total_income + total_receivables) - (total_expenses_paid + total_obligations)

    # === MONTHLY TRENDS (Last 12 months) ===
    monthly_trends = []
    for i in range(12):
        month_date = timezone.now().date().replace(day=1) - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if month_date.month == 12:
            month_end = month_date.replace(year=month_date.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month_date.replace(month=month_date.month + 1, day=1) - timedelta(days=1)

        # Monthly income
        month_fee_income = FeePaymentModel.objects.filter(
            date__range=[month_start, month_end],
            status='confirmed'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        month_other_income = IncomeModel.objects.filter(
            income_date__range=[month_start, month_end]
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        month_funding = StudentFundingModel.objects.filter(
            created_at__date__range=[month_start, month_end],
            status='confirmed'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        # Monthly expenses
        month_expenses = ExpenseModel.objects.filter(
            expense_date__range=[month_start, month_end]
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        month_salary_payments = SalaryRecord.objects.filter(
            paid_date__range=[month_start, month_end]
        ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')

        total_month_income = month_fee_income + month_other_income + month_funding
        total_month_expenses = month_expenses + month_salary_payments

        monthly_trends.append({
            'month': month_date.strftime('%b %Y'),
            'income': float(total_month_income),
            'expenses': float(total_month_expenses),
            'net': float(total_month_income - total_month_expenses)
        })

    monthly_trends.reverse()

    # === INCOME BREAKDOWN ===
    income_breakdown = [
        {'name': 'Fee Payments', 'value': float(total_fee_income)},
        {'name': 'Student Funding', 'value': float(total_funding_income)},
        {'name': 'Other Income', 'value': float(total_other_income)},
        {'name': 'Salary Repayments', 'value': float(total_loan_repayments)},
        {'name': 'Advance Settlements', 'value': float(total_advance_settlements_income)},
    ]

    # === EXPENSE BREAKDOWN ===
    expense_breakdown = [
        {'name': 'General Expenses', 'value': float(total_expenses)},
        {'name': 'Staff Salaries', 'value': float(total_salary_payments)},
        {'name': 'Supplier Payments', 'value': float(total_supplier_payments)},
        {'name': 'Purchase Advances', 'value': float(total_purchase_advance_payments)},
        {'name': 'Salary Advances', 'value': float(salary_advances_given)},
        {'name': 'Staff Refunds', 'value': float(total_advance_settlements_expense)},
    ]

    # === RECENT TRANSACTIONS ===
    recent_income = []

    # Recent fee payments
    recent_fee_payments = fee_payments.select_related(
        'invoice__student'
    ).order_by('-date')[:5]

    for payment in recent_fee_payments:
        recent_income.append({
            'type': 'Fee Payment',
            'description': f'{payment.invoice.student} - {payment.payment_mode}',
            'amount': payment.amount,
            'date': payment.date
        })

    # Recent other income
    recent_other_income = IncomeModel.objects.filter(
        combined_filter
    ).order_by('-income_date')[:3]

    for income in recent_other_income:
        recent_income.append({
            'type': 'Other Income',
            'description': income.description or income.category.name,
            'amount': income.amount,
            'date': income.income_date
        })

    # Sort by date
    recent_income.sort(key=lambda x: x['date'], reverse=True)
    recent_income = recent_income[:10]

    # Recent expenses
    recent_expenses = ExpenseModel.objects.filter(
        combined_filter
    ).select_related('category').order_by('-expense_date')[:10]

    context = {
        # Filters
        'sessions': SessionModel.objects.all(),
        'terms': TermModel.objects.all(),
        'selected_session': selected_session,
        'selected_term': selected_term,

        # Key financial metrics
        'total_income': total_income,
        'total_expenses_paid': total_expenses_paid,
        'net_cash_position': net_cash_position,
        'total_receivables': total_receivables,
        'total_obligations': total_obligations,
        'net_overall_position': net_overall_position,

        # Income breakdown
        'total_fee_income': total_fee_income,
        'total_funding_income': total_funding_income,
        'total_other_income': total_other_income,
        'total_loan_repayments': total_loan_repayments,

        # Expense breakdown
        'total_expenses': total_expenses,
        'total_salary_payments': total_salary_payments,
        'total_supplier_payments': total_supplier_payments,
        'salary_advances_given': salary_advances_given,

        # Receivables breakdown
        'total_fee_receivables': total_fee_receivables,
        'outstanding_salary_advances': outstanding_salary_advances,
        'outstanding_purchase_advances': outstanding_purchase_advances,

        # Obligations
        'unpaid_salaries': unpaid_salaries,

        # Recent activity
        'recent_income': recent_income,
        'recent_expenses': recent_expenses,

        # Chart data
        'monthly_trends_data': json.dumps(monthly_trends),
        'income_breakdown_data': json.dumps(income_breakdown),
        'expense_breakdown_data': json.dumps(expense_breakdown),
    }

    return render(request, 'finance/finance_dashboard.html', context)




# ===================================================================
# Discount Model Views (Blueprint Interface)
# ===================================================================


class DiscountListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = DiscountModel
    permission_required = 'finance.view_discountmodel'
    template_name = 'finance/discount/index.html'
    context_object_name = 'discounts'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'form' not in context:
            context['form'] = DiscountForm()

        # Add the application form
        if 'application_form' not in context:
            context['application_form'] = DiscountApplicationForm()

        # --- ADD THIS ---
        # 3. Create a data map for JS autofill
        discount_data = {
            d.pk: {'type': d.discount_type, 'amount': d.amount or 0.00}
            for d in DiscountModel.objects.all()
        }
        # Pass the data as a JSON string
        context['discount_data_json'] = json.dumps(
            discount_data,
            cls=DjangoJSONEncoder
        )
        # ----------------

        return context


class DiscountCreateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView):
    model = DiscountModel
    permission_required = 'finance.add_feepaymentmodel'
    form_class = DiscountForm

    def get_success_url(self):
        return reverse('finance_discount_list')

    def form_valid(self, form):
        messages.success(self.request, "Discount Blueprint created successfully.")
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        # Redirect GET requests to the list view for modal interface pattern
        if request.method == 'GET':
            return redirect(self.get_success_url())
        return super().dispatch(request, *args, **kwargs)


class DiscountUpdateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView):
    model = DiscountModel
    permission_required = 'finance.add_feepaymentmodel'
    form_class = DiscountForm

    def get_success_url(self):
        return reverse('finance_discount_list')

    def form_valid(self, form):
        messages.success(self.request, "Discount Blueprint updated successfully.")
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        # Redirect GET requests to the list view for modal interface pattern
        if request.method == 'GET':
            return redirect(self.get_success_url())
        return super().dispatch(request, *args, **kwargs)


class DiscountDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = DiscountModel
    permission_required = 'finance.add_feepaymentmodel'
    template_name = 'finance/discount/delete.html'
    success_url = reverse_lazy('finance_discount_list')
    context_object_name = 'discount'

    def form_valid(self, form):
        # We need custom logic here to prevent deletion if the is_protected flag is True
        if self.object.is_protected:
            messages.error(self.request, f"Cannot delete Discount '{self.object.title}'. It is linked to active discount applications.")
            return redirect(self.success_url)

        messages.success(self.request, f"Discount Blueprint '{self.object.title}' deleted successfully.")
        return super().form_valid(form)


# ===================================================================
# Discount Application Views (Context/Rate Locking Interface)
# ===================================================================

class DiscountApplicationCreateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, CreateView):
    model = DiscountApplicationModel
    permission_required = 'finance.add_feepaymentmodel'
    form_class = DiscountApplicationForm

    def get_success_url(self):
        # Assuming we redirect back to the list of Discount Blueprints
        return reverse('finance_discount_application_list')

    def form_valid(self, form):
        messages.success(self.request, "Discount rate locked for the specified term.")
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        # Redirect GET requests to the list view for modal interface pattern
        if request.method == 'GET':
            return redirect(self.get_success_url())
        return super().dispatch(request, *args, **kwargs)


class DiscountApplicationUpdateView(LoginRequiredMixin, PermissionRequiredMixin, FlashFormErrorsMixin, UpdateView):
    model = DiscountApplicationModel
    permission_required = 'finance.add_feepaymentmodel'
    form_class = DiscountApplicationForm
    pk_url_kwarg = 'application_pk' # Use a distinct keyword argument to prevent clashes

    def get_success_url(self):
        return reverse('finance_discount_application_list')

    def form_valid(self, form):
        # Note: The model's save() method prevents changing discount_type
        messages.success(self.request, "Discount rate and term updated successfully.")
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        # Redirect GET requests to the list view for modal interface pattern
        if request.method == 'GET':
            return redirect(self.get_success_url())
        return super().dispatch(request, *args, **kwargs)


class DiscountApplicationDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = DiscountApplicationModel
    permission_required = 'finance.add_feepaymentmodel'
    template_name = 'finance/discount/application_delete.html'
    success_url = reverse_lazy('finance_discount_application_list')
    context_object_name = 'application'
    pk_url_kwarg = 'application_pk' # Use a distinct keyword argument

    def form_valid(self, form):
        # We need custom logic here to prevent deletion if the is_protected flag is True
        if self.object.is_protected:
            messages.error(self.request, f"Cannot delete Discount Application for '{self.object.discount.title}' as it is linked to active student records.")
            return redirect(self.success_url)

        messages.success(self.request, f"Discount Application for '{self.object.discount.title}' deleted successfully.")
        return super().form_valid(form)


# ===================================================================
# Discount Application List View (Optional Dedicated Page)
# ===================================================================


class DiscountApplicationListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = DiscountApplicationModel
    permission_required = 'finance.view_feepaymentmodel'
    template_name = 'finance/discount/application_list.html'
    context_object_name = 'applications'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # The form will have session/term defaults from its __init__
        if 'form' not in context:
            context['form'] = DiscountApplicationForm()

        # --- ADD THIS LOGIC ---
        # Create a data map for JS autofill
        discount_data = {
            d.pk: {'type': d.discount_type, 'amount': d.amount or 0.00}
            for d in DiscountModel.objects.all()
        }
        # Pass the data as a JSON string, using Django's encoder for Decimals
        context['discount_data_json'] = json.dumps(
            discount_data,
            cls=DjangoJSONEncoder
        )
        # ---------------------

        return context


class DiscountSelectStudentView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, TemplateView):
    template_name = 'finance/discount/select_student.html'
    permission_required = 'finance.add_feepaymentmodel'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['class_list'] = ClassesModel.objects.all().order_by('name')

        student_list = StudentModel.objects.all()
        context['student_list_json'] = serializers.serialize("json", student_list)

        # Build class_list_json with sections (adapt related name if necessary)
        classes_data = []
        for cls in context['class_list']:
            # try common related names for the sections relationship, adapt if different
            if hasattr(cls, 'section'):
                secs_qs = cls.section.all()
            else:
                secs_qs = []

            sections = [{'id': s.id, 'name': getattr(s, 'name', str(s))} for s in secs_qs]
            classes_data.append({'id': cls.id, 'name': cls.name, 'sections': sections})

        context['class_list_json'] = json.dumps(classes_data)
        return context


class StudentDiscountAssignView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """Assign a discount application to a student for specific invoice items."""

    permission_required = 'finance.add_feepaymentmodel'
    template_name = 'finance/discount/assign_discount.html'
    form_class = StudentDiscountAssignForm

    def get_student(self):
        return get_object_or_404(StudentModel, pk=self.kwargs['student_pk'])

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['student'] = self.get_student()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_student()
        context['student'] = student

        # Get current school settings for display purposes
        school_setting = SchoolSettingModel.objects.first()
        context['current_session'] = school_setting.session if school_setting else None
        context['current_term'] = school_setting.term if school_setting else None

        return context

    def form_valid(self, form):
        student = self.get_student()
        discount_application = form.cleaned_data['discount_application']
        discount = discount_application.discount

        # Get selected session and term from form
        selected_session = form.cleaned_data['session']
        selected_term = form.cleaned_data['term']

        # 1. Check if student's class is eligible
        if discount.applicable_classes.exists() and student.student_class not in discount.applicable_classes.all():
            messages.error(self.request,
                           f"Student's class ({student.student_class.name}) is not eligible for this discount.")
            return self.form_invalid(form)

        # 2. Check if discount has applicable fees defined
        if not discount.applicable_fees.exists():
            messages.error(self.request, "This discount has no fees defined. Please configure the discount first.")
            return self.form_invalid(form)

        # 3. Get student's invoice for the SELECTED session/term (not discount_application's)
        try:
            invoice = InvoiceModel.objects.get(
                student=student,
                session=selected_session,
                term=selected_term
            )
        except InvoiceModel.DoesNotExist:
            messages.error(self.request,
                           f"No invoice found for {selected_session}/{selected_term.name}. "
                           f"Please generate an invoice for this session/term first.")
            return self.form_invalid(form)

        # 4. Find applicable invoice items
        applicable_fees = discount.applicable_fees.all()
        applicable_items = invoice.items.filter(
            fee_master__fee__in=applicable_fees
        )

        if not applicable_items.exists():
            messages.error(self.request,
                           "No matching fees found on student's invoice for this discount.")
            return self.form_invalid(form)

        # 5. Apply discount to each applicable item
        with transaction.atomic():
            total_discounted = Decimal('0.00')
            items_processed = 0

            for item in applicable_items:
                # Check if discount already applied to this item
                if StudentDiscountModel.objects.filter(
                        student=student,
                        discount_application=discount_application,
                        invoice_item=item
                ).exists():
                    continue  # Skip if already applied

                # Calculate discount amount
                if discount_application.discount_type == DiscountModel.DiscountType.PERCENTAGE:
                    discount_amount = (item.amount * discount_application.discount_amount) / Decimal('100')
                else:  # FIXED
                    # For fixed amount, divide equally among applicable items
                    discount_amount = discount_application.discount_amount / applicable_items.count()

                # Round to 2 decimal places
                discount_amount = discount_amount.quantize(Decimal('0.01'))

                # Create discount record
                StudentDiscountModel.objects.create(
                    student=student,
                    discount_application=discount_application,
                    invoice_item=item,
                    amount_discounted=discount_amount
                )

                total_discounted += discount_amount
                items_processed += 1

            if items_processed == 0:
                messages.warning(self.request, "This discount has already been applied to all eligible fees.")
            else:
                messages.success(self.request,
                                 f"Discount applied successfully! ₦{total_discounted:,.2f} discounted across {items_processed} fee(s) "
                                 f"for {selected_session}/{selected_term.name}.")

        return redirect('finance_student_dashboard', pk=student.pk)


class GetDiscountsAjaxView(LoginRequiredMixin, View):
    """AJAX endpoint to fetch discounts based on session, term, and student class."""

    def get(self, request, *args, **kwargs):
        session_id = request.GET.get('session_id')
        term_id = request.GET.get('term_id')
        student_pk = request.GET.get('student_pk')

        # Validate required parameters
        if not all([session_id, term_id, student_pk]):
            return JsonResponse({'error': 'Missing required parameters'}, status=400)

        try:
            student = get_object_or_404(StudentModel, pk=student_pk)

            # Build discount queryset
            queryset = DiscountApplicationModel.objects.filter(
                Q(session_id=session_id, term_id=term_id) |
                Q(session__isnull=True, term__isnull=True)  # Global discounts
            ).select_related('discount')

            # Filter by student's class
            if student.student_class:
                queryset = queryset.filter(
                    Q(discount__applicable_classes__isnull=True) |
                    Q(discount__applicable_classes=student.student_class)
                ).distinct()

            # Format response
            discounts = []
            for app in queryset:
                label = f"{app.discount.title} - {app.discount_amount}"
                label += '%' if app.discount_type == 'percentage' else ' (Fixed)'

                discounts.append({
                    'id': app.id,
                    'label': label
                })

            return JsonResponse({'discounts': discounts})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class StudentDiscountIndexView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List all student discounts with filtering by session, term, and student name."""

    permission_required = 'finance.view_feepaymentmodel'
    template_name = 'finance/discount/discount_index.html'
    context_object_name = 'discounts'
    paginate_by = 50

    def get_queryset(self):
        queryset = StudentDiscountModel.objects.select_related(
            'student',
            'discount_application__discount',
            'discount_application__session',
            'discount_application__term',
            'invoice_item__invoice'
        ).order_by('-created_at')

        # Filter by session
        session_id = self.request.GET.get('session')
        if session_id:
            queryset = queryset.filter(discount_application__session_id=session_id)

        # Filter by term
        term_id = self.request.GET.get('term')
        if term_id:
            queryset = queryset.filter(discount_application__term_id=term_id)

        # Filter by student name
        student_name = self.request.GET.get('student_name')
        if student_name:
            queryset = queryset.filter(
                Q(student__first_name__icontains=student_name) |
                Q(student__last_name__icontains=student_name) |
                Q(student__registration_number__icontains=student_name)
            )

        # Filter by discount
        discount_id = self.request.GET.get('discount')
        if discount_id:
            queryset = queryset.filter(discount_application__discount_id=discount_id)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Provide filter options
        context['sessions'] = SessionModel.objects.all().order_by('-start_year')
        context['terms'] = TermModel.objects.all().order_by('order')
        context['discount_list'] = DiscountModel.objects.all().order_by('title')

        # Preserve filter values
        context['selected_session'] = self.request.GET.get('session', '')
        context['selected_term'] = self.request.GET.get('term', '')
        context['selected_discount'] = self.request.GET.get('discount', '')
        context['student_name'] = self.request.GET.get('student_name', '')

        # Calculate summary statistics
        queryset = self.get_queryset()
        context['total_discounts'] = queryset.count()
        context['total_amount_discounted'] = queryset.aggregate(
            total=Sum('amount_discounted')
        )['total'] or Decimal('0.00')

        return context


class StudentDiscountDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a student discount with confirmation."""
    model = StudentDiscountModel
    template_name = 'finance/discount/delete_discount.html'
    permission_required = 'finance.add_feepaymentmodel'
    context_object_name = 'discount'

    def get_success_url(self):
        # Redirect to the discount index page after deletion
        return reverse('finance_discount_index')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get the student discount object
        discount = self.get_object()
        context['discount'] = discount
        context['student'] = discount.student
        return context


@login_required
@permission_required("finance.view_incomemodel", raise_exception=True)
def income_expense_report(request):
    """Display comprehensive income and expense report with net income calculation."""

    # Get date parameters, default to current month
    today = date.today()
    first_day_of_month = today.replace(day=1)
    date_from = request.GET.get('date_from', '').strip() or str(first_day_of_month)
    date_to = request.GET.get('date_to', '').strip() or str(today)
    session_filter = request.GET.get('session', '').strip()
    term_filter = request.GET.get('term', '').strip()
    download_pdf = request.GET.get('download', '') == 'pdf'

    # Validate date range
    try:
        from_date = date.fromisoformat(date_from)
        to_date = date.fromisoformat(date_to)

        if from_date > to_date:
            from_date, to_date = to_date, from_date
            date_from, date_to = date_to, date_from
    except ValueError:
        from_date = first_day_of_month
        to_date = today
        date_from = str(first_day_of_month)
        date_to = str(today)

    # Generate title based on date range
    if date_from == date_to:
        report_title = from_date.strftime("%B %d, %Y")
    else:
        report_title = f"{from_date.strftime('%B %d, %Y')} - {to_date.strftime('%B %d, %Y')}"

    # ========================================================================
    # INCOME SECTION
    # ========================================================================

    # 1. Fee Payments (by fee type)
    fee_payments_qs = FeePaymentModel.objects.filter(
        date__range=[from_date, to_date],
        status='confirmed'
    ).select_related('invoice').prefetch_related('invoice__items__fee_master__fee')

    if session_filter:
        fee_payments_qs = fee_payments_qs.filter(invoice__session_id=session_filter)
    if term_filter:
        fee_payments_qs = fee_payments_qs.filter(invoice__term_id=term_filter)

    # Calculate fee breakdown
    fee_breakdown = {}
    for payment in fee_payments_qs:
        invoice_total = payment.invoice.total_amount
        if invoice_total > 0:
            for item in payment.invoice.items.all():
                fee_name = item.fee_master.fee.name
                # Proportional allocation
                item_percentage = item.amount / invoice_total
                allocated_amount = payment.amount * item_percentage
                fee_breakdown[fee_name] = fee_breakdown.get(fee_name, Decimal('0.00')) + allocated_amount

    total_fee_payments = sum(fee_breakdown.values())
    fee_breakdown_list = [{'name': k, 'amount': v} for k, v in
                          sorted(fee_breakdown.items(), key=lambda x: x[1], reverse=True)]

    # 2. Sales Revenue (Cash/POS only - excluding wallet to avoid double counting)
    sales_qs = SaleModel.objects.filter(
        sale_date__date__range=[from_date, to_date],
        status='completed',
        payment_method__in=['cash', 'pos']
    )

    if session_filter:
        sales_qs = sales_qs.filter(session_id=session_filter)
    if term_filter:
        sales_qs = sales_qs.filter(term_id=term_filter)

    sales_data = sales_qs.values('payment_method').annotate(
        total_before_discount=Sum(F('items__quantity') * F('items__unit_price')),
        total_discount=Sum('discount')
    )

    cash_sales = Decimal('0.00')
    pos_sales = Decimal('0.00')

    for sale in sales_data:
        revenue = (sale['total_before_discount'] or Decimal('0.00')) - (sale['total_discount'] or Decimal('0.00'))
        if sale['payment_method'] == 'cash':
            cash_sales = revenue
        elif sale['payment_method'] == 'pos':
            pos_sales = revenue

    total_sales_revenue = cash_sales + pos_sales

    # 3. Wallet Funding
    student_funding_qs = StudentFundingModel.objects.filter(
        created_at__date__range=[from_date, to_date],
        status='confirmed'
    )
    staff_funding_qs = StaffFundingModel.objects.filter(
        created_at__date__range=[from_date, to_date],
        status='confirmed'
    )

    if session_filter:
        student_funding_qs = student_funding_qs.filter(session_id=session_filter)
        staff_funding_qs = staff_funding_qs.filter(session_id=session_filter)
    if term_filter:
        student_funding_qs = student_funding_qs.filter(term_id=term_filter)
        staff_funding_qs = staff_funding_qs.filter(term_id=term_filter)

    student_funding = student_funding_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    staff_funding = staff_funding_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_wallet_funding = student_funding + staff_funding

    # 4. Other Income (by category)
    other_income_qs = IncomeModel.objects.filter(
        income_date__range=[from_date, to_date]
    )

    if session_filter:
        other_income_qs = other_income_qs.filter(session_id=session_filter)
    if term_filter:
        other_income_qs = other_income_qs.filter(term_id=term_filter)

    other_income_data = other_income_qs.values('category__name').annotate(
        total=Sum('amount')
    ).order_by('-total')

    total_other_income = sum(item['total'] for item in other_income_data)

    # 5. Debt Recoveries
    loan_repayments_qs = StaffLoanRepayment.objects.filter(
        payment_date__range=[from_date, to_date]
    )

    if session_filter:
        loan_repayments_qs = loan_repayments_qs.filter(session_id=session_filter)
    if term_filter:
        loan_repayments_qs = loan_repayments_qs.filter(term_id=term_filter)

    loan_repayments = loan_repayments_qs.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')

    # Salary advance repayments from payroll (now tracked in other_deductions JSON field)
    salary_records_qs = SalaryRecord.objects.filter(
        paid_date__range=[from_date, to_date],
        payment_status='paid'  # Changed from is_paid=True
    )

    if session_filter:
        salary_records_qs = salary_records_qs.filter(session_id=session_filter)
    if term_filter:
        salary_records_qs = salary_records_qs.filter(term_id=term_filter)

    # Calculate advance repayments from other_deductions JSON field
    advance_repayments = Decimal('0.00')
    for record in salary_records_qs:
        other_deductions = record.other_deductions or {}
        advance_repayments += Decimal(str(other_deductions.get('salary_advance', 0)))

    total_debt_recoveries = loan_repayments + advance_repayments

    # TOTAL REVENUE
    total_revenue = (
            total_fee_payments +
            total_sales_revenue +
            total_wallet_funding +
            total_other_income +
            total_debt_recoveries
    )

    # ========================================================================
    # EXPENSE SECTION
    # ========================================================================

    # 1. Cost of Goods Sold (COGS)
    cogs_qs = SaleItemModel.objects.filter(
        sale__sale_date__date__range=[from_date, to_date],
        sale__status='completed'
    )

    if session_filter:
        cogs_qs = cogs_qs.filter(sale__session_id=session_filter)
    if term_filter:
        cogs_qs = cogs_qs.filter(sale__term_id=term_filter)

    cogs = cogs_qs.aggregate(
        total_cogs=Sum(F('quantity') * F('unit_cost'))
    )['total_cogs'] or Decimal('0.00')

    # 2. Salary Payments
    salary_payments_qs = SalaryRecord.objects.filter(
        paid_date__range=[from_date, to_date],
        payment_status='paid'  # Changed from is_paid=True
    )

    if session_filter:
        salary_payments_qs = salary_payments_qs.filter(session_id=session_filter)
    if term_filter:
        salary_payments_qs = salary_payments_qs.filter(term_id=term_filter)

    salary_data = salary_payments_qs.aggregate(
        total=Sum('amount_paid'),
        count=Count('id')
    )
    total_salaries = salary_data['total'] or Decimal('0.00')
    staff_count = salary_data['count']

    # 3. General Expenses (by category)
    general_expenses_qs = ExpenseModel.objects.filter(
        expense_date__range=[from_date, to_date]
    )

    if session_filter:
        general_expenses_qs = general_expenses_qs.filter(session_id=session_filter)
    if term_filter:
        general_expenses_qs = general_expenses_qs.filter(term_id=term_filter)

    general_expenses_data = general_expenses_qs.values('category__name').annotate(
        total=Sum('amount')
    ).order_by('-total')

    total_general_expenses = sum(item['total'] for item in general_expenses_data)

    # 4. Supplier Payments
    supplier_payments_qs = SupplierPaymentModel.objects.filter(
        payment_date__range=[from_date, to_date],
        status='completed'
    )

    if session_filter:
        supplier_payments_qs = supplier_payments_qs.filter(session_id=session_filter)
    if term_filter:
        supplier_payments_qs = supplier_payments_qs.filter(term_id=term_filter)

    supplier_payments = supplier_payments_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # 5. Advances & Loans Disbursed
    salary_advances_qs = SalaryAdvance.objects.filter(
        approved_date__range=[from_date, to_date],
        status='disbursed'
    )
    staff_loans_qs = StaffLoan.objects.filter(
        approved_date__range=[from_date, to_date],
        status='disbursed'
    )
    purchase_advances_qs = PurchaseAdvancePaymentModel.objects.filter(
        payment_date__range=[from_date, to_date]
    )

    if session_filter:
        salary_advances_qs = salary_advances_qs.filter(session_id=session_filter)
        staff_loans_qs = staff_loans_qs.filter(session_id=session_filter)
    if term_filter:
        salary_advances_qs = salary_advances_qs.filter(term_id=term_filter)
        staff_loans_qs = staff_loans_qs.filter(term_id=term_filter)

    salary_advances = salary_advances_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    staff_loans = staff_loans_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    purchase_advances = purchase_advances_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    total_advances = salary_advances + staff_loans + purchase_advances

    # TOTAL EXPENSES
    total_operating_expenses = (
            total_salaries +
            total_general_expenses +
            supplier_payments +
            total_advances
    )

    total_expenses = cogs + total_operating_expenses

    # GROSS PROFIT & NET INCOME
    gross_profit = total_revenue - cogs
    net_income = total_revenue - total_expenses

    # ========================================================================
    # CONTEXT DATA
    # ========================================================================

    context = {
        'report_title': report_title,
        'date_from': date_from,
        'date_to': date_to,
        'session_filter': session_filter,
        'term_filter': term_filter,

        # Income Data
        'fee_breakdown': fee_breakdown_list,
        'total_fee_payments': total_fee_payments,
        'cash_sales': cash_sales,
        'pos_sales': pos_sales,
        'total_sales_revenue': total_sales_revenue,
        'student_funding': student_funding,
        'staff_funding': staff_funding,
        'total_wallet_funding': total_wallet_funding,
        'other_income_data': other_income_data,
        'total_other_income': total_other_income,
        'loan_repayments': loan_repayments,
        'advance_repayments': advance_repayments,
        'total_debt_recoveries': total_debt_recoveries,
        'total_revenue': total_revenue,

        # Expense Data
        'cogs': cogs,
        'gross_profit': gross_profit,
        'total_salaries': total_salaries,
        'staff_count': staff_count,
        'general_expenses_data': general_expenses_data,
        'total_general_expenses': total_general_expenses,
        'supplier_payments': supplier_payments,
        'salary_advances': salary_advances,
        'staff_loans': staff_loans,
        'purchase_advances': purchase_advances,
        'total_advances': total_advances,
        'total_operating_expenses': total_operating_expenses,
        'total_expenses': total_expenses,

        # Net Result
        'net_income': net_income,

        # Filters
        'sessions': SessionModel.objects.all().order_by('-id'),
        'terms': TermModel.objects.all().order_by('order'),
    }

    # Generate PDF if requested
    if download_pdf:
        return generate_income_expense_pdf(context)

    return render(request, 'finance/reports/income_expense_report.html', context)


def generate_income_expense_pdf(context):
    """Generate PDF for income and expense report."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from io import BytesIO

        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=10,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("INCOME & EXPENSE STATEMENT", title_style))

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=14,
            textColor=colors.HexColor('#3498db'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(context['report_title'], subtitle_style))
        elements.append(Spacer(1, 0.3 * inch))

        # ====================================================================
        # INCOME SECTION
        # ====================================================================

        section_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=10,
            spaceBefore=10
        )
        elements.append(Paragraph("REVENUE (INCOME)", section_style))

        income_data = [
            ['Description', 'Amount (₦)'],
        ]

        # Fee Payments
        if context['fee_breakdown']:
            income_data.append(['Fee Payments', ''])
            for fee in context['fee_breakdown']:
                income_data.append([f"  • {fee['name']}", f"{fee['amount']:,.2f}"])
            income_data.append(['', f"{context['total_fee_payments']:,.2f}"])

        # Sales Revenue
        if context['total_sales_revenue'] > 0:
            income_data.append(['Sales Revenue (Cash/POS)', ''])
            if context['cash_sales'] > 0:
                income_data.append([f"  • Cash Sales", f"{context['cash_sales']:,.2f}"])
            if context['pos_sales'] > 0:
                income_data.append([f"  • POS Sales", f"{context['pos_sales']:,.2f}"])
            income_data.append(['', f"{context['total_sales_revenue']:,.2f}"])

        # Wallet Funding
        if context['total_wallet_funding'] > 0:
            income_data.append(['Wallet Funding', ''])
            if context['student_funding'] > 0:
                income_data.append([f"  • Student Wallets", f"{context['student_funding']:,.2f}"])
            if context['staff_funding'] > 0:
                income_data.append([f"  • Staff Wallets", f"{context['staff_funding']:,.2f}"])
            income_data.append(['', f"{context['total_wallet_funding']:,.2f}"])

        # Other Income
        if context['other_income_data']:
            income_data.append(['Other Income', ''])
            for item in context['other_income_data']:
                income_data.append([f"  • {item['category__name']}", f"{item['total']:,.2f}"])
            income_data.append(['', f"{context['total_other_income']:,.2f}"])

        # Debt Recoveries
        if context['total_debt_recoveries'] > 0:
            income_data.append(['Debt Recoveries', ''])
            if context['loan_repayments'] > 0:
                income_data.append([f"  • Loan Repayments", f"{context['loan_repayments']:,.2f}"])
            if context['advance_repayments'] > 0:
                income_data.append([f"  • Advance Repayments", f"{context['advance_repayments']:,.2f}"])
            income_data.append(['', f"{context['total_debt_recoveries']:,.2f}"])

        # Total Revenue
        income_data.append(['', ''])
        income_data.append(['TOTAL REVENUE', f"{context['total_revenue']:,.2f}"])

        # Create income table
        income_table = Table(income_data, colWidths=[4 * inch, 2 * inch])
        income_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data rows
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),

            # Total row
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#2c3e50')),
        ]))

        elements.append(income_table)
        elements.append(Spacer(1, 0.3 * inch))

        # ====================================================================
        # EXPENSE SECTION
        # ====================================================================

        elements.append(Paragraph("EXPENSES", section_style))

        expense_data = [
            ['Description', 'Amount (₦)'],
        ]

        # COGS
        if context['cogs'] > 0:
            expense_data.append(['Cost of Goods Sold', f"{context['cogs']:,.2f}"])
            expense_data.append(['', ''])
            expense_data.append(['Gross Profit (Revenue - COGS)', f"{context['gross_profit']:,.2f}"])
            expense_data.append(['', ''])
            expense_data.append(['OPERATING EXPENSES', ''])

        # Salary Payments
        if context['total_salaries'] > 0:
            expense_data.append(['Salary Payments', ''])
            expense_data.append([f"  • {context['staff_count']} Staff Members", f"{context['total_salaries']:,.2f}"])

        # General Expenses
        if context['general_expenses_data']:
            expense_data.append(['General Expenses', ''])
            for item in context['general_expenses_data']:
                expense_data.append([f"  • {item['category__name']}", f"{item['total']:,.2f}"])
            expense_data.append(['', f"{context['total_general_expenses']:,.2f}"])

        # Supplier Payments
        if context['supplier_payments'] > 0:
            expense_data.append(['Supplier Payments', f"{context['supplier_payments']:,.2f}"])

        # Advances & Loans
        if context['total_advances'] > 0:
            expense_data.append(['Advances & Loans Disbursed', ''])
            if context['salary_advances'] > 0:
                expense_data.append([f"  • Salary Advances", f"{context['salary_advances']:,.2f}"])
            if context['staff_loans'] > 0:
                expense_data.append([f"  • Staff Loans", f"{context['staff_loans']:,.2f}"])
            if context['purchase_advances'] > 0:
                expense_data.append([f"  • Purchase Advances", f"{context['purchase_advances']:,.2f}"])
            expense_data.append(['', f"{context['total_advances']:,.2f}"])

        # Total Operating Expenses
        expense_data.append(['', ''])
        expense_data.append(['Total Operating Expenses', f"{context['total_operating_expenses']:,.2f}"])

        # Total Expenses
        expense_data.append(['', ''])
        expense_data.append(['TOTAL EXPENSES (COGS + Operating)', f"{context['total_expenses']:,.2f}"])

        # Create expense table
        expense_table = Table(expense_data, colWidths=[4 * inch, 2 * inch])
        expense_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data rows
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),

            # Total row
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e74c3c')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#2c3e50')),
        ]))

        elements.append(expense_table)
        elements.append(Spacer(1, 0.3 * inch))

        # ====================================================================
        # NET INCOME
        # ====================================================================

        net_income_color = colors.HexColor('#27ae60') if context['net_income'] >= 0 else colors.HexColor('#e74c3c')
        net_income_text = "NET INCOME" if context['net_income'] >= 0 else "NET LOSS"

        net_data = [
            [net_income_text, f"₦{context['net_income']:,.2f}"]
        ]

        net_table = Table(net_data, colWidths=[4 * inch, 2 * inch])
        net_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), net_income_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('TOPPADDING', (0, 0), (-1, 0), 15),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#2c3e50')),
        ]))

        elements.append(net_table)

        # Footer note
        elements.append(Spacer(1, 0.2 * inch))
        note_style = ParagraphStyle(
            'Note',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#7f8c8d'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(
            "Note: Sales revenue excludes wallet-based payments to avoid double-counting with wallet funding.",
            note_style
        ))

        from django.utils import timezone
        elements.append(Paragraph(
            f"Generated on {timezone.now().strftime('%B %d, %Y at %H:%M')}",
            note_style
        ))

        # Build PDF
        doc.build(elements)

        # Return response
        pdf_content = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf_content, content_type='application/pdf')
        filename = f"income_expense_report_{context['date_from']}_{context['date_to']}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except ImportError:
        messages.error(request, 'PDF generation requires reportlab. Please install it.')
        return redirect('income_expense_report')


@require_http_methods(["GET"])
def get_invoice_items_json(request, invoice_id):
    """
    AJAX endpoint to get invoice items for itemized payment selection.
    Returns JSON data with item details including discounts.
    """
    try:
        # Get the parent's selected ward from session
        parent = request.user.parent_profile.parent
        ward_id = request.session.get('selected_ward_id')

        if not ward_id:
            return JsonResponse({'error': 'No ward selected'}, status=400)

        ward = parent.wards.get(pk=ward_id)

        # Get the invoice and verify it belongs to this ward
        invoice = InvoiceModel.objects.get(pk=invoice_id, student=ward)

        # Build items data
        items_data = []
        for item in invoice.items.all():
            items_data.append({
                'id': item.pk,
                'description': item.description,
                'amount': str(item.amount),
                'total_discount': str(item.total_discount),
                'amount_after_discount': str(item.amount_after_discount),
                'amount_paid': str(item.amount_paid),
                'balance': str(item.balance),
                'paid_by_sibling': bool(item.paid_by_sibling),
                'sibling_name': item.paid_by_sibling.first_name if item.paid_by_sibling else None
            })

        return JsonResponse({
            'success': True,
            'items': items_data,
            'invoice': {
                'number': invoice.invoice_number,
                'total_amount': str(invoice.total_amount),
                'total_discount': str(invoice.total_discount),
                'amount_after_discount': str(invoice.amount_after_discount),
                'balance': str(invoice.balance)
            }
        })

    except InvoiceModel.DoesNotExist:
        return JsonResponse({'error': 'Invoice not found or does not belong to your ward'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@permission_required("finance.change_feepaymentmodel", raise_exception=True)
def payment_cleanup_view(request):
    """
    Renders a page showing classes that have students with confirmed payments
    that are missing item_breakdown. The frontend will POST each class id to
    process them gradually via AJAX.
    """
    # Get current session and term info (follow your project's pattern)
    school_setting = SchoolSettingModel.objects.first()

    classes_with_payments = []

    if school_setting and school_setting.session and school_setting.term:
        # Query classes that have confirmed payments missing breakdown.
        # Note: JSONField lookups for empty dict vary by DB backend; this is a pragmatic approach:
        # find students who have at least one confirmed payment whose item_breakdown is null or empty.
        payments_qs = FeePaymentModel.objects.filter(
            status=FeePaymentModel.PaymentStatus.CONFIRMED
        ).filter(
            Q(item_breakdown__isnull=True) | Q(item_breakdown={})
        ).select_related('invoice__student', 'invoice')

        # Build a mapping of class_id -> counts
        class_counts = {}
        student_to_class = {}

        for p in payments_qs:
            invoice = getattr(p, 'invoice', None)
            if not invoice or not getattr(invoice, 'student', None):
                continue
            student = invoice.student
            student_class = getattr(student, 'student_class', None)
            if not student_class:
                continue
            cid = student_class.pk
            cname = getattr(student_class, 'name', str(cid))
            class_counts.setdefault((cid, cname), 0)
            class_counts[(cid, cname)] += 1

        # Convert to list suitable for template
        for (cid, cname), count in class_counts.items():
            classes_with_payments.append({
                'id': cid,
                'name': cname,
                'count': count
            })

        classes_with_payments.sort(key=lambda x: x['name'])

    context = {
        'school_setting': school_setting,
        'classes_with_payments': classes_with_payments,
        'total_payments': sum(c['count'] for c in classes_with_payments),
    }
    return render(request, 'finance/payment_cleanup.html', context)


@require_POST
@login_required
@permission_required("finance.change_feepaymentmodel", raise_exception=True)
def process_payment_cleanup_for_class(request):
    """
    AJAX endpoint. Processes one class at a time: finds confirmed payments
    without item_breakdown for invoices in the current session/term and writes
    a JSON mapping of {item_id: amount_allocated} to payment.item_breakdown.
    """
    class_id = request.POST.get('class_id')

    try:
        school_setting = SchoolSettingModel.objects.first()
        if not school_setting or not school_setting.session or not school_setting.term:
            return JsonResponse({'status': 'error', 'message': 'No active session/term'}, status=400)

        if not class_id:
            return JsonResponse({'status': 'error', 'message': 'class_id is required'}, status=400)

        # Students in the class
        students = StudentModel.objects.filter(student_class_id=class_id)

        cleaned_payments = 0
        skipped_payments = 0
        total_payments_scanned = 0

        # We'll process per student to avoid giant transactions; but group in a single transaction per class
        with transaction.atomic():
            for student in students:
                # Get invoices for this student in the active session/term
                invoices = InvoiceModel.objects.filter(
                    student=student,
                    session=school_setting.session,
                    term=school_setting.term
                )

                for invoice in invoices:
                    # Get confirmed payments for this invoice that lack breakdown
                    payments = FeePaymentModel.objects.filter(
                        invoice=invoice,
                        status=FeePaymentModel.PaymentStatus.CONFIRMED
                    ).filter(
                        Q(item_breakdown__isnull=True) | Q(item_breakdown={})
                    ).order_by('created_at', 'date')  # process oldest first

                    for payment in payments:
                        total_payments_scanned += 1

                        # Defensive: if payment already has non-empty breakdown, skip
                        if payment.item_breakdown and isinstance(payment.item_breakdown, dict) and len(payment.item_breakdown) > 0:
                            skipped_payments += 1
                            continue

                        remaining_payment_amount = Decimal(payment.amount or 0)
                        breakdown = {}

                        # Iterate invoice items in a deterministic order (pk)
                        invoice_items = InvoiceItemModel.objects.filter(invoice=invoice).order_by('pk').select_related('fee_master', 'fee_master__fee')

                        for item in invoice_items:
                            if remaining_payment_amount <= Decimal('0.00'):
                                break

                            # remaining on item = amount - amount_paid (don't go negative)
                            item_amount = Decimal(item.amount or 0)
                            item_paid = Decimal(item.amount_paid or 0)
                            remaining_item = item_amount - item_paid
                            if remaining_item <= Decimal('0.00'):
                                # nothing to allocate to this item
                                continue

                            alloc = min(remaining_item, remaining_payment_amount)
                            if alloc > Decimal('0.00'):
                                breakdown[str(item.pk)] = str(alloc.quantize(Decimal('0.01')))
                                remaining_payment_amount -= alloc

                                # NOTE: we're NOT mutating item.amount_paid here. We only save payment.item_breakdown.
                                # If you want to also update item.amount_paid or sibling flags, tell me and
                                # I'll extend this to safely replay/pay them.

                        # If we allocated nothing, but the payment has an amount, try a fallback:
                        if not breakdown and Decimal(payment.amount or 0) > Decimal('0.00'):
                            # Edge-case: invoice items all show fully paid (maybe data drift).
                            # Fallback: attach payment fully to the first invoice item (as last resort).
                            first_item = invoice_items.first()
                            if first_item:
                                breakdown[str(first_item.pk)] = str(Decimal(payment.amount).quantize(Decimal('0.01')))

                        # Save breakdown if we produced one
                        if breakdown:
                            payment.item_breakdown = breakdown
                            payment.save(update_fields=['item_breakdown'])
                            cleaned_payments += 1
                        else:
                            skipped_payments += 1

        return JsonResponse({
            'status': 'success',
            'class_id': class_id,
            'processed_payments': cleaned_payments,
            'skipped_payments': skipped_payments,
            'total_scanned': total_payments_scanned
        })

    except Exception as e:
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': f'Unexpected error: {str(e)}'
        }, status=500)


# ============================================================================
# GENERAL OTHER PAYMENT VIEWS (All students)
# ============================================================================

class OtherPaymentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """View all other payments across all students"""
    model = OtherPaymentModel
    permission_required = 'finance.view_feemodel'
    template_name = 'finance/other_payment/list.html'
    context_object_name = 'other_payments'
    paginate_by = 50

    def get_queryset(self):
        queryset = OtherPaymentModel.objects.select_related(
            'student', 'session', 'term', 'created_by'
        ).order_by('-created_at')

        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # Filter by session
        session_id = self.request.GET.get('session')
        if session_id:
            queryset = queryset.filter(session_id=session_id)

        # Filter by term
        term_id = self.request.GET.get('term')
        if term_id:
            queryset = queryset.filter(term_id=term_id)

        # Filter by category
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category=category)

        # Search by student name or registration number
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(student__first_name__icontains=search) |
                Q(student__last_name__icontains=search) |
                Q(student__registration_number__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sessions'] = SessionModel.objects.all().order_by('-start_year')
        context['terms'] = TermModel.objects.all().order_by('id')
        context['total_outstanding'] = sum(op.balance for op in self.get_queryset())
        return context


# ============================================================================
# STUDENT-SPECIFIC OTHER PAYMENT VIEWS
# ============================================================================

class StudentOtherPaymentIndexView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """View all other payments for a specific student"""
    model = OtherPaymentModel
    permission_required = 'finance.view_feemodel'
    template_name = 'finance/other_payment/student_index.html'
    context_object_name = 'other_payments'

    def get_student(self):
        return get_object_or_404(StudentModel, pk=self.kwargs['student_pk'])

    def get_queryset(self):
        student = self.get_student()
        return OtherPaymentModel.objects.filter(student=student).select_related(
            'session', 'term', 'created_by'
        ).order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_student()
        context['student'] = student

        # Calculate totals
        other_payments = self.get_queryset()
        context['total_amount'] = sum(op.amount for op in other_payments)
        context['total_paid'] = sum(op.amount_paid for op in other_payments)
        context['total_balance'] = sum(op.balance for op in other_payments)

        return context


class StudentOtherPaymentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new other payment/debt for a student"""
    model = OtherPaymentModel
    form_class = OtherPaymentCreateForm
    permission_required = 'finance.add_feemodel'
    template_name = 'finance/other_payment/create.html'

    def get_student(self):
        return get_object_or_404(StudentModel, pk=self.kwargs['student_pk'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student'] = self.get_student()
        return context

    def form_valid(self, form):
        student = self.get_student()
        form.instance.student = student
        form.instance.created_by = self.request.user

        # Set initial status to unpaid
        form.instance.amount_paid = Decimal('0.00')
        form.instance.status = OtherPaymentModel.Status.UNPAID

        messages.success(
            self.request,
            f"Other payment/debt of {form.instance.amount:,.2f} "
            f"created successfully for {student.first_name} {student.last_name}."
        )

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('finance_student_other_payment_index', kwargs={'student_pk': self.kwargs['student_pk']})

    @property
    def currency_symbol(self):
        return '₦' if self.object.currency == 'naira' else '$'


class StudentOtherPaymentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing other payment/debt"""
    model = OtherPaymentModel
    form_class = OtherPaymentCreateForm
    permission_required = 'finance.change_feemodel'
    template_name = 'finance/other_payment/update.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student'] = self.object.student
        return context

    def form_valid(self, form):
        messages.success(self.request, "Other payment/debt updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('finance_student_other_payment_index', kwargs={'student_pk': self.object.student.pk})


class StudentOtherPaymentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete an other payment/debt (only if no payments made)"""
    model = OtherPaymentModel
    permission_required = 'finance.delete_feemodel'
    template_name = 'finance/other_payment/delete.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student'] = self.object.student
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        # Check if any payments have been made
        if self.object.clearances.exists():
            messages.error(
                request,
                "Cannot delete this debt because payments have been made against it. "
                "Please revert all payments first."
            )
            return redirect('finance_student_other_payment_index', student_pk=self.object.student.pk)

        messages.warning(request, f"Other payment/debt deleted successfully.")
        return super().post(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('finance_student_other_payment_index', kwargs={'student_pk': self.object.student.pk})


# ============================================================================
# PAYMENT CLEARANCE VIEWS
# ============================================================================

class OtherPaymentClearanceCreateView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """Make a payment against an other payment debt"""
    form_class = OtherPaymentClearanceForm
    permission_required = 'finance.add_feemodel'
    template_name = 'finance/other_payment/pay.html'

    def get_other_payment(self):
        return get_object_or_404(OtherPaymentModel, pk=self.kwargs['other_payment_pk'])

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['other_payment'] = self.get_other_payment()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        other_payment = self.get_other_payment()
        context['other_payment'] = other_payment
        context['student'] = other_payment.student
        return context

    def form_valid(self, form):
        other_payment = self.get_other_payment()

        with transaction.atomic():
            # Create the clearance record
            clearance = form.save(commit=False)
            clearance.other_payment = other_payment
            clearance.status = OtherPaymentClearanceModel.PaymentStatus.CONFIRMED
            clearance.confirmed_by = self.request.user
            clearance.save()

            # Update the other payment's amount_paid
            other_payment.amount_paid += clearance.amount
            other_payment.save()  # This will auto-update status

        messages.success(
            self.request,
            f"Payment of {clearance.currency_symbol}{clearance.amount:,.2f} recorded successfully. "
            f"Remaining balance: {other_payment.balance:,.2f}"
        )

        return redirect('finance_student_other_payment_index', student_pk=other_payment.student.pk)


class OtherPaymentClearanceRevertView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Revert a payment clearance"""
    model = OtherPaymentClearanceModel
    permission_required = 'finance.delete_feemodel'
    template_name = 'finance/other_payment/revert.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['other_payment'] = self.object.other_payment
        context['student'] = self.object.other_payment.student
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        clearance = self.object
        other_payment = clearance.other_payment

        # Check if already reverted
        if clearance.status == OtherPaymentClearanceModel.PaymentStatus.REVERTED:
            messages.warning(request, "This payment has already been reverted.")
            return redirect('finance_student_other_payment_index', student_pk=other_payment.student.pk)

        with transaction.atomic():
            # Mark clearance as reverted
            clearance.status = OtherPaymentClearanceModel.PaymentStatus.REVERTED
            clearance.save()

            # Reduce the other payment's amount_paid
            other_payment.amount_paid -= clearance.amount
            if other_payment.amount_paid < 0:
                other_payment.amount_paid = Decimal('0.00')
            other_payment.save()  # This will auto-update status

        messages.warning(
            request,
            f"Payment of {clearance.currency_symbol}{clearance.amount:,.2f} has been reverted. "
            f"New balance: {other_payment.balance:,.2f}"
        )

        return redirect('finance_student_other_payment_index', student_pk=other_payment.student.pk)


# ============================================================================
# SALARY SETTINGS VIEWS
# ============================================================================

@login_required
@permission_required('finance.view_salarysetting', raise_exception=True)
def salary_setting_list_view(request):
    """List all salary settings"""
    settings = SalarySetting.objects.all().order_by('-is_active', '-effective_from')

    context = {
        'settings': settings,
        'page_title': 'Salary Settings'
    }
    return render(request, 'finance/salary_setting/list.html', context)


@login_required
@permission_required('finance.add_salarysetting', raise_exception=True)
def salary_setting_create_view(request):
    """Create new salary setting"""
    if request.method == 'POST':
        form = SalarySettingForm(request.POST)
        if form.is_valid():
            setting = form.save(commit=False)
            setting.created_by = request.user
            setting.save()

            messages.success(request, f'Salary setting "{setting.name}" created successfully!')
            return redirect('finance_salary_setting_detail', pk=setting.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SalarySettingForm()

    context = {
        'form': form,
        'page_title': 'Create Salary Setting',
        'action': 'Create'
    }
    return render(request, 'finance/salary_setting/form.html', context)


@login_required
@permission_required('finance.view_salarysetting', raise_exception=True)
def salary_setting_detail_view(request, pk):
    """View and manage salary setting"""
    setting = get_object_or_404(SalarySetting, pk=pk)

    # Handle actions
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'activate' and not setting.is_active:
            if request.user.has_perm('finance.change_salarysetting'):
                setting.is_active = True
                setting.save()
                messages.success(request, f'"{setting.name}" is now active.')
            else:
                messages.error(request, 'You do not have permission to activate settings.')

        elif action == 'deactivate' and setting.is_active:
            if request.user.has_perm('finance.change_salarysetting'):
                setting.is_active = False
                setting.save()
                messages.success(request, f'"{setting.name}" has been deactivated.')
            else:
                messages.error(request, 'You do not have permission to deactivate settings.')

        elif action == 'edit' and not setting.is_locked:
            return redirect('finance_salary_setting_update', pk=pk)

        return redirect('finance_salary_setting_detail', pk=pk)

    # Get usage stats
    usage_stats = {
        'structures_count': setting.staff_structures.count(),
        'records_count': setting.salary_records.count(),
    }

    # Serialize JSONFields for safe use in the template (JS will parse these <script> blocks)
    # default=str is used to safely serialize Decimal/date values if present
    basic_components_json = json.dumps(setting.basic_components or {}, default=str)
    allowances_json = json.dumps(setting.allowances or [], default=str)
    reliefs_json = json.dumps(setting.reliefs_exemptions or [], default=str)
    tax_brackets_json = json.dumps(setting.tax_brackets or [], default=str)
    income_items_json = json.dumps(setting.income_items or [], default=str)
    statutory_deductions_json = json.dumps(setting.statutory_deductions or [], default=str)
    other_deductions_json = json.dumps(setting.other_deductions_config or [], default=str)

    context = {
        'setting': setting,
        'usage_stats': usage_stats,
        'page_title': f'Salary Setting: {setting.name}',
        # serialized JSON strings for the template
        'basic_components_json': basic_components_json,
        'allowances_json': allowances_json,
        'reliefs_json': reliefs_json,
        'tax_brackets_json': tax_brackets_json,
        'income_items_json': income_items_json,
        'statutory_deductions_json': statutory_deductions_json,
        'other_deductions_json': other_deductions_json,
        # optional helper (handy in templates if you prefer)
        'can_change': request.user.has_perm('finance.change_salarysetting'),
    }
    return render(request, 'finance/salary_setting/detail.html', context)

@login_required
@permission_required('finance.change_salarysetting', raise_exception=True)
def salary_setting_update_view(request, pk):
    """Update salary setting (only if not locked)"""
    setting = get_object_or_404(SalarySetting, pk=pk)

    if setting.is_locked:
        messages.error(request, 'Cannot edit a locked salary setting.')
        return redirect('finance_salary_setting_detail', pk=pk)

    if request.method == 'POST':
        form = SalarySettingForm(request.POST, instance=setting)
        if form.is_valid():
            form.save()
            messages.success(request, f'Salary setting "{setting.name}" updated successfully!')
            return redirect('finance_salary_setting_detail', pk=setting.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SalarySettingForm(instance=setting)

    context = {
        'form': form,
        'setting': setting,
        'page_title': f'Edit: {setting.name}',
        'action': 'Update'
    }
    return render(request, 'finance/salary_setting/form.html', context)


# ============================================================================
# SALARY STRUCTURE VIEWS
# ============================================================================

@login_required
@permission_required('finance.view_salarystructure', raise_exception=True)
def salary_structure_list_view(request):
    """List all salary structures"""
    structures = SalaryStructure.objects.select_related(
        'staff__staff_profile__user', 'salary_setting'
    ).filter(is_active=True).order_by('staff__staff_id')

    # Search filter
    search = request.GET.get('search', '')
    if search:
        structures = structures.filter(
            Q(staff__staff_id__icontains=search) |
            Q(staff__staff_profile__user__first_name__icontains=search) |
            Q(staff__staff_profile__user__last_name__icontains=search)
        )

    context = {
        'structures': structures,
        'search': search,
        'page_title': 'Staff Salary Structures'
    }
    return render(request, 'finance/salary_structure/list.html', context)


@login_required
@permission_required('finance.add_salarystructure', raise_exception=True)
def salary_structure_create_view(request):
    """Create salary structure for staff"""
    if request.method == 'POST':
        form = SalaryStructureForm(request.POST)
        if form.is_valid():
            structure = form.save()
            messages.success(
                request,
                f'Salary structure created for {structure.staff} - ₦{structure.monthly_salary:,.2f}/month'
            )
            return redirect('finance_salary_structure_detail', pk=structure.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SalaryStructureForm()

    # Get all active salary settings with their configurations for preview
    salary_settings = SalarySetting.objects.filter(is_active=True)
    salary_settings_json = {}
    for setting in salary_settings:
        salary_settings_json[str(setting.id)] = {
            'id': setting.id,
            'name': setting.name,
            'basic_components': setting.basic_components,
            'allowances': setting.allowances,
            'leave_allowance_percentage': float(setting.leave_allowance_percentage),  # Add this
            'include_leave_in_gross': setting.include_leave_in_gross,  # Add this
            'reliefs_exemptions': setting.reliefs_exemptions,  # Add this
            'tax_brackets': setting.tax_brackets,  # Add this
            'statutory_deductions': setting.statutory_deductions,  # Add this
        }

    context = {
        'form': form,
        'salary_settings_json': json.dumps(salary_settings_json),
        'page_title': 'Create Salary Structure',
        'action': 'Create'
    }
    return render(request, 'finance/salary_structure/form.html', context)


@login_required
@permission_required('finance.change_salarystructure', raise_exception=True)
def salary_structure_update_view(request, pk):
    """Update salary structure"""
    structure = get_object_or_404(SalaryStructure, pk=pk)

    if request.method == 'POST':
        form = SalaryStructureForm(request.POST, instance=structure)
        if form.is_valid():
            form.save()
            messages.success(request, f'Salary structure updated for {structure.staff}')
            return redirect('finance_salary_structure_detail', pk=structure.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SalaryStructureForm(instance=structure)

    # Get all active salary settings with their configurations for preview
    salary_settings = SalarySetting.objects.filter(is_active=True)
    salary_settings_json = {}
    for setting in salary_settings:
        salary_settings_json[str(setting.id)] = {
            'id': setting.id,
            'name': setting.name,
            'basic_components': setting.basic_components,
            'allowances': setting.allowances,
            'leave_allowance_percentage': float(setting.leave_allowance_percentage),  # Add this
            'include_leave_in_gross': setting.include_leave_in_gross,  # Add this
            'reliefs_exemptions': setting.reliefs_exemptions,  # Add this
            'tax_brackets': setting.tax_brackets,  # Add this
            'statutory_deductions': setting.statutory_deductions,  # Add this
        }

    context = {
        'form': form,
        'structure': structure,
        'salary_settings_json': json.dumps(salary_settings_json),
        'page_title': f'Edit Salary: {structure.staff}',
        'action': 'Update'
    }
    return render(request, 'finance/salary_structure/form.html', context)


@login_required
@permission_required('finance.view_salarystructure', raise_exception=True)
def salary_structure_detail_view(request, pk):
    """View salary structure details with complete breakdown"""
    structure = get_object_or_404(
        SalaryStructure.objects.select_related(
            'staff__staff_profile__user',
            'salary_setting'
        ).prefetch_related(
            'staff__salary_structures__salary_setting'
        ),
        pk=pk
    )

    # Calculate complete salary breakdown
    calculation = calculate_salary_breakdown(structure)

    context = {
        'structure': structure,
        'calculation': calculation,
        'page_title': f'Salary Structure: {structure.staff}'
    }
    return render(request, 'finance/salary_structure/detail.html', context)


def calculate_salary_breakdown(structure):
    """
    Calculate complete salary breakdown for a salary structure.
    This mirrors the JavaScript calculation logic from the form.
    """
    monthly_salary = structure.monthly_salary
    annual_salary = monthly_salary * 12
    setting = structure.salary_setting

    # Round to 2 decimal places
    def round_decimal(value):
        return round(Decimal(str(value)), 2)

    # Calculate basic components
    basic_components = {}
    total_basic_percentage = Decimal('0')

    for key, component in setting.basic_components.items():
        percentage = Decimal(str(component.get('percentage', 0)))
        monthly_amount = round_decimal((monthly_salary * percentage) / Decimal('100'))
        annual_amount = round_decimal(monthly_amount * 12)

        basic_components[component['code']] = {
            'name': component['name'],
            'percentage': percentage,
            'monthly': monthly_amount,
            'annual': annual_amount
        }
        total_basic_percentage += percentage

    # Calculate leave allowance
    leave_allowance_percentage = Decimal(str(setting.leave_allowance_percentage or 0))
    annual_basic_salary = sum(comp['annual'] for comp in basic_components.values())
    leave_allowance_annual = round_decimal((annual_basic_salary * leave_allowance_percentage) / Decimal('100'))
    leave_allowance_monthly = round_decimal(leave_allowance_annual / 12)

    # Helper function to calculate amount based on component codes
    def calculate_based_on_components(based_on):
        if not based_on:
            return Decimal('0')

        if based_on.upper() == 'TOTAL':
            return sum(comp['monthly'] for comp in basic_components.values()) * 12

        if based_on.upper() == 'GROSS_INCOME':
            return Decimal('0')  # Will be handled separately

        # Handle component codes like "B+H+T"
        codes = [code.strip().upper() for code in based_on.split('+')]
        total = Decimal('0')
        for code in codes:
            for comp_code, comp_data in basic_components.items():
                if comp_code.upper() == code:
                    total += comp_data['monthly']
        return total * 12

    # Calculate other allowances
    allowances = []
    total_other_allowances_monthly = Decimal('0')
    total_other_allowances_annual = Decimal('0')

    for allowance in setting.allowances:
        if allowance.get('is_active', False):
            monthly_amount = Decimal('0')

            if allowance.get('calculation_type') == 'fixed':
                monthly_amount = round_decimal(Decimal(str(allowance.get('fixed_amount', 0))))

            elif allowance.get('calculation_type') == 'percentage':
                percentage = Decimal(str(allowance.get('percentage', 0)))
                based_on = allowance.get('based_on', 'TOTAL')

                if based_on.upper() == 'TOTAL':
                    monthly_amount = round_decimal((monthly_salary * percentage) / Decimal('100'))
                elif based_on.upper() == 'GROSS_INCOME':
                    temp_gross = monthly_salary + total_other_allowances_monthly
                    monthly_amount = round_decimal((temp_gross * percentage) / Decimal('100'))
                else:
                    base_amount = calculate_based_on_components(based_on)
                    monthly_amount = round_decimal((base_amount * percentage) / Decimal('100'))

            annual_amount = round_decimal(monthly_amount * 12)

            allowances.append({
                'name': allowance['name'],
                'monthly': monthly_amount,
                'annual': annual_amount,
                'percentage': round_decimal((monthly_amount / monthly_salary * 100) if monthly_salary > 0 else 0)
            })

            total_other_allowances_monthly += monthly_amount
            total_other_allowances_annual += annual_amount

    # Calculate gross income based on include_leave_in_gross setting
    if setting.include_leave_in_gross:
        gross_income_monthly = round_decimal(monthly_salary + total_other_allowances_monthly + leave_allowance_monthly)
        gross_income_annual = round_decimal(annual_salary + total_other_allowances_annual + leave_allowance_annual)
    else:
        gross_income_monthly = round_decimal(monthly_salary + total_other_allowances_monthly)
        gross_income_annual = round_decimal(annual_salary + total_other_allowances_annual)

    # Calculate statutory deductions
    statutory_deductions = []
    total_statutory_deductions = Decimal('0')
    total_reliefs = Decimal('0')

    for deduction in setting.statutory_deductions:
        if deduction.get('is_active', True):
            percentage = Decimal(str(deduction.get('percentage', 0)))
            base_amount = calculate_based_on_components(deduction.get('based_on', ''))
            amount = round_decimal((base_amount * percentage) / Decimal('100'))

            statutory_deductions.append({
                'name': deduction['name'],
                'percentage': percentage,
                'monthly': amount,
                'annual': round_decimal(amount * 12)
            })
            total_statutory_deductions += amount
            total_reliefs += amount
    # Calculate reliefs and exemptions
    reliefs = []

    for relief in setting.reliefs_exemptions:
        amount = Decimal('0')

        # Determine base amount
        base_amount = Decimal('0')
        based_on = relief.get('based_on', '')

        if based_on.upper() == 'GROSS_INCOME':
            base_amount = gross_income_annual
        elif based_on:
            base_amount = calculate_based_on_components(based_on)
        else:
            base_amount = gross_income_annual

        # Calculate relief amount
        if relief.get('formula_type') == 'percentage_plus_fixed':
            if relief.get('percentage'):
                amount = round_decimal((base_amount * Decimal(str(relief['percentage']))) / Decimal('100'))
            if relief.get('fixed_amount'):
                amount += round_decimal(Decimal(str(relief['fixed_amount'])))
        else:
            if relief.get('percentage'):
                amount = round_decimal((base_amount * Decimal(str(relief['percentage']))) / Decimal('100'))
            if relief.get('fixed_amount'):
                amount += round_decimal(Decimal(str(relief['fixed_amount'])))

        reliefs.append({
            'name': relief['name'],
            'amount': amount,
            'percentage': relief.get('percentage'),
            'fixed_amount': relief.get('fixed_amount')
        })
        total_reliefs += amount

    # Calculate taxable income
    taxable_income = round_decimal(gross_income_annual - total_reliefs)

    # Calculate PAYE tax using brackets
    annual_tax = Decimal('0')
    tax_breakdown = []
    remaining_income = taxable_income

    for index, bracket in enumerate(setting.tax_brackets):
        if remaining_income > 0:
            bracket_limit = bracket.get('limit')
            bracket_size = Decimal(str(bracket_limit)) if bracket_limit is not None else remaining_income

            taxable_amount = min(remaining_income, bracket_size)
            tax_rate = Decimal(str(bracket.get('rate', 0)))
            tax_amount = round_decimal((taxable_amount * tax_rate) / Decimal('100'))

            if taxable_amount > 0:
                if index == 0:
                    description = f"First {taxable_amount:,.2f}"
                elif bracket_limit is None:
                    description = f"Remaining {taxable_amount:,.2f}"
                else:
                    description = f"Next {bracket_size:,.2f}"

                tax_breakdown.append({
                    'description': description,
                    'rate': tax_rate,
                    'amount': tax_amount
                })

                annual_tax += tax_amount
                remaining_income -= taxable_amount

    monthly_tax = round_decimal(annual_tax / 12)

    # Calculate net salary and effective tax rate
    net_salary = round_decimal(gross_income_monthly - monthly_tax)
    effective_tax_rate = round_decimal((monthly_tax / gross_income_monthly * 100) if gross_income_monthly > 0 else 0)

    return {
        'monthly_salary': float(monthly_salary),
        'annual_salary': float(annual_salary),
        'basic_components': {k: {
            'name': v['name'],
            'percentage': float(v['percentage']),
            'monthly': float(v['monthly']),
            'annual': float(v['annual'])
        } for k, v in basic_components.items()},
        'leave_allowance_percentage': float(leave_allowance_percentage),
        'leave_allowance_monthly': float(leave_allowance_monthly),
        'leave_allowance_annual': float(leave_allowance_annual),
        'allowances': [{
            'name': a['name'],
            'monthly': float(a['monthly']),
            'annual': float(a['annual']),
            'percentage': float(a['percentage'])
        } for a in allowances],
        'gross_income_monthly': float(gross_income_monthly),
        'gross_income_annual': float(gross_income_annual),
        'statutory_deductions': [{
            'name': s['name'],
            'percentage': float(s['percentage']),
            'monthly': float(s['monthly'])/12,
            'annual': float(s['annual'])/12
        } for s in statutory_deductions],
        'total_statutory_deductions': float(total_statutory_deductions),
        'reliefs': [{
            'name': r['name'],
            'amount': float(r['amount']),
            'percentage': r.get('percentage'),
            'fixed_amount': r.get('fixed_amount')
        } for r in reliefs],
        'total_reliefs': float(total_reliefs),
        'taxable_income': float(taxable_income),
        'tax_breakdown': [{
            'description': t['description'],
            'rate': float(t['rate']),
            'amount': float(t['amount'])
        } for t in tax_breakdown],
        'annual_tax': float(annual_tax),
        'monthly_tax': float(monthly_tax),
        'net_salary': float(net_salary),
        'effective_tax_rate': float(effective_tax_rate)
    }


@login_required
@permission_required('finance.view_salaryrecord', raise_exception=True)
def payroll_view(request):
    """Main payroll dashboard with salary structures list"""
    today = datetime.now()
    current_year = int(request.GET.get('year', today.year))
    current_month = int(request.GET.get('month', today.month))
    status_filter = request.GET.get('status', 'all')
    search_query = request.GET.get('search', '')

    # Get all active salary structures
    structures = SalaryStructure.objects.filter(is_active=True).select_related(
        'staff__staff_profile__user', 'salary_setting', 'staff__department'
    )

    # Apply search filter if provided
    if search_query:
        structures = structures.filter(
            Q(staff__first_name__icontains=search_query) |
            Q(staff__last_name__icontains=search_query)
        )

    # Order by department name, then by staff name
    structures = structures.order_by('staff__department__name', 'staff__first_name', 'staff__last_name')

    # Check which structures have been processed for the selected month
    processed_ids = SalaryRecord.objects.filter(
        year=current_year,
        month=current_month
    ).values_list('salary_structure_id', flat=True)

    # Annotate structures with processing status
    for structure in structures:
        structure.is_processed = structure.id in processed_ids
        if structure.is_processed:
            record = SalaryRecord.objects.get(
                salary_structure=structure,
                year=current_year,
                month=current_month
            )
            structure.payment_status = record.payment_status
        else:
            structure.payment_status = 'not_processed'

    # Apply status filter
    if status_filter != 'all':
        if status_filter == 'processed':
            structures = [s for s in structures if s.is_processed]
        elif status_filter == 'pending':
            structures = [s for s in structures if s.is_processed and s.payment_status == 'pending']
        elif status_filter == 'paid':
            structures = [s for s in structures if s.is_processed and s.payment_status == 'paid']
        elif status_filter == 'not_paid':
            structures = [s for s in structures if
                          not s.is_processed or (s.is_processed and s.payment_status != 'paid')]

    import calendar
    context = {
        'structures': structures,
        'current_year': current_year,
        'current_month': current_month,
        'month_name': calendar.month_name[current_month],
        'status_filter': status_filter,
        'search_query': search_query,  # Add search query to context
        'years': range(2020, today.year + 2),
        'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'page_title': f'Payroll - {calendar.month_name[current_month]} {current_year}'
    }
    return render(request, 'finance/payroll/payroll.html', context)


@login_required
@permission_required('finance.add_salaryrecord', raise_exception=True)
def process_payroll_view(request, structure_id):
    """Process individual staff payroll with payslip creation"""
    # Get salary structure
    structure = get_object_or_404(
        SalaryStructure.objects.select_related(
            'staff__staff_profile__user', 'salary_setting'
        ),
        pk=structure_id
    )

    today = datetime.now()
    current_year = int(request.GET.get('year', today.year))
    current_month = int(request.GET.get('month', today.month))

    # Check if already processed
    existing_record = SalaryRecord.objects.filter(
        salary_structure=structure,
        year=current_year,
        month=current_month
    ).first()

    # Parse JSON fields for template use
    additional_income = {}
    other_deductions = {}

    if existing_record:
        # Safely parse additional_income
        if existing_record.additional_income:
            try:
                if isinstance(existing_record.additional_income, str):
                    additional_income = json.loads(existing_record.additional_income)
                else:
                    additional_income = existing_record.additional_income
            except (json.JSONDecodeError, TypeError):
                additional_income = {}

        # Safely parse other_deductions
        if existing_record.other_deductions:
            try:
                if isinstance(existing_record.other_deductions, str):
                    other_deductions = json.loads(existing_record.other_deductions)
                else:
                    other_deductions = existing_record.other_deductions
            except (json.JSONDecodeError, TypeError):
                other_deductions = {}

    if request.method == 'POST':
        # Get form data
        bonus = Decimal(request.POST.get('bonus', '0'))
        notes = request.POST.get('notes', '')

        # Get allowances
        allowances = {}
        for allowance_config in structure.salary_setting.income_items:
            if allowance_config.get('is_active', True):
                allowance_name = allowance_config['name']
                allowance_value = request.POST.get(f'allowance_{allowance_name}', '0')
                if allowance_value:
                    allowances[allowance_name] = Decimal(allowance_value)

        # Get deductions
        deductions = {}
        for deduction_config in structure.salary_setting.other_deductions_config:
            if not deduction_config.get('linked_to'):  # Only manual deductions
                deduction_name = deduction_config['name']
                deduction_value = request.POST.get(f'deduction_{deduction_name}', '0')
                if deduction_value:
                    deductions[deduction_name] = Decimal(deduction_value)

        # Calculate salary using SalaryCalculator - BUT WITHOUT additional_income and custom_deductions
        calculator = SalaryCalculator(structure, current_month, current_year)
        salary_data = calculator.calculate_complete_salary(
            bonus=bonus,
            custom_deductions=deductions,  # ← PASS THE DEDUCTIONS
            additional_income=allowances  # ← PASS THE ALLOWANCES
        )

        # Helper function to convert Decimals in nested structures
        def convert_decimals(obj):
            if isinstance(obj, Decimal):
                return str(obj)
            elif isinstance(obj, dict):
                return {k: convert_decimals(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [convert_decimals(item) for item in obj]
            return obj

        # Calculate totals manually
        # total_additional_income = sum(allowances.values()) if allowances else Decimal('0')
        # total_other_deductions = sum(deductions.values()) if deductions else Decimal('0')
        #
        # # Calculate final values
        # gross_salary = salary_data.get('gross_salary', Decimal('0')) + total_additional_income
        # net_salary = gross_salary - salary_data.get('total_statutory_deductions', Decimal('0')) - salary_data.get(
        #     'monthly_tax', Decimal('0')) - total_other_deductions

        # Use calculated values directly
        gross_salary = salary_data['total_income']  # Already includes allowances
        net_salary = salary_data['net_salary']  # Already includes all deductions

        # Create or update salary record
        if existing_record:
            record = existing_record
            record.bonus = bonus
            record.notes = notes
            record.additional_income = convert_decimals(allowances)
            record.other_deductions = convert_decimals(deductions)
        else:
            record = SalaryRecord(
                staff=structure.staff,
                salary_structure=structure,
                salary_setting=structure.salary_setting,
                month=current_month,
                year=current_year,
                monthly_salary=structure.monthly_salary,
                annual_salary=structure.annual_salary,
                basic_components_breakdown=convert_decimals(salary_data.get('basic_components', {})),
                allowances_breakdown=convert_decimals(salary_data.get('allowances', {})),
                bonus=bonus,
                additional_income=convert_decimals(allowances),
                other_deductions=convert_decimals(deductions),
                notes=notes,
                created_by=request.user
            )

        # Update calculated fields
        record.total_income = gross_salary
        record.gross_salary = gross_salary
        record.statutory_deductions = convert_decimals(salary_data.get('statutory_deductions', {}))
        record.total_statutory_deductions = Decimal(salary_data.get('total_statutory_deductions', '0'))
        record.total_other_deductions = salary_data['total_other_deductions']
        record.annual_gross_income = Decimal(salary_data.get('annual_gross_income', '0'))
        record.total_reliefs = Decimal(salary_data.get('total_reliefs', '0'))
        record.taxable_income = Decimal(salary_data.get('taxable_income', '0'))
        record.annual_tax = Decimal(salary_data.get('annual_tax', '0'))
        record.monthly_tax = Decimal(salary_data.get('monthly_tax', '0'))
        record.total_taxation = Decimal(salary_data.get('monthly_tax', '0'))
        record.effective_tax_rate = Decimal(salary_data.get('effective_tax_rate', '0'))
        record.net_salary = net_salary

        record.save()

        messages.success(request, f'Payroll for {structure.staff} processed successfully!')
        return redirect(reverse('finance_salary_record_detail', kwargs={'pk': record.id}))

    # Prepare context for GET request
    context = {
        'structure': structure,
        'existing_record': existing_record,
        'additional_income': additional_income,
        'other_deductions': other_deductions,
        'current_year': current_year,
        'current_month': current_month,
        'month_name': calendar.month_name[current_month],
        'page_title': f'Process Payroll - {structure.staff}',
        # Pass salary setting data to JavaScript
        'salary_setting_data': json.dumps({
            'statutory_deductions': structure.salary_setting.statutory_deductions,
            'tax_brackets': structure.salary_setting.tax_brackets,
            'reliefs_exemptions': structure.salary_setting.reliefs_exemptions
        })
    }
    return render(request, 'finance/payroll/process.html', context)


@login_required
@permission_required('finance.view_salaryrecord', raise_exception=True)
def salary_record_detail_view(request, pk):
    """View salary record (payslip) - using SalaryCalculator for consistency"""
    import json
    import calendar
    from decimal import Decimal

    # Get the salary record
    record = get_object_or_404(
        SalaryRecord.objects.select_related(
            'staff__staff_profile__user',
            'salary_structure',
            'salary_setting'
        ),
        pk=pk
    )

    structure = record.salary_structure

    # Parse JSON fields for additional data
    def parse_json_field(field):
        if field and isinstance(field, str):
            try:
                return json.loads(field)
            except (json.JSONDecodeError, TypeError):
                return {}
        return field or {}

    additional_income = parse_json_field(record.additional_income)
    other_deductions = parse_json_field(record.other_deductions)

    # Use SalaryCalculator to recalculate (ensures consistency with process view)
    calculator = SalaryCalculator(structure, record.month, record.year)
    salary_data = calculator.calculate_complete_salary(
        bonus=record.bonus,
        custom_deductions=other_deductions,  # ← Pass the parsed deductions
        additional_income=additional_income  # ← Pass the parsed allowances

    )

    # Build income breakdown list from calculator results
    income_breakdown = []

    # Add basic components - CORRECTED KEY NAME
    basic_components = salary_data.get('basic_components_breakdown', {})
    if basic_components:
        for code, component in basic_components.items():
            if isinstance(component, dict):
                income_breakdown.append({
                    'name': component.get('name', code),
                    'amount': float(component.get('amount', 0)),
                    'percentage': float(component.get('percentage', 0)),
                    'type': 'basic'
                })

    # Add allowances - CORRECTED KEY NAME
    allowances = salary_data.get('allowances_breakdown', {})
    if allowances:
        for allowance_name, allowance_data in allowances.items():
            if isinstance(allowance_data, dict):
                amount = float(allowance_data.get('amount', 0))
                if amount > 0:
                    income_breakdown.append({
                        'name': allowance_name,
                        'amount': amount,
                        'percentage': 0,
                        'type': 'allowance'
                    })

    # Add bonus
    if record.bonus > 0:
        income_breakdown.append({
            'name': 'Bonus',
            'amount': float(record.bonus),
            'percentage': 0,
            'type': 'bonus'
        })

    # Add additional income
    if additional_income:
        for name, amount in additional_income.items():
            amount_val = float(amount) if amount else 0
            if amount_val > 0:
                income_breakdown.append({
                    'name': name,
                    'amount': amount_val,
                    'percentage': 0,
                    'type': 'additional'
                })

    # Build statutory deductions breakdown from calculator
    statutory_breakdown = []
    statutory_deductions = salary_data.get('statutory_deductions', {})

    if statutory_deductions:
        for name, deduction in statutory_deductions.items():
            if isinstance(deduction, dict):
                statutory_breakdown.append({
                    'name': name,
                    'amount': float(deduction.get('amount', 0)),
                    'percentage': float(deduction.get('percentage', 0)),
                    'based_on': deduction.get('based_on', '')
                })

    # Build other deductions breakdown
    other_breakdown = []
    other_deductions_calc = salary_data.get('other_deductions', {})
    if other_deductions_calc:
        for name, deduction in other_deductions_calc.items():
            if isinstance(deduction, dict):
                amount_val = float(deduction.get('amount', 0))
            else:
                amount_val = float(deduction) if deduction else 0

            if amount_val > 0:
                other_breakdown.append({
                    'name': name,
                    'amount': amount_val
                })

    # Get calculated totals from salary_data (from calculator)
    calculated_totals = {
        'total_income': float(salary_data.get('total_income', 0)),
        'total_statutory': float(salary_data.get('total_statutory_deductions', 0)),
        'total_other': float(salary_data.get('total_other_deductions', 0)),
        'total_tax': float(salary_data.get('total_taxation', 0)),
        'net_salary': float(salary_data.get('net_salary', 0))
    }

    # Prepare context
    context = {
        'record': record,
        'structure': structure,
        'income_breakdown': income_breakdown,
        'statutory_breakdown': statutory_breakdown,
        'other_breakdown': other_breakdown,
        'calculated_totals': calculated_totals,
        'page_title': f'Payslip - {record.staff} - {record.month_name} {record.year}',
        # Pass salary setting data for any JavaScript calculations
        'salary_setting_data': json.dumps({
            'basic_components': structure.salary_setting.basic_components,
            'statutory_deductions': structure.salary_setting.statutory_deductions,
            'tax_brackets': structure.salary_setting.tax_brackets,
            'reliefs_exemptions': structure.salary_setting.reliefs_exemptions
        })
    }

    return render(request, 'finance/payroll/record_detail.html', context)


@login_required
@permission_required('finance.view_salaryrecord', raise_exception=True)
def download_payslip_pdf(request, pk):
    """Generate and download payslip as PDF - Optimized for single page"""
    import json

    # Get the salary record
    record = get_object_or_404(
        SalaryRecord.objects.select_related(
            'staff__staff_profile__user',
            'salary_structure',
            'salary_setting'
        ),
        pk=pk
    )

    structure = record.salary_structure

    # Get school info
    school_info = SchoolInfoModel.objects.first()

    # Parse JSON fields
    def parse_json_field(field):
        if field and isinstance(field, str):
            try:
                return json.loads(field)
            except (json.JSONDecodeError, TypeError):
                return {}
        return field or {}

    additional_income = parse_json_field(record.additional_income)
    other_deductions = parse_json_field(record.other_deductions)

    # Recalculate using calculator
    calculator = SalaryCalculator(structure, record.month, record.year)
    salary_data = calculator.calculate_complete_salary(
        bonus=record.bonus,
        custom_deductions=other_deductions,
        additional_income=additional_income
    )

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response[
        'Content-Disposition'] = f'attachment; filename="payslip_{record.staff.staff_id}_{record.month}_{record.year}.pdf"'

    # Create the PDF object with tighter margins
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        topMargin=0.4 * inch,
        bottomMargin=0.4 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch
    )
    elements = []

    # Styles
    styles = getSampleStyleSheet()

    # School Info at the top (if available)
    if school_info:
        school_style = ParagraphStyle(
            'SchoolInfo',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#2c3e50'),
            alignment=TA_CENTER,
            spaceAfter=3
        )
        elements.append(Paragraph(f"<b>{school_info.name}</b>", school_style))
        elements.append(Paragraph(f"{school_info.mobile} | {school_info.email}", school_style))
        elements.append(Paragraph(f"{school_info.address}", school_style))
        elements.append(Spacer(1, 0.1 * inch))

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a237e'),
        spaceAfter=8,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("PAYSLIP", title_style))
    elements.append(Spacer(1, 0.1 * inch))

    # Staff Information (Compact)
    staff_data = [
        ['Staff Information', '', 'Payment Details', ''],
        ['Name:', str(record.staff), 'Period:', f'{record.month_name} {record.year}'],
        ['Staff ID:', record.staff.staff_id, 'Payment Date:', str(record.paid_date) if record.paid_date else 'Pending'],
        ['Department:', 'N/A', 'Status:', record.get_payment_status_display()],
    ]

    staff_table = Table(staff_data, colWidths=[1.3 * inch, 2 * inch, 1.3 * inch, 2 * inch])
    staff_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#e3f2fd')),
        ('BACKGROUND', (2, 0), (3, 0), colors.HexColor('#e3f2fd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(staff_table)
    elements.append(Spacer(1, 0.1 * inch))

    # Bank Details (if available) - Compact
    if structure.bank_name:
        bank_data = [[
            f"Bank: {structure.bank_name} | Account: {structure.account_number} | Name: {structure.account_name}"
        ]]
        bank_table = Table(bank_data, colWidths=[6.6 * inch])
        bank_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f5f5f5')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(bank_table)
        elements.append(Spacer(1, 0.1 * inch))

    # ONE BIG COMPREHENSIVE TABLE
    # Build all rows for the unified table
    table_data = []

    # Header
    table_data.append(['Description', 'Details', 'Amount (N)'])

    # INCOME SECTION
    table_data.append(['INCOME', '', ''])

    # Basic Salary Components
    basic_components = salary_data.get('basic_components_breakdown', {})
    for code, component in basic_components.items():
        if isinstance(component, dict):
            percentage = component.get('percentage', 0)
            amount = float(component.get('amount', 0))
            table_data.append([
                f"  {component.get('name', code)}",
                f"{percentage:.2f}%",
                f"{amount:,.2f}"
            ])

    # Allowances (from additional income breakdown)
    allowances = salary_data.get('allowances_breakdown', {})
    if allowances:
        for allowance_name, allowance_data in allowances.items():
            if isinstance(allowance_data, dict):
                amount = float(allowance_data.get('amount', 0))
                if amount > 0:
                    table_data.append([
                        f"  {allowance_name}",
                        'Allowance',
                        f"{amount:,.2f}"
                    ])

    # Bonus
    if record.bonus > 0:
        table_data.append([
            '  Bonus',
            '',
            f"{float(record.bonus):,.2f}"
        ])

    # Additional Income
    if additional_income:
        for name, amount in additional_income.items():
            amount_val = float(amount) if amount else 0
            if amount_val > 0:
                table_data.append([
                    f"  {name}",
                    'Additional',
                    f"{amount_val:,.2f}"
                ])

    # Total Income Row
    table_data.append([
        'TOTAL INCOME (A)',
        '',
        f"{float(salary_data['total_income']):,.2f}"
    ])

    # DEDUCTIONS SECTION
    table_data.append(['DEDUCTIONS', '', ''])

    # Statutory Deductions
    statutory_deductions = salary_data.get('statutory_deductions', {})
    for name, deduction in statutory_deductions.items():
        if isinstance(deduction, dict):
            percentage = deduction.get('percentage', 0)
            based_on = deduction.get('based_on', '')
            amount = float(deduction.get('amount', 0))

            details = ''
            if percentage > 0:
                details = f"{percentage:.2f}% of {based_on}"

            table_data.append([
                f"  {name}",
                details,
                f"{amount:,.2f}"
            ])

    # Other Deductions
    if other_deductions:
        for name, amount in other_deductions.items():
            amount_val = float(amount) if amount else 0
            if amount_val > 0:
                table_data.append([
                    f"  {name}",
                    'Other Deduction',
                    f"{amount_val:,.2f}"
                ])

    # Taxation
    table_data.append([
        '  PAYE Tax',
        'Monthly Tax',
        f"{float(record.monthly_tax):,.2f}"
    ])

    if record.other_taxes > 0:
        table_data.append([
            '  Other Taxes',
            '',
            f"{float(record.other_taxes):,.2f}"
        ])

    # Total Deductions Row
    total_deductions = (
            float(salary_data['total_statutory_deductions']) +
            float(salary_data['total_other_deductions']) +
            float(salary_data['total_taxation'])
    )
    table_data.append([
        'TOTAL DEDUCTIONS (B)',
        '',
        f"{total_deductions:,.2f}"
    ])

    # NET SALARY ROW (Final)
    table_data.append([
        'NET SALARY (A - B)',
        '',
        f"{float(salary_data['net_salary']):,.2f}"
    ])

    # Create the unified table
    unified_table = Table(table_data, colWidths=[3.2 * inch, 2 * inch, 1.4 * inch])

    # Apply comprehensive styling
    table_style = [
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (1, 0), 'LEFT'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),

        # All content
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]

    # Find section headers and total rows for special styling
    income_row = None
    deductions_row = None
    total_income_row = None
    total_deductions_row = None
    net_salary_row = None

    for idx, row in enumerate(table_data):
        if row[0] == 'INCOME':
            income_row = idx
            table_style.extend([
                ('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#e3f2fd')),
                ('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'),
                ('FONTSIZE', (0, idx), (-1, idx), 9),
            ])
        elif row[0] == 'DEDUCTIONS':
            deductions_row = idx
            table_style.extend([
                ('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#ffe0e0')),
                ('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'),
                ('FONTSIZE', (0, idx), (-1, idx), 9),
            ])
        elif row[0] == 'TOTAL INCOME (A)':
            total_income_row = idx
            table_style.extend([
                ('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#b3d9ff')),
                ('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'),
                ('FONTSIZE', (0, idx), (-1, idx), 9),
            ])
        elif row[0] == 'TOTAL DEDUCTIONS (B)':
            total_deductions_row = idx
            table_style.extend([
                ('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#ffb3b3')),
                ('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'),
                ('FONTSIZE', (0, idx), (-1, idx), 9),
            ])
        elif row[0] == 'NET SALARY (A - B)':
            net_salary_row = idx
            table_style.extend([
                ('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#4CAF50')),
                ('TEXTCOLOR', (0, idx), (-1, idx), colors.white),
                ('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'),
                ('FONTSIZE', (0, idx), (-1, idx), 10),
            ])

    unified_table.setStyle(TableStyle(table_style))
    elements.append(unified_table)

    # Notes (if any) - Compact
    if record.notes:
        elements.append(Spacer(1, 0.1 * inch))
        notes_style = ParagraphStyle(
            'Notes',
            parent=styles['Normal'],
            fontSize=8,
            leading=10
        )
        elements.append(Paragraph(f"<b>Notes:</b> {record.notes}", notes_style))

    # Build PDF
    doc.build(elements)

    return response


@login_required
@permission_required('finance.change_salaryrecord', raise_exception=True)
def mark_as_paid_view(request, pk):
    """Mark salary record as paid"""
    record = get_object_or_404(SalaryRecord, pk=pk)

    if request.method == 'POST':
        record.payment_status = SalaryRecord.PaymentStatus.PAID
        record.paid_date = date.today()
        record.paid_by = request.user
        if record.amount_paid == 0:
            record.amount_paid = record.net_salary
        record.save()

        messages.success(request, f'Salary for {record.staff} marked as paid!')
        return redirect('finance_salary_record_detail', pk=pk)

    context = {
        'record': record,
        'page_title': f'Mark as Paid - {record.staff}'
    }
    return render(request, 'finance/payroll/mark_as_paid.html', context)


@login_required
@permission_required('finance.view_salaryrecord', raise_exception=True)
def payroll_dashboard_view(request):
    """Payroll dashboard with monthly summary and statistics"""
    from .models import SalaryRecord, SalaryStructure

    # Get current or requested month/year
    today = datetime.now()
    current_year = int(request.GET.get('year', today.year))
    current_month = int(request.GET.get('month', today.month))

    # Get all records for the selected month
    records = SalaryRecord.objects.filter(
        year=current_year,
        month=current_month
    ).select_related('staff', 'salary_structure')

    # Calculate statistics
    total_staff = SalaryStructure.objects.filter(is_active=True).count()
    processed_count = records.count()
    unprocessed_count = total_staff - processed_count

    # Payment status breakdown
    paid_records = records.filter(payment_status=SalaryRecord.PaymentStatus.PAID)
    partially_paid_records = records.filter(payment_status=SalaryRecord.PaymentStatus.PARTIALLY_PAID)
    pending_records = records.filter(payment_status=SalaryRecord.PaymentStatus.PENDING)

    paid_count = paid_records.count()
    partially_paid_count = partially_paid_records.count()
    pending_count = pending_records.count()

    # Financial totals
    total_gross = records.aggregate(total=Sum('total_income'))['total'] or Decimal('0.00')
    total_net = records.aggregate(total=Sum('net_salary'))['total'] or Decimal('0.00')
    total_statutory = records.aggregate(total=Sum('total_statutory_deductions'))['total'] or Decimal('0.00')
    total_tax = records.aggregate(total=Sum('monthly_tax'))['total'] or Decimal('0.00')
    total_other_deductions = records.aggregate(total=Sum('total_other_deductions'))['total'] or Decimal('0.00')

    # Amount paid vs outstanding
    total_amount_paid = records.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
    total_outstanding = total_net - total_amount_paid

    # Average statistics
    avg_gross = records.aggregate(avg=Avg('total_income'))['avg'] or Decimal('0.00')
    avg_net = records.aggregate(avg=Avg('net_salary'))['avg'] or Decimal('0.00')
    avg_tax_rate = records.aggregate(avg=Avg('effective_tax_rate'))['avg'] or Decimal('0.00')

    # Top earners (by net salary)
    top_earners = records.order_by('-net_salary')[:5]

    # Recent payments
    recent_payments = records.filter(
        payment_status=SalaryRecord.PaymentStatus.PAID,
        paid_date__isnull=False
    ).order_by('-paid_date')[:10]

    # Staff without salary records (unprocessed)
    processed_staff_ids = records.values_list('staff_id', flat=True)
    unprocessed_staff = SalaryStructure.objects.filter(
        is_active=True
    ).exclude(staff_id__in=processed_staff_ids).select_related('staff')[:10]

    # Month navigation data
    months = [
        {'value': i, 'name': calendar.month_name[i]}
        for i in range(1, 13)
    ]

    years = list(range(today.year - 2, today.year + 2))

    # Payment completion percentage
    payment_completion = 0
    if processed_count > 0:
        payment_completion = (paid_count / processed_count) * 100

    # Processing completion percentage
    processing_completion = 0
    if total_staff > 0:
        processing_completion = (processed_count / total_staff) * 100

    context = {
        'page_title': f'Payroll Dashboard - {calendar.month_name[current_month]} {current_year}',
        'current_month': current_month,
        'current_year': current_year,
        'month_name': calendar.month_name[current_month],
        'months': months,
        'years': years,

        # Staff counts
        'total_staff': total_staff,
        'processed_count': processed_count,
        'unprocessed_count': unprocessed_count,
        'processing_completion': processing_completion,

        # Payment status
        'paid_count': paid_count,
        'partially_paid_count': partially_paid_count,
        'pending_count': pending_count,
        'payment_completion': payment_completion,

        # Financial totals
        'total_gross': total_gross,
        'total_net': total_net,
        'total_statutory': total_statutory,
        'total_tax': total_tax,
        'total_other_deductions': total_other_deductions,
        'total_amount_paid': total_amount_paid,
        'total_outstanding': total_outstanding,

        # Averages
        'avg_gross': avg_gross,
        'avg_net': avg_net,
        'avg_tax_rate': avg_tax_rate,

        # Lists
        'records': records,
        'top_earners': top_earners,
        'recent_payments': recent_payments,
        'unprocessed_staff': unprocessed_staff,
    }

    return render(request, 'finance/payroll/dashboard.html', context)


@login_required
@permission_required('finance.view_salaryrecord', raise_exception=True)
def bulk_payroll_view(request):
    """Bulk payroll processing page"""
    from datetime import datetime
    import calendar

    today = datetime.now()
    current_year = int(request.GET.get('year', today.year))
    current_month = int(request.GET.get('month', today.month))

    # Get all active salary structures with staff and setting
    structures = SalaryStructure.objects.filter(
        is_active=True
    ).select_related(
        'staff',
        'salary_setting'
    ).order_by(
        'staff__department',  # Sort by department first
        'staff__first_name',  # Then by first name
        'staff__last_name'  # Then by last name
    )

    # Get salary setting for this month (use the active one)
    active_setting = SalarySetting.objects.filter(is_active=True).first()

    if not active_setting:
        messages.error(request, 'No active salary setting found. Please activate a salary setting first.')
        return redirect('finance_payroll_dashboard')

    # Prepare staff data with calculations
    staff_data = []

    for structure in structures:
        # Check if already processed
        existing_record = SalaryRecord.objects.filter(
            staff=structure.staff,
            month=current_month,
            year=current_year
        ).first()

        # Calculate salary breakdown
        calculator = SalaryCalculator(structure, current_month, current_year)
        salary_calc = calculator.calculate_complete_salary()

        # Parse existing deductions and allowances if record exists
        existing_deductions = {}
        existing_allowances = {}
        bonus = Decimal('0.00')
        amount_paid = Decimal('0.00')
        is_paid = False

        if existing_record:
            # Parse JSON fields
            if existing_record.other_deductions:
                if isinstance(existing_record.other_deductions, str):
                    existing_deductions = json.loads(existing_record.other_deductions)
                else:
                    existing_deductions = existing_record.other_deductions

            if existing_record.additional_income:
                if isinstance(existing_record.additional_income, str):
                    existing_allowances = json.loads(existing_record.additional_income)
                else:
                    existing_allowances = existing_record.additional_income

            bonus = existing_record.bonus
            amount_paid = existing_record.amount_paid
            is_paid = existing_record.payment_status == SalaryRecord.PaymentStatus.PAID

        staff_data.append({
            'structure': structure,
            'staff': structure.staff,
            'monthly_salary': structure.monthly_salary,
            'statutory_deductions': salary_calc['total_statutory_deductions'],
            'paye_tax': salary_calc['monthly_tax'],
            'net_monthly': salary_calc['net_salary'],
            'bonus': bonus,
            'existing_deductions': existing_deductions,
            'existing_allowances': existing_allowances,
            'amount_paid': amount_paid,
            'is_paid': is_paid,
            'has_record': existing_record is not None,
            'record_id': existing_record.id if existing_record else None,
        })

    # Get deduction and allowance configurations from active setting
    deduction_configs = active_setting.other_deductions_config or []
    allowance_configs = active_setting.income_items or []

    # Filter only manual deductions (not linked to external systems)
    manual_deduction_configs = [
        d for d in deduction_configs
        if not d.get('linked_to')
    ]

    context = {
        'staff_data': staff_data,
        'current_year': current_year,
        'current_month': current_month,
        'month_name': calendar.month_name[current_month],
        'years': range(2020, today.year + 2),
        'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'deduction_configs': manual_deduction_configs,
        'allowance_configs': allowance_configs,
        'active_setting': active_setting,
        'page_title': f'Bulk Payroll - {calendar.month_name[current_month]} {current_year}'
    }

    return render(request, 'finance/payroll/bulk_payroll.html', context)


@login_required
@permission_required('finance.add_salaryrecord', raise_exception=True)
@require_http_methods(["POST"])
def bulk_payroll_save(request):
    """AJAX endpoint to save bulk payroll data"""
    import json
    from decimal import Decimal
    try:
        data = json.loads(request.body)
        staff_records = data.get('staff_records', [])
        current_year = int(data.get('year'))
        current_month = int(data.get('month'))

        results = []

        for record_data in staff_records:
            try:
                structure_id = record_data.get('structure_id')
                structure = SalaryStructure.objects.get(id=structure_id)

                # Get or create salary record
                salary_record, created = SalaryRecord.objects.get_or_create(
                    staff=structure.staff,
                    month=current_month,
                    year=current_year,
                    defaults={
                        'salary_structure': structure,
                        'salary_setting': structure.salary_setting,
                        'monthly_salary': structure.monthly_salary,
                        'annual_salary': structure.annual_salary,
                        'created_by': request.user,
                    }
                )

                # Parse deductions and allowances
                deductions = record_data.get('deductions', {})
                allowances = record_data.get('allowances', {})
                bonus = Decimal(str(record_data.get('bonus', '0')))
                amount_paid = Decimal(str(record_data.get('amount_paid', '0')))
                mark_as_paid = record_data.get('mark_as_paid', False)

                # Calculate complete salary
                calculator = SalaryCalculator(structure, current_month, current_year)
                salary_calc = calculator.calculate_complete_salary(
                    bonus=bonus,
                    custom_deductions=deductions,
                    additional_income=allowances
                )

                # Helper to convert Decimals in nested structures
                def convert_decimals(obj):
                    if isinstance(obj, Decimal):
                        return str(obj)
                    elif isinstance(obj, dict):
                        return {k: convert_decimals(v) for k, v in obj.items()}
                    elif isinstance(obj, list):
                        return [convert_decimals(item) for item in obj]
                    return obj

                # Update salary record
                salary_record.bonus = bonus
                salary_record.additional_income = convert_decimals(allowances)
                salary_record.other_deductions = convert_decimals(deductions)
                salary_record.basic_components_breakdown = convert_decimals(
                    salary_calc.get('basic_components_breakdown', {}))
                salary_record.allowances_breakdown = convert_decimals(salary_calc.get('allowances_breakdown', {}))
                salary_record.total_income = Decimal(str(salary_calc['total_income']))
                salary_record.gross_salary = Decimal(str(salary_calc['total_income']))
                salary_record.statutory_deductions = convert_decimals(salary_calc.get('statutory_deductions', {}))
                salary_record.total_statutory_deductions = Decimal(str(salary_calc['total_statutory_deductions']))
                salary_record.total_other_deductions = Decimal(str(salary_calc['total_other_deductions']))
                salary_record.annual_gross_income = Decimal(str(salary_calc['annual_gross_income']))
                salary_record.total_reliefs = Decimal(str(salary_calc['total_reliefs']))
                salary_record.taxable_income = Decimal(str(salary_calc['taxable_income']))
                salary_record.annual_tax = Decimal(str(salary_calc['annual_tax']))
                salary_record.monthly_tax = Decimal(str(salary_calc['monthly_tax']))
                salary_record.total_taxation = Decimal(str(salary_calc['total_taxation']))
                salary_record.effective_tax_rate = Decimal(str(salary_calc['effective_tax_rate']))
                salary_record.net_salary = Decimal(str(salary_calc['net_salary']))
                salary_record.amount_paid = amount_paid

                # Calculate actual take home (net salary - other deductions + allowances + bonus)
                total_allowances = sum(Decimal(str(v)) for v in allowances.values() if v)
                actual_take_home = (
                        salary_record.net_salary -
                        salary_record.total_other_deductions +
                        total_allowances +
                        bonus
                )

                # Update payment status based on actual take home
                if mark_as_paid and round(amount_paid) >= round(actual_take_home):
                    salary_record.payment_status = SalaryRecord.PaymentStatus.PAID
                    salary_record.paid_date = date.today()
                    salary_record.paid_by = request.user
                elif amount_paid > 0:
                    salary_record.payment_status = SalaryRecord.PaymentStatus.PARTIALLY_PAID
                else:
                    salary_record.payment_status = SalaryRecord.PaymentStatus.PENDING

                salary_record.save()

                results.append({
                    'success': True,
                    'staff_name': str(structure.staff),
                    'record_id': salary_record.id
                })

            except Exception as e:
                results.append({
                    'success': False,
                    'staff_name': record_data.get('staff_name', 'Unknown'),
                    'error': str(e)
                })

        return JsonResponse({
            'success': True,
            'results': results,
            'total': len(staff_records),
            'successful': sum(1 for r in results if r['success']),
            'failed': sum(1 for r in results if not r['success'])
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@permission_required('finance.add_salaryrecord', raise_exception=True)
@require_http_methods(["POST"])
def auto_save_payroll_row(request):
    """AJAX endpoint to auto-save individual payroll row"""
    import json
    from decimal import Decimal

    try:
        data = json.loads(request.body)
        structure_id = data.get('structure_id')
        current_year = int(data.get('year'))
        current_month = int(data.get('month'))

        structure = SalaryStructure.objects.get(id=structure_id)

        # Get or create salary record
        salary_record, created = SalaryRecord.objects.get_or_create(
            staff=structure.staff,
            month=current_month,
            year=current_year,
            defaults={
                'salary_structure': structure,
                'salary_setting': structure.salary_setting,
                'monthly_salary': structure.monthly_salary,
                'annual_salary': structure.annual_salary,
                'created_by': request.user,
            }
        )

        # Parse data
        deductions = data.get('deductions', {})
        allowances = data.get('allowances', {})
        bonus = Decimal(str(data.get('bonus', '0')))
        amount_paid = Decimal(str(data.get('amount_paid', '0')))

        # Calculate salary
        calculator = SalaryCalculator(structure, current_month, current_year)
        salary_calc = calculator.calculate_complete_salary(
            bonus=bonus,
            custom_deductions=deductions,
            additional_income=allowances
        )

        # Helper to convert Decimals
        def convert_decimals(obj):
            if isinstance(obj, Decimal):
                return str(obj)
            elif isinstance(obj, dict):
                return {k: convert_decimals(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [convert_decimals(item) for item in obj]
            return obj

        # Update record
        salary_record.bonus = bonus
        salary_record.additional_income = convert_decimals(allowances)
        salary_record.other_deductions = convert_decimals(deductions)
        salary_record.amount_paid = amount_paid
        salary_record.net_salary = Decimal(str(salary_calc['net_salary']))
        salary_record.total_other_deductions = Decimal(str(salary_calc['total_other_deductions']))

        # Calculate actual take home and update payment status
        total_allowances = sum(Decimal(str(v)) for v in allowances.values() if v)
        actual_take_home = (
                salary_record.net_salary -
                salary_record.total_other_deductions +
                total_allowances +
                bonus
        )

        # Update payment status based on actual take home
        if amount_paid >= actual_take_home:
            salary_record.payment_status = SalaryRecord.PaymentStatus.PAID
            salary_record.paid_date = date.today()
            salary_record.paid_by = request.user
        elif amount_paid > 0:
            salary_record.payment_status = SalaryRecord.PaymentStatus.PARTIALLY_PAID
        else:
            salary_record.payment_status = SalaryRecord.PaymentStatus.PENDING

        salary_record.save()

        return JsonResponse({
            'success': True,
            'message': f'Auto-saved for {structure.staff}',
            'record_id': salary_record.id
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


class BonusListView(LoginRequiredMixin, ListView):
    """List all bonuses with filtering options"""
    model = Bonus
    template_name = 'finance/bonus/bonus_list.html'
    context_object_name = 'bonuses'
    paginate_by = 20

    def get_queryset(self):
        queryset = Bonus.objects.all()

        # Get filter parameters
        month = self.request.GET.get('month')
        year = self.request.GET.get('year')
        search_query = self.request.GET.get('search', '')
        category_filter = self.request.GET.get('category', '')
        status_filter = self.request.GET.get('status', '')

        # Apply filters
        if month:
            queryset = queryset.filter(month=month)

        if year:
            queryset = queryset.filter(year=year)

        if category_filter:
            queryset = queryset.filter(category=category_filter)

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        if search_query:
            queryset = queryset.filter(
                Q(volunteer_name__icontains=search_query) |
                Q(staff__first_name__icontains=search_query) |
                Q(staff__last_name__icontains=search_query) |
                Q(staff__staff_id__icontains=search_query) |
                Q(notes__icontains=search_query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get filter parameters
        month = self.request.GET.get('month')
        year = self.request.GET.get('year')
        search_query = self.request.GET.get('search', '')
        category_filter = self.request.GET.get('category', '')
        status_filter = self.request.GET.get('status', '')

        # Initialize filter form with current values
        filter_data = {
            'month': month,
            'year': year,
            'search': search_query,
            'category': category_filter,
            'status': status_filter,
        }
        context['filter_form'] = BonusFilterForm(filter_data)

        # Set page title
        context['page_title'] = 'Bonus Management'

        # Calculate summary statistics
        queryset = self.get_queryset()
        context['total_amount'] = queryset.aggregate(total=Sum('amount'))['total'] or 0
        context['paid_amount'] = queryset.filter(status='paid').aggregate(total=Sum('amount'))['total'] or 0
        context['unpaid_amount'] = queryset.filter(status='unpaid').aggregate(total=Sum('amount'))['total'] or 0

        # Calculate category breakdown
        category_breakdown = {}
        for category in Bonus.BonusCategory.choices:
            cat_value = category[0]
            cat_label = category[1]
            total = queryset.filter(category=cat_value).aggregate(total=Sum('amount'))['total'] or 0
            category_breakdown[cat_value] = {
                'label': cat_label,
                'total': total
            }
        context['category_breakdown'] = category_breakdown

        return context


class BonusDetailView(LoginRequiredMixin, DetailView):
    """View bonus details"""
    model = Bonus
    template_name = 'finance/bonus/bonus_detail.html'
    context_object_name = 'bonus'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Bonus Details - {self.object}'
        return context


class BonusCreateView(LoginRequiredMixin, CreateView):
    """Create a new bonus"""
    model = Bonus
    form_class = BonusForm
    template_name = 'finance/bonus/bonus_form.html'
    success_url = reverse_lazy('finance_bonus_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Bonus'
        context['action'] = 'Create'
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Bonus created successfully!')
        return super().form_valid(form)


class BonusUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing bonus"""
    model = Bonus
    form_class = BonusForm
    template_name = 'finance/bonus/bonus_form.html'
    success_url = reverse_lazy('finance_bonus_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Bonus - {self.object}'
        context['action'] = 'Update'
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Bonus updated successfully!')
        return super().form_valid(form)


class BonusDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a bonus"""
    model = Bonus
    template_name = 'finance/bonus/bonus_confirm_delete.html'
    success_url = reverse_lazy('finance_bonus_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Delete Bonus - {self.object}'
        return context

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Bonus deleted successfully!')
        return super().delete(request, *args, **kwargs)


@login_required
def mark_bonus_as_paid_view(request, pk):
    """Mark a bonus as paid"""
    bonus = get_object_or_404(Bonus, pk=pk)

    if request.method == 'POST':
        bonus.status = Bonus.Status.PAID
        bonus.save()
        messages.success(request, f'Bonus for {bonus} marked as paid!')
        return redirect('finance_bonus_detail', pk=pk)

    return render(request, 'finance/bonus/mark_as_paid.html', {
        'bonus': bonus,
        'page_title': f'Mark as Paid - {bonus}'
    })


@login_required
def staff_search_view(request):
    """AJAX endpoint for searching staff"""
    query = request.GET.get('q', '')

    if len(query) < 2:
        return JsonResponse({'results': []})

    staff_list = StaffModel.objects.filter(
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(staff_id__icontains=query)
    ).values('id', 'first_name', 'last_name', 'staff_id')

    results = []
    for staff in staff_list:
        results.append({
            'id': staff['id'],
            'text': f"{staff['first_name']} {staff['last_name']} ({staff['staff_id']})"
        })

    return JsonResponse({'results': results})


@login_required
def bonus_report_view(request):
    """Generate a report of bonuses with category breakdown"""
    # Get filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')
    search_query = request.GET.get('search', '')

    # Initialize filter form with current values
    filter_data = {
        'month': month,
        'year': year,
        'search': search_query,
    }
    filter_form = BonusFilterForm(filter_data)

    # Start with all bonuses
    queryset = Bonus.objects.all()

    # Apply filters
    if month:
        queryset = queryset.filter(month=month)

    if year:
        queryset = queryset.filter(year=year)

    if search_query:
        queryset = queryset.filter(
            Q(volunteer_name__icontains=search_query) |
            Q(staff__first_name__icontains=search_query) |
            Q(staff__last_name__icontains=search_query) |
            Q(staff__staff_id__icontains=search_query)
        )

    # Calculate summary statistics
    total_amount = queryset.aggregate(total=Sum('amount'))['total'] or 0
    paid_amount = queryset.filter(status='paid').aggregate(total=Sum('amount'))['total'] or 0
    unpaid_amount = queryset.filter(status='unpaid').aggregate(total=Sum('amount'))['total'] or 0

    # Calculate category breakdown
    category_breakdown = {}
    for category in Bonus.BonusCategory.choices:
        cat_value = category[0]
        cat_label = category[1]
        total = queryset.filter(category=cat_value).aggregate(total=Sum('amount'))['total'] or 0
        category_breakdown[cat_value] = {
            'label': cat_label,
            'total': total,
            'count': queryset.filter(category=cat_value).count()
        }

    # Group bonuses by recipient for detailed breakdown
    recipient_breakdown = {}
    for bonus in queryset:
        recipient_name = bonus.staff.__str__() if bonus.type == 'staff' else bonus.volunteer_name
        recipient_type = bonus.type

        if recipient_name not in recipient_breakdown:
            recipient_breakdown[recipient_name] = {
                'type': recipient_type,
                'categories': {},
                'total': 0
            }

        category = bonus.category
        if category not in recipient_breakdown[recipient_name]['categories']:
            recipient_breakdown[recipient_name]['categories'][category] = 0

        recipient_breakdown[recipient_name]['categories'][category] += bonus.amount
        recipient_breakdown[recipient_name]['total'] += bonus.amount

    context = {
        'page_title': 'Bonus Report',
        'filter_form': filter_form,
        'bonuses': queryset,
        'total_amount': total_amount,
        'paid_amount': paid_amount,
        'unpaid_amount': unpaid_amount,
        'category_breakdown': category_breakdown,
        'recipient_breakdown': recipient_breakdown,
        'month': month,
        'year': year,
        'search_query': search_query,
    }

    return render(request, 'finance/bonus/bonus_report.html', context)


@login_required
def bonus_report_pdf_view(request):
    """Generate and download bonus report as PDF"""
    # Get filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')
    search_query = request.GET.get('search', '')

    # Start with all bonuses
    queryset = Bonus.objects.all()

    # Apply filters
    if month:
        queryset = queryset.filter(month=month)

    if year:
        queryset = queryset.filter(year=year)

    if search_query:
        queryset = queryset.filter(
            Q(volunteer_name__icontains=search_query) |
            Q(staff__first_name__icontains=search_query) |
            Q(staff__last_name__icontains=search_query) |
            Q(staff__staff_id__icontains=search_query)
        )

    # Calculate summary statistics
    total_amount = queryset.aggregate(total=Sum('amount'))['total'] or 0
    paid_amount = queryset.filter(status='paid').aggregate(total=Sum('amount'))['total'] or 0
    unpaid_amount = queryset.filter(status='unpaid').aggregate(total=Sum('amount'))['total'] or 0

    # Calculate category breakdown
    category_breakdown = []
    for category in Bonus.BonusCategory.choices:
        cat_value = category[0]
        cat_label = category[1]
        total = queryset.filter(category=cat_value).aggregate(total=Sum('amount'))['total'] or 0
        count = queryset.filter(category=cat_value).count()
        category_breakdown.append([cat_label, count, f"₦{total:,.2f}"])

    # Group bonuses by recipient for detailed breakdown
    recipient_breakdown = []
    recipient_data = {}

    for bonus in queryset:
        recipient_name = bonus.staff.__str__() if bonus.type == 'staff' else bonus.volunteer_name
        recipient_type = bonus.type

        if recipient_name not in recipient_data:
            recipient_data[recipient_name] = {
                'type': recipient_type,
                'categories': {},
                'total': 0
            }

        category = bonus.category
        if category not in recipient_data[recipient_name]['categories']:
            recipient_data[recipient_name]['categories'][category] = 0

        recipient_data[recipient_name]['categories'][category] += bonus.amount
        recipient_data[recipient_name]['total'] += bonus.amount

    # Convert to list for table
    for name, data in recipient_data.items():
        recipient_breakdown.append([name, data['type'], f"₦{data['total']:,.2f}"])

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    filename = f"bonus_report_{year}_{month}.pdf" if month and year else "bonus_report.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a237e'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1a237e'),
        spaceAfter=12,
        spaceBefore=20
    )

    # Title
    title = "BONUS REPORT"
    if month and year:
        title += f" - {calendar.month_name[int(month)]} {year}"
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.2 * inch))

    # Summary section
    elements.append(Paragraph("Summary", heading_style))

    summary_data = [
        ['Total Amount', f"₦{total_amount:,.2f}"],
        ['Paid Amount', f"₦{paid_amount:,.2f}"],
        ['Unpaid Amount', f"₦{unpaid_amount:,.2f}"],
    ]

    summary_table = Table(summary_data, colWidths=[3 * inch, 3 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f5f5f5')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3 * inch))

    # Category breakdown section
    elements.append(Paragraph("Category Breakdown", heading_style))

    category_data = [['Category', 'Count', 'Total Amount']]
    category_data.extend(category_breakdown)

    category_table = Table(category_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch])
    category_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    elements.append(category_table)
    elements.append(Spacer(1, 0.3 * inch))

    # Recipient breakdown section
    elements.append(Paragraph("Recipient Breakdown", heading_style))

    recipient_data = [['Recipient Name', 'Type', 'Total Amount']]
    recipient_data.extend(recipient_breakdown)

    recipient_table = Table(recipient_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch])
    recipient_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    elements.append(recipient_table)

    # Build PDF
    doc.build(elements)

    return response


@login_required
@permission_required('finance.view_salaryrecord', raise_exception=True)
def annual_payroll_list_view(request):
    """View all staff annual payroll for a selected year"""
    from datetime import datetime
    import calendar

    today = datetime.now()
    selected_year = int(request.GET.get('year', today.year))
    search_query = request.GET.get('search', '')

    # Get all salary records for the selected year
    records = SalaryRecord.objects.filter(
        year=selected_year
    ).select_related(
        'staff__staff_profile__user',
        'salary_structure',
        'staff__department'
    ).order_by('staff__first_name', 'staff__last_name')

    # Group records by staff
    from collections import defaultdict
    staff_annual_data = defaultdict(lambda: {
        'staff': None,
        'structure': None,
        'department': None,
        'records': [],
        'months_covered': [],
        'total_income': 0,
        'total_deductions': 0,
        'total_net': 0,
        'months_count': 0
    })

    for record in records:
        staff_id = record.staff.id
        data = staff_annual_data[staff_id]

        data['staff'] = record.staff
        data['structure'] = record.salary_structure
        data['department'] = record.staff.department.name if record.staff.department else 'N/A'
        data['records'].append(record)
        data['months_covered'].append(calendar.month_abbr[record.month])
        data['total_income'] += float(record.gross_salary)
        # Fix: Calculate total deductions by summing statutory and other deductions
        data['total_deductions'] += float(record.total_statutory_deductions + record.total_other_deductions)
        data['total_net'] += float(record.net_salary)
        data['months_count'] += 1

    # Convert to list and sort
    staff_list = list(staff_annual_data.values())

    # Apply search filter if provided
    if search_query:
        staff_list = [
            s for s in staff_list
            if search_query.lower() in s['staff'].first_name.lower()
               or search_query.lower() in s['staff'].last_name.lower()
               or search_query.lower() in str(s['staff'].staff_id).lower()
        ]

    context = {
        'staff_list': staff_list,
        'selected_year': selected_year,
        'search_query': search_query,
        'years': range(2020, today.year + 2),
        'page_title': f'Annual Payroll Reports - {selected_year}'
    }

    return render(request, 'finance/payroll/annual_payroll_list.html', context)


@login_required
@permission_required('finance.view_salaryrecord', raise_exception=True)
def annual_payroll_detail_view(request, structure_id):
    """View individual staff annual payroll for a selected year"""
    from datetime import datetime
    import calendar
    import json

    today = datetime.now()
    selected_year = int(request.GET.get('year', today.year))

    # Get salary structure
    structure = get_object_or_404(
        SalaryStructure.objects.select_related(
            'staff__staff_profile__user',
            'salary_setting',
            'staff__department'
        ),
        pk=structure_id
    )

    # Get all salary records for this staff for the selected year
    records = SalaryRecord.objects.filter(
        salary_structure=structure,
        year=selected_year
    ).order_by('month')

    if not records.exists():
        from django.contrib import messages
        messages.warning(request, f'No payroll records found for {structure.staff} in {selected_year}.')
        return redirect('finance_annual_payroll_list')

    # Helper function to parse JSON fields
    def parse_json_field(field):
        if field and isinstance(field, str):
            try:
                return json.loads(field)
            except (json.JSONDecodeError, TypeError):
                return {}
        return field or {}

    # Aggregate annual data
    annual_data = {
        'months_covered': [],
        'months_count': 0,
        'total_gross_income': 0,
        'total_basic_salary': 0,
        'total_allowances': 0,
        'total_bonus': 0,
        'total_additional_income': 0,
        'total_statutory_deductions': 0,
        'total_other_deductions': 0,
        'total_paye': 0,
        'total_other_taxes': 0,
        'total_net_salary': 0,
        'basic_components_breakdown': {},
        'allowances_breakdown': {},
        'statutory_breakdown': {},
        'other_deductions_breakdown': {},
    }

    for record in records:
        # Track months
        annual_data['months_covered'].append(calendar.month_name[record.month])
        annual_data['months_count'] += 1

        # Aggregate totals
        annual_data['total_gross_income'] += float(record.gross_salary)
        annual_data['total_bonus'] += float(record.bonus)
        annual_data['total_paye'] += float(record.monthly_tax)
        annual_data['total_other_taxes'] += float(record.other_taxes)
        annual_data['total_net_salary'] += float(record.net_salary)

        # Parse and aggregate additional income
        additional_income = parse_json_field(record.additional_income)
        for key, value in additional_income.items():
            annual_data['total_additional_income'] += float(value or 0)

        # Parse and aggregate other deductions
        other_deductions = parse_json_field(record.other_deductions)
        for key, value in other_deductions.items():
            deduction_amount = float(value or 0)
            annual_data['total_other_deductions'] += deduction_amount

            # Track breakdown
            if key not in annual_data['other_deductions_breakdown']:
                annual_data['other_deductions_breakdown'][key] = 0
            annual_data['other_deductions_breakdown'][key] += deduction_amount

        # Recalculate using calculator to get component breakdowns
        calculator = SalaryCalculator(structure, record.month, record.year)
        salary_data = calculator.calculate_complete_salary(
            bonus=record.bonus,
            custom_deductions=other_deductions,
            additional_income=additional_income
        )

        # Aggregate basic components
        basic_components = salary_data.get('basic_components_breakdown', {})
        for code, component in basic_components.items():
            if isinstance(component, dict):
                if code not in annual_data['basic_components_breakdown']:
                    annual_data['basic_components_breakdown'][code] = {
                        'name': component.get('name', code),
                        'amount': 0
                    }
                annual_data['basic_components_breakdown'][code]['amount'] += float(component.get('amount', 0))

        # Aggregate allowances
        allowances = salary_data.get('allowances_breakdown', {})
        for allowance_name, allowance_data in allowances.items():
            if isinstance(allowance_data, dict):
                amount = float(allowance_data.get('amount', 0))
                if allowance_name not in annual_data['allowances_breakdown']:
                    annual_data['allowances_breakdown'][allowance_name] = 0
                annual_data['allowances_breakdown'][allowance_name] += amount

        # Aggregate statutory deductions
        statutory_deductions = salary_data.get('statutory_deductions', {})
        for name, deduction in statutory_deductions.items():
            if isinstance(deduction, dict):
                amount = float(deduction.get('amount', 0))
                if name not in annual_data['statutory_breakdown']:
                    annual_data['statutory_breakdown'][name] = 0
                annual_data['statutory_breakdown'][name] += amount
                annual_data['total_statutory_deductions'] += amount

    # Calculate total allowances from breakdown
    annual_data['total_allowances'] = sum(annual_data['allowances_breakdown'].values())

    # Calculate total basic salary from breakdown
    annual_data['total_basic_salary'] = sum(
        comp['amount'] for comp in annual_data['basic_components_breakdown'].values()
    )

    context = {
        'structure': structure,
        'selected_year': selected_year,
        'annual_data': annual_data,
        'years': range(2020, today.year + 2),
        'page_title': f'Annual Payroll - {structure.staff} - {selected_year}'
    }

    return render(request, 'finance/payroll/annual_payroll_detail.html', context)


@login_required
@permission_required('finance.view_salaryrecord', raise_exception=True)
def download_annual_payslip_pdf(request, structure_id):
    """Generate and download annual payslip as PDF"""
    from datetime import datetime
    import calendar
    import json

    today = datetime.now()
    selected_year = int(request.GET.get('year', today.year))

    # Get salary structure
    structure = get_object_or_404(
        SalaryStructure.objects.select_related(
            'staff__staff_profile__user',
            'salary_setting',
            'staff__department'
        ),
        pk=structure_id
    )

    # Get school info
    school_info = SchoolInfoModel.objects.first()

    # Get all salary records for this staff for the selected year
    records = SalaryRecord.objects.filter(
        salary_structure=structure,
        year=selected_year
    ).order_by('month')

    if not records.exists():
        from django.contrib import messages
        messages.error(request, f'No payroll records found for {structure.staff} in {selected_year}.')
        return redirect('finance_annual_payroll_detail', structure_id=structure_id)

    # Helper function to parse JSON fields
    def parse_json_field(field):
        if field and isinstance(field, str):
            try:
                return json.loads(field)
            except (json.JSONDecodeError, TypeError):
                return {}
        return field or {}

    # Aggregate annual data (same logic as detail view)
    annual_data = {
        'months_covered': [],
        'months_count': 0,
        'total_gross_income': 0,
        'total_basic_salary': 0,
        'total_allowances': 0,
        'total_bonus': 0,
        'total_additional_income': 0,
        'total_statutory_deductions': 0,
        'total_other_deductions': 0,
        'total_paye': 0,
        'total_other_taxes': 0,
        'total_net_salary': 0,
        'basic_components_breakdown': {},
        'allowances_breakdown': {},
        'statutory_breakdown': {},
        'other_deductions_breakdown': {},
    }

    for record in records:
        # Track months
        annual_data['months_covered'].append(calendar.month_abbr[record.month])
        annual_data['months_count'] += 1

        # Aggregate totals
        annual_data['total_gross_income'] += float(record.gross_salary)
        annual_data['total_bonus'] += float(record.bonus)
        annual_data['total_paye'] += float(record.monthly_tax)
        annual_data['total_other_taxes'] += float(record.other_taxes)
        annual_data['total_net_salary'] += float(record.net_salary)

        # Parse and aggregate additional income
        additional_income = parse_json_field(record.additional_income)
        for key, value in additional_income.items():
            annual_data['total_additional_income'] += float(value or 0)

        # Parse and aggregate other deductions
        other_deductions = parse_json_field(record.other_deductions)
        for key, value in other_deductions.items():
            deduction_amount = float(value or 0)
            annual_data['total_other_deductions'] += deduction_amount

            # Track breakdown
            if key not in annual_data['other_deductions_breakdown']:
                annual_data['other_deductions_breakdown'][key] = 0
            annual_data['other_deductions_breakdown'][key] += deduction_amount

        # Recalculate using calculator to get component breakdowns
        calculator = SalaryCalculator(structure, record.month, record.year)
        salary_data = calculator.calculate_complete_salary(
            bonus=record.bonus,
            custom_deductions=other_deductions,
            additional_income=additional_income
        )

        # Aggregate basic components
        basic_components = salary_data.get('basic_components_breakdown', {})
        for code, component in basic_components.items():
            if isinstance(component, dict):
                if code not in annual_data['basic_components_breakdown']:
                    annual_data['basic_components_breakdown'][code] = {
                        'name': component.get('name', code),
                        'amount': 0
                    }
                annual_data['basic_components_breakdown'][code]['amount'] += float(component.get('amount', 0))

        # Aggregate allowances
        allowances = salary_data.get('allowances_breakdown', {})
        for allowance_name, allowance_data in allowances.items():
            if isinstance(allowance_data, dict):
                amount = float(allowance_data.get('amount', 0))
                if allowance_name not in annual_data['allowances_breakdown']:
                    annual_data['allowances_breakdown'][allowance_name] = 0
                annual_data['allowances_breakdown'][allowance_name] += amount

        # Aggregate statutory deductions
        statutory_deductions = salary_data.get('statutory_deductions', {})
        for name, deduction in statutory_deductions.items():
            if isinstance(deduction, dict):
                amount = float(deduction.get('amount', 0))
                if name not in annual_data['statutory_breakdown']:
                    annual_data['statutory_breakdown'][name] = 0
                annual_data['statutory_breakdown'][name] += amount
                annual_data['total_statutory_deductions'] += amount

    # Calculate total allowances from breakdown
    annual_data['total_allowances'] = sum(annual_data['allowances_breakdown'].values())

    # Calculate total basic salary from breakdown
    annual_data['total_basic_salary'] = sum(
        comp['amount'] for comp in annual_data['basic_components_breakdown'].values()
    )

    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response[
        'Content-Disposition'] = f'attachment; filename="annual_payslip_{structure.staff.staff_id}_{selected_year}.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        topMargin=0.4 * inch,
        bottomMargin=0.4 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch
    )
    elements = []

    # Styles
    styles = getSampleStyleSheet()

    # School Info at the top (if available)
    if school_info:
        school_style = ParagraphStyle(
            'SchoolInfo',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#2c3e50'),
            alignment=TA_CENTER,
            spaceAfter=3
        )
        elements.append(Paragraph(f"<b>{school_info.name}</b>", school_style))
        elements.append(Paragraph(f"{school_info.mobile} | {school_info.email}", school_style))
        elements.append(Paragraph(f"{school_info.address}", school_style))
        elements.append(Spacer(1, 0.15 * inch))

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1a237e'),
        spaceAfter=8,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f"ANNUAL PAYSLIP - {selected_year}", title_style))
    elements.append(Spacer(1, 0.15 * inch))

    # Staff Information
    staff_data = [
        ['Staff Information', '', 'Annual Summary', ''],
        ['Name:', str(structure.staff), 'Year:', str(selected_year)],
        ['Staff ID:', structure.staff.staff_id, 'Months Covered:', f"{annual_data['months_count']} months"],
        ['Department:', structure.staff.department.name if structure.staff.department else 'N/A',
         'Months:', ', '.join(annual_data['months_covered'])],
    ]

    staff_table = Table(staff_data, colWidths=[1.3 * inch, 2 * inch, 1.3 * inch, 2 * inch])
    staff_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#e3f2fd')),
        ('BACKGROUND', (2, 0), (3, 0), colors.HexColor('#e3f2fd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(staff_table)
    elements.append(Spacer(1, 0.15 * inch))

    # Bank Details (if available)
    if structure.bank_name:
        bank_data = [[
            f"Bank: {structure.bank_name} | Account: {structure.account_number} | Name: {structure.account_name}"
        ]]
        bank_table = Table(bank_data, colWidths=[6.6 * inch])
        bank_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f5f5f5')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(bank_table)
        elements.append(Spacer(1, 0.15 * inch))

    # Income Breakdown Section
    income_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1a237e'),
        spaceAfter=6,
        spaceBefore=10
    )
    elements.append(Paragraph("INCOME BREAKDOWN", income_title_style))

    # Basic Salary Components
    if annual_data['basic_components_breakdown']:
        basic_data = [['Component', 'Amount (N)']]
        for code, component in annual_data['basic_components_breakdown'].items():
            basic_data.append([component['name'], f"{component['amount']:,.2f}"])
        basic_data.append(['', ''])
        basic_data.append(['Total Basic Salary', f"{annual_data['total_basic_salary']:,.2f}"])

        basic_table = Table(basic_data, colWidths=[4.5 * inch, 2.1 * inch])
        basic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e3f2fd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 1), (-1, -3), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -3), 0.5, colors.grey),
            ('LINEABOVE', (0, -2), (-1, -2), 1, colors.HexColor('#1a237e')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f5e9')),
        ]))
        elements.append(basic_table)
        elements.append(Spacer(1, 0.1 * inch))

    # Allowances
    if annual_data['allowances_breakdown']:
        allowances_data = [['Allowance', 'Amount (N)']]
        for name, amount in annual_data['allowances_breakdown'].items():
            allowances_data.append([name, f"{amount:,.2f}"])
        allowances_data.append(['', ''])
        allowances_data.append(['Total Allowances', f"{annual_data['total_allowances']:,.2f}"])

        allowances_table = Table(allowances_data, colWidths=[4.5 * inch, 2.1 * inch])
        allowances_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e3f2fd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 1), (-1, -3), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -3), 0.5, colors.grey),
            ('LINEABOVE', (0, -2), (-1, -2), 1, colors.HexColor('#1a237e')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f5e9')),
        ]))
        elements.append(allowances_table)
        elements.append(Spacer(1, 0.1 * inch))

    # Bonus and Additional Income
    if annual_data['total_bonus'] > 0 or annual_data['total_additional_income'] > 0:
        other_income_data = [['Income Type', 'Amount (N)']]
        if annual_data['total_bonus'] > 0:
            other_income_data.append(['Bonus', f"{annual_data['total_bonus']:,.2f}"])
        if annual_data['total_additional_income'] > 0:
            other_income_data.append(['Additional Income', f"{annual_data['total_additional_income']:,.2f}"])
        other_income_data.append(['', ''])
        other_income_data.append(['Total Other Income',
                                f"{annual_data['total_bonus'] + annual_data['total_additional_income']:,.2f}"])

        other_income_table = Table(other_income_data, colWidths=[4.5 * inch, 2.1 * inch])
        other_income_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e3f2fd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 1), (-1, -3), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -3), 0.5, colors.grey),
            ('LINEABOVE', (0, -2), (-1, -2), 1, colors.HexColor('#1a237e')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f5e9')),
        ]))
        elements.append(other_income_table)
        elements.append(Spacer(1, 0.1 * inch))

    # Deductions Breakdown Section
    elements.append(Paragraph("DEDUCTIONS BREAKDOWN", income_title_style))

    # Statutory Deductions
    if annual_data['statutory_breakdown']:
        statutory_data = [['Deduction', 'Amount (N)']]
        for name, amount in annual_data['statutory_breakdown'].items():
            statutory_data.append([name, f"{amount:,.2f}"])
        statutory_data.append(['', ''])
        statutory_data.append(['Total Statutory Deductions', f"{annual_data['total_statutory_deductions']:,.2f}"])

        statutory_table = Table(statutory_data, colWidths=[4.5 * inch, 2.1 * inch])
        statutory_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e3f2fd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 1), (-1, -3), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -3), 0.5, colors.grey),
            ('LINEABOVE', (0, -2), (-1, -2), 1, colors.HexColor('#1a237e')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ffebee')),
        ]))
        elements.append(statutory_table)
        elements.append(Spacer(1, 0.1 * inch))

    # Other Deductions
    if annual_data['other_deductions_breakdown']:
        other_ded_data = [['Deduction', 'Amount (N)']]
        for name, amount in annual_data['other_deductions_breakdown'].items():
            other_ded_data.append([name, f"{amount:,.2f}"])
        other_ded_data.append(['', ''])
        other_ded_data.append(['Total Other Deductions', f"{annual_data['total_other_deductions']:,.2f}"])

        other_ded_table = Table(other_ded_data, colWidths=[4.5 * inch, 2.1 * inch])
        other_ded_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e3f2fd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 1), (-1, -3), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -3), 0.5, colors.grey),
            ('LINEABOVE', (0, -2), (-1, -2), 1, colors.HexColor('#1a237e')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ffebee')),
        ]))
        elements.append(other_ded_table)
        elements.append(Spacer(1, 0.1 * inch))

    # Tax Breakdown
    tax_data = [['Tax Type', 'Amount (N)']]
    tax_data.append(['PAYE Tax', f"{annual_data['total_paye']:,.2f}"])
    if annual_data['total_other_taxes'] > 0:
        tax_data.append(['Other Taxes', f"{annual_data['total_other_taxes']:,.2f}"])
    tax_data.append(['', ''])
    tax_data.append(['Total Tax', f"{annual_data['total_paye'] + annual_data['total_other_taxes']:,.2f}"])

    tax_table = Table(tax_data, colWidths=[4.5 * inch, 2.1 * inch])
    tax_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e3f2fd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 1), (-1, -3), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -3), 0.5, colors.grey),
        ('LINEABOVE', (0, -2), (-1, -2), 1, colors.HexColor('#1a237e')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ffebee')),
    ]))
    elements.append(tax_table)
    elements.append(Spacer(1, 0.15 * inch))

    # Annual Summary Table
    summary_data = [
        ['Description', 'Amount (N)'],
        ['Total Gross Income', f"{annual_data['total_gross_income']:,.2f}"],
        ['Total Statutory Deductions', f"{annual_data['total_statutory_deductions']:,.2f}"],
        ['Total Other Deductions', f"{annual_data['total_other_deductions']:,.2f}"],
        ['Total PAYE Tax', f"{annual_data['total_paye']:,.2f}"],
        ['Total Other Taxes', f"{annual_data['total_other_taxes']:,.2f}"],
        ['', ''],
        ['Total Net Salary', f"{annual_data['total_net_salary']:,.2f}"],
    ]

    summary_table = Table(summary_data, colWidths=[4.5 * inch, 2.1 * inch])

    # Apply styling
    summary_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),

        # Content rows
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Separator row
        ('LINEABOVE', (0, -2), (-1, -2), 2, colors.HexColor('#1a237e')),

        # Net Salary row (last row)
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('BOX', (0, -1), (-1, -1), 1.5, colors.HexColor('#2e7d32')),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 0.2 * inch))

    # Footer note
    note_style = ParagraphStyle(
        'Note',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(
        f"This is an automatically generated annual payslip for the year {selected_year}. "
        f"It represents the total of all {annual_data['months_count']} month(s) processed for this period.",
        note_style
    ))

    # Build PDF
    doc.build(elements)

    return response


@login_required
@permission_required('finance.view_salaryrecord', raise_exception=True)
def salary_management_report_view(request):
    """Salary management report showing totals for selected period"""
    from datetime import datetime
    import calendar
    from decimal import Decimal
    from django.db.models import Sum, Q

    today = datetime.now()

    # Get filter parameters
    search_query = request.GET.get('search', '')
    report_type = request.GET.get('report_type', 'single')  # 'single' or 'range'

    # Single month parameters
    single_month = int(request.GET.get('month', today.month))
    single_year = int(request.GET.get('year', today.year))

    # Date range parameters
    from_month = int(request.GET.get('from_month', today.month))
    from_year = int(request.GET.get('from_year', today.year))
    to_month = int(request.GET.get('to_month', today.month))
    to_year = int(request.GET.get('to_year', today.year))

    # Build query for salary records
    records_query = SalaryRecord.objects.all()

    # Apply period filter
    if report_type == 'single':
        records_query = records_query.filter(month=single_month, year=single_year)
        period_display = f"{calendar.month_name[single_month]} {single_year}"
        bonus_query = Q(month=single_month, year=single_year)
    else:
        # Date range filter
        records_query = records_query.filter(
            Q(year=from_year, month__gte=from_month) |
            Q(year__gt=from_year, year__lt=to_year) |
            Q(year=to_year, month__lte=to_month)
        ).filter(year__gte=from_year, year__lte=to_year)
        period_display = f"{calendar.month_name[from_month]} {from_year} - {calendar.month_name[to_month]} {to_year}"

        bonus_query = (
                              Q(year=from_year, month__gte=from_month) |
                              Q(year__gt=from_year, year__lt=to_year) |
                              Q(year=to_year, month__lte=to_month)
                      ) & Q(year__gte=from_year, year__lte=to_year)

    # Apply search filter
    if search_query:
        records_query = records_query.filter(
            Q(staff__first_name__icontains=search_query) |
            Q(staff__last_name__icontains=search_query) |
            Q(staff__staff_id__icontains=search_query)
        )

    records = records_query.select_related('staff', 'salary_structure', 'salary_setting')

    # Initialize aggregated data
    report_data = {
        'total_gross_salary': Decimal('0.00'),
        'basic_components': {},
        'allowances': {},
        'additional_income': {},
        'statutory_deductions': {},
        'other_deductions': {},
        'total_paye_tax': Decimal('0.00'),
        'total_reliefs': Decimal('0.00'),
        'total_take_home': Decimal('0.00'),
        'total_paid': Decimal('0.00'),
        'total_pending': Decimal('0.00'),
    }

    # Aggregate data from salary records
    for record in records:
        # Gross salary
        report_data['total_gross_salary'] += record.total_income

        # Basic components
        basic_breakdown = record.basic_components_breakdown
        if basic_breakdown:
            if isinstance(basic_breakdown, str):
                import json
                basic_breakdown = json.loads(basic_breakdown)

            for code, component in basic_breakdown.items():
                if isinstance(component, dict):
                    name = component.get('name', code)
                    amount = Decimal(str(component.get('amount', 0)))

                    if name not in report_data['basic_components']:
                        report_data['basic_components'][name] = Decimal('0.00')
                    report_data['basic_components'][name] += amount

        # Allowances
        allowances_breakdown = record.allowances_breakdown
        if allowances_breakdown:
            if isinstance(allowances_breakdown, str):
                import json
                allowances_breakdown = json.loads(allowances_breakdown)

            for allowance_name, allowance_data in allowances_breakdown.items():
                if isinstance(allowance_data, dict):
                    amount = Decimal(str(allowance_data.get('amount', 0)))

                    if allowance_name not in report_data['allowances']:
                        report_data['allowances'][allowance_name] = Decimal('0.00')
                    report_data['allowances'][allowance_name] += amount

        # Additional income
        additional = record.additional_income
        if additional:
            if isinstance(additional, str):
                import json
                additional = json.loads(additional)

            for name, amount in additional.items():
                amount_val = Decimal(str(amount)) if amount else Decimal('0.00')

                if name not in report_data['additional_income']:
                    report_data['additional_income'][name] = Decimal('0.00')
                report_data['additional_income'][name] += amount_val

        # Statutory deductions
        statutory = record.statutory_deductions
        if statutory:
            if isinstance(statutory, str):
                import json
                statutory = json.loads(statutory)

            for name, deduction in statutory.items():
                if isinstance(deduction, dict):
                    amount = Decimal(str(deduction.get('amount', 0)))
                else:
                    amount = Decimal(str(deduction)) if deduction else Decimal('0.00')

                if name not in report_data['statutory_deductions']:
                    report_data['statutory_deductions'][name] = Decimal('0.00')
                report_data['statutory_deductions'][name] += amount

        # Other deductions
        other_ded = record.other_deductions
        if other_ded:
            if isinstance(other_ded, str):
                import json
                other_ded = json.loads(other_ded)

            for name, amount in other_ded.items():
                amount_val = Decimal(str(amount)) if amount else Decimal('0.00')

                if name not in report_data['other_deductions']:
                    report_data['other_deductions'][name] = Decimal('0.00')
                report_data['other_deductions'][name] += amount_val

        # Tax and reliefs
        report_data['total_paye_tax'] += record.monthly_tax
        report_data['total_reliefs'] += record.total_reliefs

        # Take home and paid
        report_data['total_take_home'] += record.net_salary
        report_data['total_paid'] += record.amount_paid

    # Calculate pending
    report_data['total_pending'] = report_data['total_take_home'] - report_data['total_paid']

    # Get bonuses
    bonuses = Bonus.objects.filter(bonus_query)

    if search_query:
        bonuses = bonuses.filter(
            Q(staff__first_name__icontains=search_query) |
            Q(staff__last_name__icontains=search_query) |
            Q(staff__staff_id__icontains=search_query) |
            Q(volunteer_name__icontains=search_query)
        )

    # Aggregate bonus data
    staff_bonuses_total = bonuses.filter(type=Bonus.BonusType.STAFF).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    volunteer_bonuses_total = bonuses.filter(type=Bonus.BonusType.VOLUNTEER).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    total_bonuses = staff_bonuses_total + volunteer_bonuses_total

    # Bonus breakdown by category
    bonus_categories = {}
    for category in Bonus.BonusCategory.choices:
        category_total = bonuses.filter(category=category[0]).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')

        if category_total > 0:
            bonus_categories[category[1]] = category_total

    bonus_data = {
        'staff_total': staff_bonuses_total,
        'volunteer_total': volunteer_bonuses_total,
        'total': total_bonuses,
        'categories': bonus_categories
    }

    # Filter out zero values
    report_data['allowances'] = {k: v for k, v in report_data['allowances'].items() if v > 0}
    report_data['additional_income'] = {k: v for k, v in report_data['additional_income'].items() if v > 0}
    report_data['other_deductions'] = {k: v for k, v in report_data['other_deductions'].items() if v > 0}

    context = {
        'report_data': report_data,
        'bonus_data': bonus_data,
        'period_display': period_display,
        'report_type': report_type,
        'search_query': search_query,
        'single_month': single_month,
        'single_year': single_year,
        'from_month': from_month,
        'from_year': from_year,
        'to_month': to_month,
        'to_year': to_year,
        'years': range(2020, today.year + 2),
        'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'page_title': f'Salary Management Report - {period_display}'
    }

    return render(request, 'finance/payroll/salary_report.html', context)


@login_required
@permission_required('finance.view_salaryrecord', raise_exception=True)
def download_salary_report_pdf(request):
    """Generate and download salary management report as PDF"""
    from datetime import datetime
    import calendar
    from decimal import Decimal
    from django.db.models import Sum, Q
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    today = datetime.now()

    # Get filter parameters (same as view)
    search_query = request.GET.get('search', '')
    report_type = request.GET.get('report_type', 'single')

    single_month = int(request.GET.get('month', today.month))
    single_year = int(request.GET.get('year', today.year))

    from_month = int(request.GET.get('from_month', today.month))
    from_year = int(request.GET.get('from_year', today.year))
    to_month = int(request.GET.get('to_month', today.month))
    to_year = int(request.GET.get('to_year', today.year))

    # Build query (same logic as view)
    records_query = SalaryRecord.objects.all()

    if report_type == 'single':
        records_query = records_query.filter(month=single_month, year=single_year)
        period_display = f"{calendar.month_name[single_month]} {single_year}"
        bonus_query = Q(month=single_month, year=single_year)
    else:
        records_query = records_query.filter(
            Q(year=from_year, month__gte=from_month) |
            Q(year__gt=from_year, year__lt=to_year) |
            Q(year=to_year, month__lte=to_month)
        ).filter(year__gte=from_year, year__lte=to_year)
        period_display = f"{calendar.month_name[from_month]} {from_year} - {calendar.month_name[to_month]} {to_year}"

        bonus_query = (
                              Q(year=from_year, month__gte=from_month) |
                              Q(year__gt=from_year, year__lt=to_year) |
                              Q(year=to_year, month__lte=to_month)
                      ) & Q(year__gte=from_year, year__lte=to_year)

    if search_query:
        records_query = records_query.filter(
            Q(staff__first_name__icontains=search_query) |
            Q(staff__last_name__icontains=search_query) |
            Q(staff__staff_id__icontains=search_query)
        )

    records = records_query.select_related('staff', 'salary_structure', 'salary_setting')

    # Aggregate data (same logic as view)
    report_data = {
        'total_gross_salary': Decimal('0.00'),
        'basic_components': {},
        'allowances': {},
        'additional_income': {},
        'statutory_deductions': {},
        'other_deductions': {},
        'total_paye_tax': Decimal('0.00'),
        'total_reliefs': Decimal('0.00'),
        'total_take_home': Decimal('0.00'),
        'total_paid': Decimal('0.00'),
        'total_pending': Decimal('0.00'),
    }

    for record in records:
        report_data['total_gross_salary'] += record.total_income

        basic_breakdown = record.basic_components_breakdown
        if basic_breakdown:
            if isinstance(basic_breakdown, str):
                import json
                basic_breakdown = json.loads(basic_breakdown)

            for code, component in basic_breakdown.items():
                if isinstance(component, dict):
                    name = component.get('name', code)
                    amount = Decimal(str(component.get('amount', 0)))

                    if name not in report_data['basic_components']:
                        report_data['basic_components'][name] = Decimal('0.00')
                    report_data['basic_components'][name] += amount

        allowances_breakdown = record.allowances_breakdown
        if allowances_breakdown:
            if isinstance(allowances_breakdown, str):
                import json
                allowances_breakdown = json.loads(allowances_breakdown)

            for allowance_name, allowance_data in allowances_breakdown.items():
                if isinstance(allowance_data, dict):
                    amount = Decimal(str(allowance_data.get('amount', 0)))

                    if allowance_name not in report_data['allowances']:
                        report_data['allowances'][allowance_name] = Decimal('0.00')
                    report_data['allowances'][allowance_name] += amount

        additional = record.additional_income
        if additional:
            if isinstance(additional, str):
                import json
                additional = json.loads(additional)

            for name, amount in additional.items():
                amount_val = Decimal(str(amount)) if amount else Decimal('0.00')

                if name not in report_data['additional_income']:
                    report_data['additional_income'][name] = Decimal('0.00')
                report_data['additional_income'][name] += amount_val

        statutory = record.statutory_deductions
        if statutory:
            if isinstance(statutory, str):
                import json
                statutory = json.loads(statutory)

            for name, deduction in statutory.items():
                if isinstance(deduction, dict):
                    amount = Decimal(str(deduction.get('amount', 0)))
                else:
                    amount = Decimal(str(deduction)) if deduction else Decimal('0.00')

                if name not in report_data['statutory_deductions']:
                    report_data['statutory_deductions'][name] = Decimal('0.00')
                report_data['statutory_deductions'][name] += amount

        other_ded = record.other_deductions
        if other_ded:
            if isinstance(other_ded, str):
                import json
                other_ded = json.loads(other_ded)

            for name, amount in other_ded.items():
                amount_val = Decimal(str(amount)) if amount else Decimal('0.00')

                if name not in report_data['other_deductions']:
                    report_data['other_deductions'][name] = Decimal('0.00')
                report_data['other_deductions'][name] += amount_val

        report_data['total_paye_tax'] += record.monthly_tax
        report_data['total_reliefs'] += record.total_reliefs
        report_data['total_take_home'] += record.net_salary
        report_data['total_paid'] += record.amount_paid

    report_data['total_pending'] = report_data['total_take_home'] - report_data['total_paid']

    # Get bonuses
    bonuses = Bonus.objects.filter(bonus_query)

    if search_query:
        bonuses = bonuses.filter(
            Q(staff__first_name__icontains=search_query) |
            Q(staff__last_name__icontains=search_query) |
            Q(staff__staff_id__icontains=search_query) |
            Q(volunteer_name__icontains=search_query)
        )

    staff_bonuses_total = bonuses.filter(type=Bonus.BonusType.STAFF).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    volunteer_bonuses_total = bonuses.filter(type=Bonus.BonusType.VOLUNTEER).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    total_bonuses = staff_bonuses_total + volunteer_bonuses_total

    bonus_categories = {}
    for category in Bonus.BonusCategory.choices:
        category_total = bonuses.filter(category=category[0]).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')

        if category_total > 0:
            bonus_categories[category[1]] = category_total

    # Filter out zeros
    report_data['allowances'] = {k: v for k, v in report_data['allowances'].items() if v > 0}
    report_data['additional_income'] = {k: v for k, v in report_data['additional_income'].items() if v > 0}
    report_data['other_deductions'] = {k: v for k, v in report_data['other_deductions'].items() if v > 0}

    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="salary_report_{period_display.replace(" ", "_")}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch
    )
    elements = []
    styles = getSampleStyleSheet()

    # School info
    school_info = SchoolInfoModel.objects.first()
    if school_info:
        school_style = ParagraphStyle(
            'SchoolInfo',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#2c3e50'),
            alignment=TA_CENTER,
            spaceAfter=3
        )
        elements.append(Paragraph(f"<b>{school_info.name}</b>", school_style))
        elements.append(Paragraph(f"{school_info.mobile} | {school_info.email}", school_style))
        elements.append(Paragraph(f"{school_info.address}", school_style))
        elements.append(Spacer(1, 0.2 * inch))

    # Title
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a237e'),
        alignment=TA_CENTER,
        spaceAfter=5
    )
    elements.append(Paragraph("SALARY MANAGEMENT REPORT", title_style))
    elements.append(Paragraph(f"<b>Period: {period_display}</b>", ParagraphStyle(
        'Period',
        parent=styles['Normal'],
        fontSize=12,
        alignment=TA_CENTER,
        spaceAfter=10
    )))
    elements.append(Spacer(1, 0.2 * inch))

    # Build comprehensive table
    table_data = []
    table_data.append(['DESCRIPTION', 'AMOUNT (₦)'])

    # INCOME SECTION
    table_data.append(['INCOME', ''])
    table_data.append(['Total Gross Salary', f"{float(report_data['total_gross_salary']):,.2f}"])

    # Basic components
    if report_data['basic_components']:
        table_data.append(['Basic Components:', ''])
        for name, amount in report_data['basic_components'].items():
            table_data.append([f"  {name}", f"{float(amount):,.2f}"])

    # Allowances
    if report_data['allowances']:
        table_data.append(['Allowances:', ''])
        for name, amount in report_data['allowances'].items():
            table_data.append([f"  {name}", f"{float(amount):,.2f}"])

    # Additional income
    if report_data['additional_income']:
        table_data.append(['Additional Income:', ''])
        for name, amount in report_data['additional_income'].items():
            table_data.append([f"  {name}", f"{float(amount):,.2f}"])

    # BONUSES
    table_data.append(['BONUSES', ''])
    table_data.append(['Staff Bonuses Total', f"{float(staff_bonuses_total):,.2f}"])
    table_data.append(['Volunteer Bonuses Total', f"{float(volunteer_bonuses_total):,.2f}"])
    table_data.append(['Total Bonuses', f"{float(total_bonuses):,.2f}"])

    if bonus_categories:
        table_data.append(['  Breakdown by Category:', ''])
        for category, amount in bonus_categories.items():
            table_data.append([f"    {category}", f"{float(amount):,.2f}"])

    # DEDUCTIONS SECTION
    table_data.append(['DEDUCTIONS', ''])

    # Statutory
    if report_data['statutory_deductions']:
        table_data.append(['Statutory Deductions:', ''])
        for name, amount in report_data['statutory_deductions'].items():
            table_data.append([f"  {name}", f"{float(amount):,.2f}"])

    # Other deductions
    if report_data['other_deductions']:
        table_data.append(['Other Deductions:', ''])
        for name, amount in report_data['other_deductions'].items():
            table_data.append([f"  {name}", f"{float(amount):,.2f}"])

    # Tax and reliefs
    table_data.append(['Total PAYE Tax', f"{float(report_data['total_paye_tax']):,.2f}"])
    table_data.append(['Total Reliefs/Exemptions', f"{float(report_data['total_reliefs']):,.2f}"])

    # SUMMARY
    table_data.append(['SUMMARY', ''])
    table_data.append(['Total Take-Home Salary', f"{float(report_data['total_take_home']):,.2f}"])
    table_data.append(['Total Amount Paid', f"{float(report_data['total_paid']):,.2f}"])
    table_data.append(['Total Pending/Unpaid', f"{float(report_data['total_pending']):,.2f}"])

    # Create table
    report_table = Table(table_data, colWidths=[6 * inch, 2.5 * inch])

    # Styling
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]

    # Find and style section headers
    for idx, row in enumerate(table_data):
        if row[0] in ['INCOME', 'BONUSES', 'DEDUCTIONS', 'SUMMARY']:
            table_style.extend([
                ('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#e3f2fd')),
                ('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'),
                ('FONTSIZE', (0, idx), (-1, idx), 10),
            ])
        elif row[0] in ['Total Gross Salary', 'Total Bonuses', 'Total Take-Home Salary', 'Total Amount Paid',
                        'Total Pending/Unpaid']:
            table_style.extend([
                ('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#b3d9ff')),
                ('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'),
            ])

    report_table.setStyle(TableStyle(table_style))
    elements.append(report_table)

    # Build PDF
    doc.build(elements)

    return response


@login_required
@permission_required('finance.view_salaryrecord', raise_exception=True)
def bank_payment_export_view(request):
    """Bank payment export view - shows salary data in bank transfer format"""
    from datetime import datetime, date
    import calendar

    today = datetime.now()

    # Get filter parameters
    selected_month = int(request.GET.get('month', today.month))
    selected_year = int(request.GET.get('year', today.year))
    search_query = request.GET.get('search', '')
    bank_filter = request.GET.get('bank', '')
    sort_by = request.GET.get('sort', 'name')  # name, bank, department

    # Get form inputs for due date and debit account
    due_date = request.GET.get('due_date', str(date.today()))
    debit_account = request.GET.get('debit_account', '')

    # Get salary records for the selected month
    records = SalaryRecord.objects.filter(
        month=selected_month,
        year=selected_year
    ).select_related('staff', 'salary_structure', 'staff__department')

    # Apply search filter
    if search_query:
        records = records.filter(
            Q(staff__first_name__icontains=search_query) |
            Q(staff__last_name__icontains=search_query) |
            Q(staff__staff_id__icontains=search_query)
        )

    # Apply bank filter
    if bank_filter:
        records = records.filter(salary_structure__bank_name=bank_filter)

    # Build payment data
    payment_data = []
    banks_with_staff = set()

    for record in records:
        structure = record.salary_structure

        # Skip if no bank details
        if not structure.bank_name or not structure.account_number:
            continue

        banks_with_staff.add(structure.bank_name)

        # Generate staff ref: YYYYMMSTAFF_ID_PADDED
        staff_id_padded = str(record.staff.staff_id).zfill(3)
        staff_ref = f"{selected_year}{str(selected_month).zfill(2)}{staff_id_padded}"

        payment_data.append({
            'staff_ref': staff_ref,
            'beneficiary_name': str(record.staff),
            'amount': record.net_salary,
            'payment_due_date': due_date,
            'beneficiary_code': structure.beneficiary_code or '',
            'beneficiary_account_number': structure.account_number,
            'branch_sort_code': structure.branch_sort_code or '',
            'debit_account': debit_account,
            'bank_name': structure.bank_name,
            'staff_name': str(record.staff),
            'department': record.staff.department.name if record.staff.department else 'N/A',
        })

    # Sort the data
    if sort_by == 'name':
        payment_data.sort(key=lambda x: x['staff_name'])
    elif sort_by == 'bank':
        payment_data.sort(key=lambda x: (x['bank_name'], x['staff_name']))
    elif sort_by == 'department':
        payment_data.sort(key=lambda x: (x['department'], x['staff_name']))

    # Get list of banks with staff (sorted)
    banks_list = sorted(list(banks_with_staff))

    context = {
        'payment_data': payment_data,
        'banks_list': banks_list,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'month_name': calendar.month_name[selected_month],
        'search_query': search_query,
        'bank_filter': bank_filter,
        'sort_by': sort_by,
        'due_date': due_date,
        'debit_account': debit_account,
        'years': range(2020, today.year + 2),
        'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'page_title': f'Bank Payment Export - {calendar.month_name[selected_month]} {selected_year}',
        'total_amount': sum(item['amount'] for item in payment_data),
        'total_count': len(payment_data)
    }

    return render(request, 'finance/payroll/bank_payment_export.html', context)


@login_required
@permission_required('finance.view_salaryrecord', raise_exception=True)
def download_bank_payment_excel(request):
    """Download bank payment data as Excel file"""
    from datetime import datetime, date
    import calendar
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    today = datetime.now()

    # Get filter parameters (same as view)
    selected_month = int(request.GET.get('month', today.month))
    selected_year = int(request.GET.get('year', today.year))
    search_query = request.GET.get('search', '')
    bank_filter = request.GET.get('bank', '')
    sort_by = request.GET.get('sort', 'name')
    due_date = request.GET.get('due_date', str(date.today()))
    debit_account = request.GET.get('debit_account', '')

    # Get salary records
    records = SalaryRecord.objects.filter(
        month=selected_month,
        year=selected_year
    ).select_related('staff', 'salary_structure', 'staff__department')

    if search_query:
        records = records.filter(
            Q(staff__first_name__icontains=search_query) |
            Q(staff__last_name__icontains=search_query) |
            Q(staff__staff_id__icontains=search_query)
        )

    if bank_filter:
        records = records.filter(salary_structure__bank_name=bank_filter)

    # Build payment data
    payment_data = []

    for record in records:
        structure = record.salary_structure

        if not structure.bank_name or not structure.account_number:
            continue

        staff_id_padded = str(record.staff.staff_id).zfill(3)
        staff_ref = f"{selected_year}{str(selected_month).zfill(2)}{staff_id_padded}"

        payment_data.append({
            'staff_ref': staff_ref,
            'beneficiary_name': str(record.staff),
            'amount': float(record.net_salary),
            'payment_due_date': due_date,
            'beneficiary_code': structure.beneficiary_code or '',
            'beneficiary_account_number': structure.account_number,
            'branch_sort_code': structure.branch_sort_code or '',
            'debit_account': debit_account,
            'bank_name': structure.bank_name,
            'staff_name': str(record.staff),
            'department': record.staff.department.name if record.staff.department else 'N/A',
        })

    # Sort
    if sort_by == 'name':
        payment_data.sort(key=lambda x: x['staff_name'])
    elif sort_by == 'bank':
        payment_data.sort(key=lambda x: (x['bank_name'], x['staff_name']))
    elif sort_by == 'department':
        payment_data.sort(key=lambda x: (x['department'], x['staff_name']))

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Payment"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        'staff ref',
        'beneficiary name',
        'Amount',
        'payment due date',
        'beneficiary code',
        'beneficiary account number',
        'branch sort code',
        'debit account'
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Write data
    for row_num, item in enumerate(payment_data, 2):
        ws.cell(row=row_num, column=1).value = item['staff_ref']
        ws.cell(row=row_num, column=2).value = item['beneficiary_name']
        ws.cell(row=row_num, column=3).value = item['amount']
        ws.cell(row=row_num, column=4).value = item['payment_due_date']
        ws.cell(row=row_num, column=5).value = item['beneficiary_code']
        ws.cell(row=row_num, column=6).value = item['beneficiary_account_number']
        ws.cell(row=row_num, column=7).value = item['branch_sort_code']
        ws.cell(row=row_num, column=8).value = item['debit_account']

        # Apply borders
        for col in range(1, 9):
            ws.cell(row=row_num, column=col).border = border

    # Adjust column widths
    column_widths = [15, 30, 15, 18, 18, 25, 18, 18]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"bank_payment_{calendar.month_name[selected_month]}_{selected_year}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def staff_monthly_payroll_view(request):
    """
    View for staff to see their own monthly payroll
    """
    today = datetime.now()
    selected_year = int(request.GET.get('year', today.year))
    selected_month = int(request.GET.get('month', today.month))

    # Check if user has a staff profile
    try:
        staff_member = request.user.staff_profile.staff
    except (StaffProfileModel.DoesNotExist, AttributeError):
        messages.warning(request, "Your user account is not linked to a staff profile.")
        return render(request, 'finance/payroll/staff_monthly.html', {
            'page_title': 'My Monthly Payroll',
            'no_profile': True,
            'years': range(2020, today.year + 2),
            'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
            'selected_year': selected_year,
            'selected_month': selected_month,
        })

    # Check if staff has an active salary structure
    try:
        structure = SalaryStructure.objects.get(staff=staff_member, is_active=True)
    except SalaryStructure.DoesNotExist:
        messages.warning(request, "You do not have an active salary structure configured yet.")
        return render(request, 'finance/payroll/staff_monthly.html', {
            'page_title': 'My Monthly Payroll',
            'no_structure': True,
            'years': range(2020, today.year + 2),
            'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
            'selected_year': selected_year,
            'selected_month': selected_month,
        })

    # Try to get the salary record for the selected month
    try:
        record = SalaryRecord.objects.select_related(
            'staff__staff_profile__user',
            'salary_structure',
            'salary_setting'
        ).get(
            staff=staff_member,
            year=selected_year,
            month=selected_month
        )

        # Build context similar to salary_record_detail_view
        context = {
            'record': record,
            'structure': structure,
            'page_title': f'My Payslip - {calendar.month_name[selected_month]} {selected_year}',
            'years': range(2020, today.year + 2),
            'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
            'selected_year': selected_year,
            'selected_month': selected_month,
            'has_record': True,
        }

    except SalaryRecord.DoesNotExist:
        messages.info(
            request,
            f"No payroll record found for {calendar.month_name[selected_month]} {selected_year}. "
            "Please check with HR or select a different period."
        )
        context = {
            'structure': structure,
            'page_title': 'My Monthly Payroll',
            'years': range(2020, today.year + 2),
            'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
            'selected_year': selected_year,
            'selected_month': selected_month,
            'has_record': False,
        }

    return render(request, 'finance/payroll/staff_monthly.html', context)


@login_required
def staff_annual_payroll_view(request):
    """
    View for staff to see their own annual payroll
    """
    today = datetime.now()
    selected_year = int(request.GET.get('year', today.year))

    # Check if user has a staff profile
    try:
        staff_member = request.user.staff_profile.staff
    except (StaffProfileModel.DoesNotExist, AttributeError):
        messages.warning(request, "Your user account is not linked to a staff profile.")
        return render(request, 'finance/payroll/staff_annual.html', {
            'page_title': 'My Annual Payroll',
            'no_profile': True,
            'years': range(2020, today.year + 2),
            'selected_year': selected_year,
        })

    # Check if staff has an active salary structure
    try:
        structure = SalaryStructure.objects.select_related(
            'staff__staff_profile__user',
            'salary_setting',
            'staff__department'
        ).get(staff=staff_member, is_active=True)
    except SalaryStructure.DoesNotExist:
        messages.warning(request, "You do not have an active salary structure configured yet.")
        return render(request, 'finance/payroll/staff_annual.html', {
            'page_title': 'My Annual Payroll',
            'no_structure': True,
            'years': range(2020, today.year + 2),
            'selected_year': selected_year,
        })

    # Get all salary records for this staff for the selected year
    records = SalaryRecord.objects.filter(
        staff=staff_member,
        year=selected_year
    ).order_by('month')

    if not records.exists():
        messages.info(
            request,
            f"No payroll records found for {selected_year}. "
            "Please check with HR or select a different year."
        )
        return render(request, 'finance/payroll/staff_annual.html', {
            'structure': structure,
            'page_title': 'My Annual Payroll',
            'years': range(2020, today.year + 2),
            'selected_year': selected_year,
            'has_records': False,
        })

    # Calculate annual data (reuse logic from annual_payroll_detail_view)
    import json
    from decimal import Decimal

    def parse_json_field(field):
        if field and isinstance(field, str):
            try:
                return json.loads(field)
            except (json.JSONDecodeError, TypeError):
                return {}
        return field or {}

    annual_data = {
        'months_covered': [],
        'months_count': 0,
        'total_gross_income': 0,
        'total_basic_salary': 0,
        'total_allowances': 0,
        'total_bonus': 0,
        'total_additional_income': 0,
        'total_statutory_deductions': 0,
        'total_other_deductions': 0,
        'total_paye': 0,
        'total_other_taxes': 0,
        'total_net_salary': 0,
        'basic_components_breakdown': {},
        'allowances_breakdown': {},
        'statutory_breakdown': {},
        'other_deductions_breakdown': {},
    }

    for record in records:
        # Track months
        annual_data['months_covered'].append(calendar.month_name[record.month])
        annual_data['months_count'] += 1

        # Aggregate totals
        annual_data['total_gross_income'] += float(record.gross_salary)
        annual_data['total_bonus'] += float(record.bonus)
        annual_data['total_paye'] += float(record.monthly_tax)
        annual_data['total_other_taxes'] += float(record.other_taxes)
        annual_data['total_net_salary'] += float(record.net_salary)

        # Parse and aggregate additional income
        additional_income = parse_json_field(record.additional_income)
        for key, value in additional_income.items():
            annual_data['total_additional_income'] += float(value or 0)

        # Parse and aggregate other deductions
        other_deductions = parse_json_field(record.other_deductions)
        for key, value in other_deductions.items():
            deduction_amount = float(value or 0)
            annual_data['total_other_deductions'] += deduction_amount

            # Track breakdown
            if key not in annual_data['other_deductions_breakdown']:
                annual_data['other_deductions_breakdown'][key] = 0
            annual_data['other_deductions_breakdown'][key] += deduction_amount

        # Recalculate using calculator to get component breakdowns
        calculator = SalaryCalculator(structure, record.month, record.year)
        salary_data = calculator.calculate_complete_salary(
            bonus=record.bonus,
            custom_deductions=other_deductions,
            additional_income=additional_income
        )

        # Aggregate basic components
        basic_components = salary_data.get('basic_components_breakdown', {})
        for code, component in basic_components.items():
            if isinstance(component, dict):
                if code not in annual_data['basic_components_breakdown']:
                    annual_data['basic_components_breakdown'][code] = {
                        'name': component.get('name', code),
                        'amount': 0
                    }
                annual_data['basic_components_breakdown'][code]['amount'] += float(component.get('amount', 0))

        # Aggregate allowances
        allowances = salary_data.get('allowances_breakdown', {})
        for allowance_name, allowance_data in allowances.items():
            if isinstance(allowance_data, dict):
                amount = float(allowance_data.get('amount', 0))
                if allowance_name not in annual_data['allowances_breakdown']:
                    annual_data['allowances_breakdown'][allowance_name] = 0
                annual_data['allowances_breakdown'][allowance_name] += amount

        # Aggregate statutory deductions
        statutory_deductions = salary_data.get('statutory_deductions', {})
        for name, deduction in statutory_deductions.items():
            if isinstance(deduction, dict):
                amount = float(deduction.get('amount', 0))
                if name not in annual_data['statutory_breakdown']:
                    annual_data['statutory_breakdown'][name] = 0
                annual_data['statutory_breakdown'][name] += amount
                annual_data['total_statutory_deductions'] += amount

    # Calculate total allowances from breakdown
    annual_data['total_allowances'] = sum(annual_data['allowances_breakdown'].values())

    # Calculate total basic salary from breakdown
    annual_data['total_basic_salary'] = sum(
        comp['amount'] for comp in annual_data['basic_components_breakdown'].values()
    )

    context = {
        'structure': structure,
        'selected_year': selected_year,
        'annual_data': annual_data,
        'years': range(2020, today.year + 2),
        'page_title': f'My Annual Payroll - {selected_year}',
        'has_records': True,
    }

    return render(request, 'finance/payroll/staff_annual.html', context)


@login_required
def staff_bonus_list_view(request):
    """
    View for staff to see their own bonuses
    """
    # Check if user has a staff profile
    try:
        staff_member = request.user.staff_profile.staff
    except (StaffProfileModel.DoesNotExist, AttributeError):
        messages.warning(request, "Your user account is not linked to a staff profile.")
        return render(request, 'finance/bonus/staff_bonus_list.html', {
            'page_title': 'My Bonuses',
            'no_profile': True,
        })

    # Get all bonuses for this staff member
    bonuses = Bonus.objects.filter(
        type='staff',
        staff=staff_member
    ).order_by('-year', '-month')

    # Calculate summary
    from django.db.models import Sum
    total_amount = bonuses.aggregate(total=Sum('amount'))['total'] or 0
    paid_amount = bonuses.filter(status='paid').aggregate(total=Sum('amount'))['total'] or 0
    unpaid_amount = bonuses.filter(status='unpaid').aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'bonuses': bonuses,
        'page_title': 'My Bonuses',
        'total_amount': total_amount,
        'paid_amount': paid_amount,
        'unpaid_amount': unpaid_amount,
        'has_profile': True,
    }

    return render(request, 'finance/bonus/staff_bonus_list.html', context)


@login_required
def staff_bonus_detail_view(request, pk):
    """
    View for staff to see details of their own bonus
    """
    # Check if user has a staff profile
    try:
        staff_member = request.user.staff_profile.staff
    except (StaffProfileModel.DoesNotExist, AttributeError):
        messages.warning(request, "Your user account is not linked to a staff profile.")
        return redirect('staff_bonus_list')

    # Get the bonus and ensure it belongs to this staff member
    bonus = get_object_or_404(
        Bonus,
        pk=pk,
        type='staff',
        staff=staff_member
    )

    context = {
        'bonus': bonus,
        'page_title': f'Bonus Details - {bonus.get_category_display()}',
    }

    return render(request, 'finance/bonus/staff_bonus_detail.html', context)